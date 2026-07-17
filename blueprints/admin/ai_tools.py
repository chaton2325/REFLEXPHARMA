"""Outils (function-calling) que l'assistante IA de ReflexPharma peut appeler pour
repondre avec de vraies donnees de la pharmacie (ventes, chiffre d'affaires, stock,
clients, employes) plutot que des reponses generiques.

Chaque fonction `tool_*` retourne un dict JSON-serialisable. `call_ai_tool()` est le
point d'entree unique utilise par la vue Flask : il capture toute exception pour ne
jamais faire planter la conversation, et renvoie {'error': ...} a la place.

Permissions : la liste d'outils envoyee au modele n'est PAS filtree en amont (tous
les utilisateurs voient la meme liste d'outils, ce qui evite au modele d'halluciner un
nom d'outil different quand un tool attendu semble manquant, et economise des
allers-retours). A la place, CHAQUE fonction `tool_*` verifie elle-meme, en tout
debut d'execution, si l'utilisateur appelant a la permission du module concerne
(via `_check_access`). Superadmin et admin ont acces a tout, sans exception.
"""

import os
import re
import unicodedata
import uuid
from datetime import datetime, date, timedelta
from xml.sax.saxutils import escape as xml_escape

from flask import url_for
from sqlalchemy import func, or_

from extensions import db
from models.vente import Vente, VenteLigne
from models.stock import Stock
from models.stock_exit_log import StockExitLog
from models.produit import Produit
from models.fournisseur import Fournisseur
from models.groupe_fournisseur import GroupeFournisseur
from models.client import Client
from models.groupe_client import GroupeClient
from models.inventaire import Inventaire, InventaireLigne
from models.commande import Commande, CommandeLigne
from models.declaration_impot import DeclarationImpot
from models.setting import Setting
from .bon_commande_pdf import build_bon_commande_pdf
from .finance_reports import (
    compute_solde_actuel, compute_totaux_operations, query_operations_financieres,
    build_operations_financieres_pdf, build_operations_financieres_excel
)
from models.user import User
from models.poste import Poste
from utils.permissions import FEATURES


# ---------------------------------------------------------------------------
# Resolution des periodes (evite de faire faire des calculs de dates au LLM)
# ---------------------------------------------------------------------------

PERIODES_VALIDES = [
    'aujourd_hui', 'hier', 'cette_semaine', 'semaine_derniere',
    'ce_mois', 'mois_dernier', 'cette_annee', 'personnalise'
]

_PERIODE_DESC = (
    "Periode a analyser. Valeurs possibles : aujourd_hui, hier, cette_semaine, "
    "semaine_derniere, ce_mois, mois_dernier, cette_annee, ou 'personnalise' "
    "(fournir alors aussi date_debut et date_fin au format AAAA-MM-JJ)."
)

_MODULE_DESC = (
    "Cle du module/fonctionnalite dont on veut verifier les acces. Valeurs possibles : "
    + ', '.join(f"{cle} ({nom})" for cle, nom in FEATURES.items())
)


def _resolve_periode(periode, date_debut=None, date_fin=None):
    today = date.today()
    periode = (periode or 'aujourd_hui').strip()

    if periode == 'personnalise':
        if not date_debut or not date_fin:
            raise ValueError("date_debut et date_fin (format AAAA-MM-JJ) sont requis pour periode='personnalise'.")
        try:
            start = datetime.strptime(date_debut, '%Y-%m-%d').date()
            end = datetime.strptime(date_fin, '%Y-%m-%d').date()
        except ValueError:
            raise ValueError("Format de date invalide, utilisez AAAA-MM-JJ.")
        label = f"du {start.strftime('%d/%m/%Y')} au {end.strftime('%d/%m/%Y')}"
    elif periode == 'aujourd_hui':
        start = end = today
        label = "aujourd'hui"
    elif periode == 'hier':
        start = end = today - timedelta(days=1)
        label = "hier"
    elif periode == 'cette_semaine':
        start = today - timedelta(days=today.weekday())
        end = today
        label = "cette semaine"
    elif periode == 'semaine_derniere':
        start = today - timedelta(days=today.weekday() + 7)
        end = start + timedelta(days=6)
        label = "la semaine dernière"
    elif periode == 'ce_mois':
        start = today.replace(day=1)
        end = today
        label = "ce mois-ci"
    elif periode == 'mois_dernier':
        first_this_month = today.replace(day=1)
        end = first_this_month - timedelta(days=1)
        start = end.replace(day=1)
        label = "le mois dernier"
    elif periode == 'cette_annee':
        start = today.replace(month=1, day=1)
        end = today
        label = "cette année"
    else:
        raise ValueError(f"Periode inconnue: {periode}. Valeurs valides: {', '.join(PERIODES_VALIDES)}")

    start_dt = datetime.combine(start, datetime.min.time())
    end_dt = datetime.combine(end, datetime.max.time())
    return start_dt, end_dt, label


def _round2(value):
    return round(float(value or 0), 2)


_CONDITIONNEMENT_LABELS = {
    1: 'Cond 1 — unité seule',
    2: 'Cond 2 — unité + sous-unité',
    3: 'Cond 3 — unité + sous-unité + sous-sous-unité',
}


def _conditionnement_label(n):
    """'Conditionnement' = structure d'emballage du produit (combien de niveaux :
    unite / sous-unite / sous-sous-unite), PAS la quantite en stock."""
    return _CONDITIONNEMENT_LABELS.get(n, f'Cond {n}' if n is not None else 'Non précisé')


# ---------------------------------------------------------------------------
# Permissions : verifiees directement par chaque outil (pas de filtrage centralise
# en amont). Superadmin et admin ont acces a tout, sans exception ; les autres
# roles suivent has_permission() (override individuel, puis poste, sinon refuse).
# ---------------------------------------------------------------------------

def _user_has_ai_access(user, feature):
    if user is None:
        return False
    if user.role in ('superadmin', 'admin'):
        return True
    return user.has_permission(feature)


def _check_access(user, feature):
    """A appeler en tout debut de chaque fonction d'outil. Renvoie un dict d'erreur
    si l'acces est refuse, sinon None."""
    if _user_has_ai_access(user, feature):
        return None
    return {
        'error': (
            f"Accès refusé : cette information relève du module "
            f"'{FEATURES.get(feature, feature)}', auquel cet utilisateur n'a pas accès. "
            "Explique-le poliment sans donner de detail, n'essaie pas un autre outil."
        )
    }


# ---------------------------------------------------------------------------
# Ventes / chiffre d'affaires
# ---------------------------------------------------------------------------

def tool_chiffre_affaires_periode(user, periode='aujourd_hui', date_debut=None, date_fin=None):
    denied = _check_access(user, 'stats_ventes')
    if denied:
        return denied
    start_dt, end_dt, label = _resolve_periode(periode, date_debut, date_fin)
    count, ttc, ht = db.session.query(
        func.count(Vente.id),
        func.coalesce(func.sum(Vente.total_ttc), 0.0),
        func.coalesce(func.sum(Vente.total_ht), 0.0),
    ).filter(
        Vente.created_at >= start_dt,
        Vente.created_at <= end_dt,
        Vente.statut != 'annulee'
    ).one()
    count = int(count or 0)
    ttc = float(ttc or 0)
    return {
        'periode': label,
        'nombre_ventes': count,
        'chiffre_affaires_ttc': _round2(ttc),
        'chiffre_affaires_ht': _round2(ht),
        'panier_moyen_ttc': _round2(ttc / count) if count else 0.0,
    }


def tool_comparer_ca_mois_precedent(user):
    denied = _check_access(user, 'stats_ventes')
    if denied:
        return denied
    curr = tool_chiffre_affaires_periode(user, 'ce_mois')
    prev = tool_chiffre_affaires_periode(user, 'mois_dernier')
    variation_montant = curr['chiffre_affaires_ttc'] - prev['chiffre_affaires_ttc']
    variation_pourcentage = (
        _round2(variation_montant / prev['chiffre_affaires_ttc'] * 100)
        if prev['chiffre_affaires_ttc'] else None
    )
    return {
        'ce_mois': curr,
        'mois_dernier': prev,
        'variation_montant_ttc': _round2(variation_montant),
        'variation_pourcentage': variation_pourcentage,
        'tendance': 'hausse' if variation_montant > 0 else ('baisse' if variation_montant < 0 else 'stable'),
    }


def tool_prevision_chiffre_affaires(user, horizon_jours=7):
    denied = _check_access(user, 'stats_ventes')
    if denied:
        return denied
    horizon_jours = max(1, min(int(horizon_jours or 7), 60))
    jours_historique = 30
    today = date.today()
    start = today - timedelta(days=jours_historique - 1)
    start_dt = datetime.combine(start, datetime.min.time())
    end_dt = datetime.combine(today, datetime.max.time())

    rows = db.session.query(
        func.date(Vente.created_at),
        func.coalesce(func.sum(Vente.total_ttc), 0.0)
    ).filter(
        Vente.created_at >= start_dt,
        Vente.created_at <= end_dt,
        Vente.statut != 'annulee'
    ).group_by(func.date(Vente.created_at)).all()

    daily_totals = {}
    for day_val, total in rows:
        day_key = datetime.strptime(day_val, '%Y-%m-%d').date() if isinstance(day_val, str) else day_val
        daily_totals[day_key] = float(total or 0)

    series = []
    curr = start
    while curr <= today:
        series.append(daily_totals.get(curr, 0.0))
        curr += timedelta(days=1)

    n = len(series)
    if n < 4:
        return {'erreur': "Pas assez d'historique de ventes (minimum 4 jours) pour calculer une prevision fiable."}

    xs = list(range(n))
    mean_x = sum(xs) / n
    mean_y = sum(series) / n
    num = sum((x - mean_x) * (y - mean_y) for x, y in zip(xs, series))
    den = sum((x - mean_x) ** 2 for x in xs)
    slope = num / den if den else 0.0
    intercept = mean_y - slope * mean_x

    target_date = today + timedelta(days=horizon_jours)
    predicted_last_day = max(slope * (n - 1 + horizon_jours) + intercept, 0.0)
    predicted_cumulative = sum(
        max(slope * (n - 1 + i) + intercept, 0.0) for i in range(1, horizon_jours + 1)
    )

    return {
        'methode': "regression lineaire (moindres carres) sur le chiffre d'affaires quotidien des 30 derniers jours",
        'jours_historique_utilises': n,
        'tendance': 'croissante' if slope > 0 else ('decroissante' if slope < 0 else 'stable'),
        'variation_quotidienne_moyenne_ttc': _round2(slope),
        'date_cible': target_date.strftime('%Y-%m-%d'),
        'ca_ttc_estime_ce_jour_la': _round2(predicted_last_day),
        'ca_ttc_estime_cumule_sur_la_periode': _round2(predicted_cumulative),
    }


# ---------------------------------------------------------------------------
# Employes (statistiques de vente)
# ---------------------------------------------------------------------------

def tool_employe_du_mois(user, periode='ce_mois'):
    denied = _check_access(user, 'stats_ventes')
    if denied:
        return denied
    if periode not in ('ce_mois', 'mois_dernier'):
        periode = 'ce_mois'
    start_dt, end_dt, label = _resolve_periode(periode)

    rows = db.session.query(
        Vente.auteur_id, Vente.auteur_nom, Vente.auteur_prenom,
        func.count(Vente.id), func.coalesce(func.sum(Vente.total_ttc), 0.0)
    ).filter(
        Vente.created_at >= start_dt,
        Vente.created_at <= end_dt,
        Vente.statut != 'annulee'
    ).group_by(Vente.auteur_id, Vente.auteur_nom, Vente.auteur_prenom).order_by(
        func.sum(Vente.total_ttc).desc()
    ).limit(5).all()

    classement = [
        {
            'employe': (f'{prenom or ""} {nom or ""}'.strip() or 'Employé inconnu'),
            'nombre_ventes': int(count or 0),
            'chiffre_affaires_ttc': _round2(ttc),
        }
        for _, nom, prenom, count, ttc in rows
    ]

    return {
        'periode': label,
        'classement': classement,
        'employe_du_mois': classement[0]['employe'] if classement else None,
    }


# ---------------------------------------------------------------------------
# Produits / stock
# ---------------------------------------------------------------------------

def tool_nombre_produits(user):
    denied = _check_access(user, 'gestion_produits')
    if denied:
        return denied
    total = db.session.query(func.count(Produit.id)).scalar() or 0
    return {'nombre_produits_catalogue': int(total)}


def tool_liste_produits(user, recherche=None, limite=50):
    """Liste le CATALOGUE de produits (fiche produit : nom, code, fournisseur, rayon,
    famille). Ne contient aucune quantite en stock — pour ca, voir stock_produit /
    produits_stock_faible, qui portent sur une notion differente (le stock physique)."""
    denied = _check_access(user, 'gestion_produits')
    if denied:
        return denied
    limite = max(1, min(int(limite or 50), 200))
    query = Produit.query
    recherche = (recherche or '').strip()
    if recherche:
        query = query.filter(
            or_(
                Produit.nom.ilike(f'%{recherche}%'),
                Produit.code_produit.ilike(f'%{recherche}%')
            )
        )

    total = query.count()
    produits = query.order_by(Produit.nom.asc()).limit(limite).all()

    return {
        'nombre_total_correspondant': total,
        'produits': [
            {
                'produit': p.nom,
                'code_produit': p.code_produit,
                'fournisseur': p.fournisseur.nom if p.fournisseur else None,
                'rayon': p.rayon.nom if p.rayon else None,
                'famille': p.famille.nom if p.famille else None,
                'conditionnement': p.conditionnement,
                'conditionnement_label': _conditionnement_label(p.conditionnement),
            }
            for p in produits
        ],
    }


def tool_liste_fournisseurs(user, recherche=None, limite=50):
    """Liste les fournisseurs avec leur coefficient et taux de TVA effectifs (propres au
    fournisseur, ou herites de son groupe fournisseur si non personnalises)."""
    denied = _check_access(user, 'gestion_fournisseurs')
    if denied:
        return denied
    limite = max(1, min(int(limite or 50), 200))
    query = Fournisseur.query
    recherche = (recherche or '').strip()
    if recherche:
        query = query.filter(
            or_(
                Fournisseur.nom.ilike(f'%{recherche}%'),
                Fournisseur.prefixe.ilike(f'%{recherche}%')
            )
        )

    total = query.count()
    fournisseurs = query.order_by(Fournisseur.nom.asc()).limit(limite).all()

    return {
        'nombre_total_correspondant': total,
        'fournisseurs': [
            {
                'fournisseur': f.nom,
                'prefixe': f.prefixe,
                'groupe_fournisseur': f.groupe.nom if f.groupe else None,
                'coefficient': _round2(f.effectif_coefficient),
                'coefficient_personnalise': f.coefficient is not None,
                'tva_pourcentage': _round2(f.effectif_tva),
                'tva_personnalisee': f.tva is not None,
                'contact': f.contact,
            }
            for f in fournisseurs
        ],
    }


def tool_liste_groupes_fournisseurs(user, recherche=None):
    """Liste les groupes fournisseurs avec leur coefficient et TVA par defaut (herites par
    les fournisseurs du groupe qui n'ont pas de valeur personnalisee)."""
    denied = _check_access(user, 'gestion_groupes_fournisseurs')
    if denied:
        return denied
    query = GroupeFournisseur.query
    recherche = (recherche or '').strip()
    if recherche:
        query = query.filter(GroupeFournisseur.nom.ilike(f'%{recherche}%'))

    groupes = query.order_by(GroupeFournisseur.nom.asc()).all()
    return {
        'nombre_groupes': len(groupes),
        'groupes': [
            {
                'groupe_fournisseur': g.nom,
                'coefficient_par_defaut': _round2(g.coefficient_defaut),
                'tva_par_defaut_pourcentage': _round2(g.tva_defaut),
                'nombre_fournisseurs': len(g.fournisseurs),
            }
            for g in groupes
        ],
    }


def tool_stock_produit(user, nom_produit):
    denied = _check_access(user, 'gestion_stock')
    if denied:
        return denied
    nom_produit = (nom_produit or '').strip()
    if not nom_produit:
        raise ValueError("Le parametre nom_produit est requis.")

    produits = Produit.query.filter(
        or_(
            Produit.nom.ilike(f'%{nom_produit}%'),
            Produit.code_produit.ilike(f'%{nom_produit}%')
        )
    ).limit(5).all()

    if not produits:
        return {'trouve': False, 'message': f"Aucun produit ne correspond a '{nom_produit}'."}

    resultats = []
    for p in produits:
        u, su, ssu, nb_lots, prochaine_peremption = db.session.query(
            func.coalesce(func.sum(Stock.quantite_unites), 0),
            func.coalesce(func.sum(Stock.quantite_sous_unites), 0),
            func.coalesce(func.sum(Stock.quantite_sous_sous_unites), 0),
            func.count(Stock.id),
            func.min(Stock.date_peremption),
        ).filter(Stock.produit_id == p.id).one()
        resultats.append({
            'produit': p.nom,
            'code_produit': p.code_produit,
            'conditionnement': p.conditionnement,
            'conditionnement_label': _conditionnement_label(p.conditionnement),
            'quantite_unites': int(u or 0),
            'quantite_sous_unites': int(su or 0),
            'quantite_sous_sous_unites': int(ssu or 0),
            'nombre_lots': int(nb_lots or 0),
            'prochaine_peremption': prochaine_peremption.strftime('%Y-%m-%d') if prochaine_peremption else None,
        })

    return {'trouve': True, 'resultats': resultats}


def tool_produits_stock_faible(user, seuil=10, limite=15):
    denied = _check_access(user, 'gestion_stock')
    if denied:
        return denied
    seuil = max(0, int(seuil or 10))
    limite = max(1, min(int(limite or 15), 50))

    total_expr = (
        func.coalesce(func.sum(Stock.quantite_unites), 0)
        + func.coalesce(func.sum(Stock.quantite_sous_unites), 0)
        + func.coalesce(func.sum(Stock.quantite_sous_sous_unites), 0)
    )

    rows = db.session.query(
        Produit.nom, Produit.code_produit, total_expr.label('total')
    ).outerjoin(Stock, Stock.produit_id == Produit.id).group_by(
        Produit.id, Produit.nom, Produit.code_produit
    ).having(total_expr <= seuil).order_by(total_expr.asc()).limit(limite).all()

    return {
        'seuil': seuil,
        'nombre_produits_concernes': len(rows),
        'produits': [
            {'produit': nom, 'code_produit': code, 'quantite_totale': int(total or 0)}
            for nom, code, total in rows
        ],
    }


def tool_produits_peremption_proche(user, nb_jours=30, limite=15):
    denied = _check_access(user, 'gestion_stock')
    if denied:
        return denied
    nb_jours = max(1, min(int(nb_jours or 30), 365))
    limite = max(1, min(int(limite or 15), 50))
    today = date.today()
    limite_date = today + timedelta(days=nb_jours)

    rows = db.session.query(Stock, Produit.nom, Produit.code_produit).join(
        Produit, Stock.produit_id == Produit.id
    ).filter(
        Stock.date_peremption >= today,
        Stock.date_peremption <= limite_date
    ).order_by(Stock.date_peremption.asc()).limit(limite).all()

    return {
        'horizon_jours': nb_jours,
        'nombre_lots_concernes': len(rows),
        'lots': [
            {
                'produit': nom,
                'code_produit': code,
                'date_peremption': stock.date_peremption.strftime('%Y-%m-%d'),
                'quantite_totale': stock.quantite_totale,
                'code_suivi': stock.code_suivi,
            }
            for stock, nom, code in rows
        ],
    }


def tool_top_produits_vendus(user, periode='ce_mois', critere='chiffre_affaires', limite=5, date_debut=None, date_fin=None):
    denied = _check_access(user, 'stats_ventes')
    if denied:
        return denied
    start_dt, end_dt, label = _resolve_periode(periode, date_debut, date_fin)
    limite = max(1, min(int(limite or 5), 20))

    query = db.session.query(
        VenteLigne.produit_nom,
        func.coalesce(func.sum(VenteLigne.quantite), 0.0),
        func.coalesce(func.sum(VenteLigne.total_ttc), 0.0)
    ).filter(
        VenteLigne.created_at >= start_dt,
        VenteLigne.created_at <= end_dt
    ).group_by(VenteLigne.produit_nom)

    if critere == 'quantite':
        query = query.order_by(func.sum(VenteLigne.quantite).desc())
    else:
        critere = 'chiffre_affaires'
        query = query.order_by(func.sum(VenteLigne.total_ttc).desc())

    rows = query.limit(limite).all()

    return {
        'periode': label,
        'critere': critere,
        'classement': [
            {'produit': nom, 'quantite_vendue': _round2(qte), 'chiffre_affaires_ttc': _round2(ttc)}
            for nom, qte, ttc in rows
        ],
    }


# ---------------------------------------------------------------------------
# Sorties de stock (ventes, pertes, corrections, transferts...)
# ---------------------------------------------------------------------------

def _quantite_sortie(row):
    return (
        (row.quantite_unites_sortie or 0)
        + (row.quantite_sous_unites_sortie or 0)
        + (row.quantite_sous_sous_unites_sortie or 0)
    )


def tool_sorties_stock_periode(user, periode='aujourd_hui', date_debut=None, date_fin=None):
    denied = _check_access(user, 'stats_sorties_stock')
    if denied:
        return denied
    start_dt, end_dt, label = _resolve_periode(periode, date_debut, date_fin)

    count, total_ht, total_ttc, u, su, ssu = db.session.query(
        func.count(StockExitLog.id),
        func.coalesce(func.sum(StockExitLog.total_sortie_ht), 0.0),
        func.coalesce(func.sum(StockExitLog.total_sortie_ttc), 0.0),
        func.coalesce(func.sum(StockExitLog.quantite_unites_sortie), 0),
        func.coalesce(func.sum(StockExitLog.quantite_sous_unites_sortie), 0),
        func.coalesce(func.sum(StockExitLog.quantite_sous_sous_unites_sortie), 0),
    ).filter(
        StockExitLog.created_at >= start_dt,
        StockExitLog.created_at <= end_dt
    ).one()

    raisons = db.session.query(
        StockExitLog.reason_nom,
        func.count(StockExitLog.id),
        func.coalesce(func.sum(StockExitLog.total_sortie_ttc), 0.0)
    ).filter(
        StockExitLog.created_at >= start_dt,
        StockExitLog.created_at <= end_dt
    ).group_by(StockExitLog.reason_nom).order_by(func.sum(StockExitLog.total_sortie_ttc).desc()).all()

    return {
        'periode': label,
        'nombre_sorties': int(count or 0),
        'quantite_totale_sortie': int((u or 0) + (su or 0) + (ssu or 0)),
        'valeur_totale_ht': _round2(total_ht),
        'valeur_totale_ttc': _round2(total_ttc),
        'repartition_par_raison': [
            {'raison': raison or 'Non précisée', 'nombre': int(n), 'valeur_ttc': _round2(v)}
            for raison, n, v in raisons
        ],
    }


def tool_dernieres_sorties_stock(user, periode='aujourd_hui', date_debut=None, date_fin=None, limite=10):
    denied = _check_access(user, 'stats_sorties_stock')
    if denied:
        return denied
    limite = max(1, min(int(limite or 10), 50))
    start_dt, end_dt, label = _resolve_periode(periode, date_debut, date_fin)

    rows = StockExitLog.query.filter(
        StockExitLog.created_at >= start_dt,
        StockExitLog.created_at <= end_dt
    ).order_by(StockExitLog.created_at.desc()).limit(limite).all()

    return {
        'periode': label,
        'nombre': len(rows),
        'sorties': [
            {
                'produit': r.produit_nom,
                'quantite': _quantite_sortie(r),
                'raison': r.reason_nom,
                'effectue_par': f'{r.user_prenom} {r.user_nom}'.strip(),
                'valeur_ttc': _round2(r.total_sortie_ttc),
                'date_heure': r.created_at.strftime('%Y-%m-%d %H:%M') if r.created_at else None,
            }
            for r in rows
        ],
    }


def tool_sorties_stock_produit(user, nom_produit, periode='ce_mois', date_debut=None, date_fin=None):
    denied = _check_access(user, 'stats_sorties_stock')
    if denied:
        return denied
    nom_produit = (nom_produit or '').strip()
    if not nom_produit:
        raise ValueError("Le parametre nom_produit est requis.")
    start_dt, end_dt, label = _resolve_periode(periode, date_debut, date_fin)

    rows = db.session.query(
        StockExitLog.produit_nom,
        func.count(StockExitLog.id),
        func.coalesce(func.sum(StockExitLog.quantite_unites_sortie), 0),
        func.coalesce(func.sum(StockExitLog.quantite_sous_unites_sortie), 0),
        func.coalesce(func.sum(StockExitLog.quantite_sous_sous_unites_sortie), 0),
        func.coalesce(func.sum(StockExitLog.total_sortie_ttc), 0.0),
    ).filter(
        StockExitLog.created_at >= start_dt,
        StockExitLog.created_at <= end_dt,
        StockExitLog.produit_nom.ilike(f'%{nom_produit}%')
    ).group_by(StockExitLog.produit_nom).all()

    if not rows:
        return {'trouve': False, 'message': f"Aucune sortie de stock pour '{nom_produit}' sur cette periode ({label})."}

    return {
        'trouve': True,
        'periode': label,
        'resultats': [
            {
                'produit': nom,
                'nombre_sorties': int(n),
                'quantite_totale': int((u or 0) + (su or 0) + (ssu or 0)),
                'valeur_ttc': _round2(v),
            }
            for nom, n, u, su, ssu, v in rows
        ],
    }


# ---------------------------------------------------------------------------
# Commandes fournisseurs
# ---------------------------------------------------------------------------

_STATUTS_COMMANDE = ('en_cours', 'livree', 'annulee')


def _serialize_commande(c, avec_lignes=False):
    lignes = c.lignes
    data = {
        'numero': c.numero,
        'statut': c.statut,
        'fournisseur': c.fournisseur_nom,
        'note': c.note,
        'est_une_relance_de': c.relance_de_numero,
        'creee_le': c.created_at.strftime('%Y-%m-%d %H:%M') if c.created_at else None,
        'creee_par': c.created_by_nom,
        'livree_le': c.livree_at.strftime('%Y-%m-%d %H:%M') if c.livree_at else None,
        'receptionnee_par': c.livree_by_nom,
        'nombre_lignes': len(lignes),
        'quantite_totale_commandee': int(c.total_commande or 0),
        'montant_commande_ht': _round2(c.montant_commande_ht),
    }
    if c.statut == 'livree':
        data.update({
            'quantite_totale_livree': int(c.total_livre or 0),
            'montant_livre_ht': _round2(c.montant_livre_ht),
            'nombre_lignes_avec_ecart': c.nb_lignes_ecart,
            'total_unites_manquantes': int(c.total_manquant or 0),
        })
    if avec_lignes:
        data['lignes'] = [
            {
                'produit': l.produit_nom,
                'code_produit': l.produit_code,
                'quantite_commandee': int(l.quantite_commandee or 0),
                'quantite_livree': int(l.quantite_livree) if l.quantite_livree is not None else None,
                'ecart': int(l.ecart) if l.ecart is not None else None,
                'prix_unite_ht': _round2(l.prix_unite_ht),
                'montant_commande_ht': _round2(l.montant_commande_ht),
                'stock_unites_au_moment_de_la_commande': int(l.stock_unites_au_moment or 0),
            }
            for l in lignes
        ]
    return data


def tool_liste_commandes(user, statut=None, fournisseur=None, periode=None, date_debut=None, date_fin=None, limite=10):
    denied = _check_access(user, 'gestion_commandes')
    if denied:
        return denied
    limite = max(1, min(int(limite or 10), 50))
    query = Commande.query

    statut = (statut or '').strip()
    if statut:
        if statut not in _STATUTS_COMMANDE:
            raise ValueError(f"Statut inconnu: '{statut}'. Valeurs valides : {', '.join(_STATUTS_COMMANDE)}.")
        query = query.filter(Commande.statut == statut)

    fournisseur = (fournisseur or '').strip()
    if fournisseur:
        query = query.filter(Commande.fournisseur_nom.ilike(f'%{fournisseur}%'))

    label_periode = None
    if periode:
        start_dt, end_dt, label_periode = _resolve_periode(periode, date_debut, date_fin)
        query = query.filter(Commande.created_at >= start_dt, Commande.created_at <= end_dt)

    total = query.count()
    commandes = query.order_by(Commande.created_at.desc()).limit(limite).all()

    return {
        'periode': label_periode or 'toutes périodes',
        'nombre_total_correspondant': total,
        'commandes': [_serialize_commande(c) for c in commandes],
    }


def _trouver_commande(recherche=None):
    """Retrouve une commande par numero (partiel) puis nom de fournisseur ; sans
    recherche, la plus recente. Retourne (commande, dict_erreur)."""
    recherche = (recherche or '').strip()
    if recherche:
        commande = Commande.query.filter(
            Commande.numero.ilike(f'%{recherche}%')
        ).order_by(Commande.created_at.desc()).first()
        if not commande:
            commande = Commande.query.filter(
                Commande.fournisseur_nom.ilike(f'%{recherche}%')
            ).order_by(Commande.created_at.desc()).first()
        if not commande:
            return None, {'trouve': False, 'message': f"Aucune commande ne correspond a '{recherche}' (ni par numero, ni par fournisseur)."}
    else:
        commande = Commande.query.order_by(Commande.created_at.desc()).first()
        if not commande:
            return None, {'trouve': False, 'message': "Aucune commande enregistree pour le moment."}
    return commande, None


def tool_detail_commande(user, recherche=None):
    denied = _check_access(user, 'gestion_commandes')
    if denied:
        return denied
    commande, erreur = _trouver_commande(recherche)
    if erreur:
        return erreur

    relances = Commande.query.filter_by(relance_de_numero=commande.numero).order_by(Commande.created_at.asc()).all()
    return {
        'trouve': True,
        **_serialize_commande(commande, avec_lignes=True),
        'relances_de_cette_commande': [
            {'numero': r.numero, 'statut': r.statut, 'creee_le': r.created_at.strftime('%Y-%m-%d %H:%M') if r.created_at else None}
            for r in relances
        ],
    }


def tool_commandes_produit(user, nom_produit, limite=10):
    denied = _check_access(user, 'gestion_commandes')
    if denied:
        return denied
    nom_produit = (nom_produit or '').strip()
    if not nom_produit:
        raise ValueError("Le parametre nom_produit est requis.")
    limite = max(1, min(int(limite or 10), 50))

    rows = db.session.query(CommandeLigne, Commande).join(
        Commande, CommandeLigne.commande_id == Commande.id
    ).filter(
        or_(
            CommandeLigne.produit_nom.ilike(f'%{nom_produit}%'),
            CommandeLigne.produit_code.ilike(f'%{nom_produit}%')
        )
    ).order_by(Commande.created_at.desc()).limit(limite).all()

    if not rows:
        return {'trouve': False, 'message': f"Aucune commande ne contient de produit correspondant a '{nom_produit}'."}

    return {
        'trouve': True,
        'nombre_lignes': len(rows),
        'commandes': [
            {
                'numero': c.numero,
                'date': c.created_at.strftime('%Y-%m-%d %H:%M') if c.created_at else None,
                'fournisseur': c.fournisseur_nom,
                'statut': c.statut,
                'produit': l.produit_nom,
                'quantite_commandee': int(l.quantite_commandee or 0),
                'quantite_livree': int(l.quantite_livree) if l.quantite_livree is not None else None,
                'ecart': int(l.ecart) if l.ecart is not None else None,
            }
            for l, c in rows
        ],
    }


def tool_stats_commandes(user, periode='ce_mois', date_debut=None, date_fin=None):
    denied = _check_access(user, 'gestion_commandes')
    if denied:
        return denied
    start_dt, end_dt, label = _resolve_periode(periode, date_debut, date_fin)
    commandes = Commande.query.filter(
        Commande.created_at >= start_dt,
        Commande.created_at <= end_dt
    ).all()

    par_statut = {s: 0 for s in _STATUTS_COMMANDE}
    montant_commande = 0.0
    montant_livre = 0.0
    nb_avec_ecart = 0
    unites_manquantes = 0
    par_fournisseur = {}
    for c in commandes:
        par_statut[c.statut] = par_statut.get(c.statut, 0) + 1
        montant_commande += c.montant_commande_ht
        if c.statut == 'livree':
            montant_livre += c.montant_livre_ht
            if c.a_ecart:
                nb_avec_ecart += 1
            unites_manquantes += c.total_manquant
        entry = par_fournisseur.setdefault(c.fournisseur_nom, {'nombre_commandes': 0, 'montant_commande_ht': 0.0})
        entry['nombre_commandes'] += 1
        entry['montant_commande_ht'] += c.montant_commande_ht

    top_fournisseurs = sorted(
        (
            {'fournisseur': nom, 'nombre_commandes': v['nombre_commandes'], 'montant_commande_ht': _round2(v['montant_commande_ht'])}
            for nom, v in par_fournisseur.items()
        ),
        key=lambda e: e['montant_commande_ht'], reverse=True
    )[:5]

    return {
        'periode': label,
        'nombre_commandes': len(commandes),
        'par_statut': par_statut,
        'montant_total_commande_ht': _round2(montant_commande),
        'montant_total_livre_ht': _round2(montant_livre),
        'nombre_commandes_livrees_avec_ecart': nb_avec_ecart,
        'total_unites_manquantes_a_la_livraison': int(unites_manquantes),
        'top_fournisseurs_par_montant': top_fournisseurs,
    }


def tool_generer_bon_commande_pdf(user, recherche=None):
    """Genere le bon de commande PDF OFFICIEL d'une commande : passe par exactement
    le meme constructeur que l'export du module Commandes (bon_commande_pdf.py),
    donc le document est identique a celui genere par l'application."""
    denied = _check_access(user, 'gestion_commandes')
    if denied:
        return denied
    commande, erreur = _trouver_commande(recherche)
    if erreur:
        return erreur

    titre = f'Bon de commande {commande.numero}'
    filepath, filename, nom_public = _preparer_fichier_rapport(titre, 'pdf')
    build_bon_commande_pdf(
        commande, filepath,
        tire_par=f'{user.nom} {user.prenom}' if user else 'Assistant IA',
        pharmacy_name=Setting.get_value('pharmacy_name', 'REFLEXPHARMA'))
    return {
        'pdf_genere': True,
        'trouve': True,
        'titre': titre,
        'numero_commande': commande.numero,
        'fournisseur': commande.fournisseur_nom,
        'statut': commande.statut,
        'nom_fichier': nom_public,
        'url_telechargement': url_for('admin.assistant_download_report', filename=filename),
    }


# ---------------------------------------------------------------------------
# Finance (chiffre d'affaires, benefice, solde de tresorerie, operations
# financieres manuelles). Le benefice est toujours une donnee pure deduite des
# ventes ; le solde est un cumul ajuste par les operations manuelles, qui
# n'ecrivent jamais dans Vente/VenteLigne (voir finance_reports.py).
# ---------------------------------------------------------------------------

def tool_resume_finance(user, periode='ce_mois', date_debut=None, date_fin=None):
    """Chiffre d'affaires et benefice sur la periode demandee, et solde de
    tresorerie actuel. Le solde est TOUJOURS global (cumul du benefice depuis
    toujours, ajuste par les encaissements/decaissements manuels) : il ne
    depend pas de la periode demandee ici, contrairement au CA et au benefice."""
    denied = _check_access(user, 'gestion_finance')
    if denied:
        return denied
    start_dt, end_dt, label = _resolve_periode(periode, date_debut, date_fin)
    ventes = _ventes_periode_fiscale(start_dt, end_dt)
    totaux = _totaux_fiscaux(ventes)
    encaissements_periode, decaissements_periode = compute_totaux_operations(start_dt, end_dt)

    return {
        'periode': label,
        'nombre_ventes': totaux['nombre_ventes'],
        'chiffre_affaires_ttc': totaux['total_ttc'],
        'chiffre_affaires_ht': totaux['total_ht'],
        'benefice_periode': totaux['total_benefice'],
        'encaissements_manuels_sur_la_periode': _round2(encaissements_periode),
        'decaissements_manuels_sur_la_periode': _round2(decaissements_periode),
        'solde_actuel': _round2(compute_solde_actuel()),
        'note_solde': (
            "Le solde actuel est un cumul depuis toujours (benefice total + encaissements - "
            "decaissements), independant de la periode demandee ci-dessus."
        ),
    }


def tool_operations_financieres(user, type_operation=None, periode=None, date_debut=None, date_fin=None, limite=20):
    """Liste les operations financieres manuelles (encaissements/decaissements),
    les plus recentes en premier. Sans periode : tout l'historique."""
    denied = _check_access(user, 'gestion_finance')
    if denied:
        return denied
    limite = max(1, min(int(limite or 20), 100))
    type_operation = (type_operation or '').strip()
    if type_operation and type_operation not in ('encaissement', 'decaissement'):
        raise ValueError("Parametre 'type_operation' invalide. Valeurs valides : encaissement, decaissement.")

    start_dt = end_dt = None
    label = 'toutes périodes'
    if periode:
        start_dt, end_dt, label = _resolve_periode(periode, date_debut, date_fin)

    operations = query_operations_financieres(start_dt, end_dt, type_operation or None)[:limite]
    return {
        'periode': label,
        'nombre_operations': len(operations),
        'operations': [
            {
                'type': o.type,
                'montant': _round2(o.montant),
                'raison': o.raison,
                'note': o.note,
                'enregistre_par': o.created_by_nom,
                'date': o.created_at.strftime('%Y-%m-%d %H:%M') if o.created_at else None,
            }
            for o in operations
        ],
    }


def tool_generer_export_operations_financieres_pdf(user, periode=None, date_debut=None, date_fin=None):
    """Genere l'export PDF OFFICIEL des operations financieres : passe par
    exactement le meme constructeur que l'export du module Finance
    (finance_reports.py), donc le document est identique a celui de
    l'application. Sans periode : tout l'historique des operations."""
    denied = _check_access(user, 'gestion_finance')
    if denied:
        return denied
    start_dt = end_dt = None
    label = 'toutes périodes'
    if periode:
        start_dt, end_dt, label = _resolve_periode(periode, date_debut, date_fin)

    operations = query_operations_financieres(start_dt, end_dt)
    titre = 'Opérations financières'
    filepath, filename, nom_public = _preparer_fichier_rapport(titre, 'pdf')
    build_operations_financieres_pdf(
        filepath, operations, label,
        tire_par=f'{user.nom} {user.prenom}' if user else 'Assistant IA',
        pharmacy_name=Setting.get_value('pharmacy_name', 'REFLEXPHARMA'),
        solde_actuel=compute_solde_actuel())
    return {
        'pdf_genere': True,
        'titre': titre,
        'periode': label,
        'nombre_operations': len(operations),
        'nom_fichier': nom_public,
        'url_telechargement': url_for('admin.assistant_download_report', filename=filename),
    }


def tool_generer_export_operations_financieres_excel(user, periode=None, date_debut=None, date_fin=None):
    """Comme tool_generer_export_operations_financieres_pdf, mais produit le
    classeur Excel officiel des operations financieres."""
    denied = _check_access(user, 'gestion_finance')
    if denied:
        return denied
    start_dt = end_dt = None
    label = 'toutes périodes'
    if periode:
        start_dt, end_dt, label = _resolve_periode(periode, date_debut, date_fin)

    operations = query_operations_financieres(start_dt, end_dt)
    titre = 'Opérations financières'
    filepath, filename, nom_public = _preparer_fichier_rapport(titre, 'xlsx')
    build_operations_financieres_excel(filepath, operations, label, solde_actuel=compute_solde_actuel())
    return {
        'excel_genere': True,
        'titre': titre,
        'periode': label,
        'nombre_operations': len(operations),
        'nom_fichier': nom_public,
        'url_telechargement': url_for('admin.assistant_download_report', filename=filename),
    }


# ---------------------------------------------------------------------------
# Clients
# ---------------------------------------------------------------------------

def tool_top_clients(user, periode='ce_mois', limite=5, date_debut=None, date_fin=None):
    denied = _check_access(user, 'gestion_clients')
    if denied:
        return denied
    start_dt, end_dt, label = _resolve_periode(periode, date_debut, date_fin)
    limite = max(1, min(int(limite or 5), 20))

    rows = db.session.query(
        Vente.client_nom, Vente.client_prenom,
        func.count(Vente.id), func.coalesce(func.sum(Vente.total_ttc), 0.0)
    ).filter(
        Vente.created_at >= start_dt,
        Vente.created_at <= end_dt,
        Vente.statut != 'annulee',
        Vente.client_id.isnot(None)
    ).group_by(Vente.client_id, Vente.client_nom, Vente.client_prenom).order_by(
        func.sum(Vente.total_ttc).desc()
    ).limit(limite).all()

    return {
        'periode': label,
        'classement': [
            {
                'client': (f'{prenom or ""} {nom or ""}'.strip() or 'Client inconnu'),
                'nombre_achats': int(count or 0),
                'chiffre_affaires_ttc': _round2(ttc),
            }
            for nom, prenom, count, ttc in rows
        ],
    }


def tool_solde_client(user, recherche):
    denied = _check_access(user, 'gestion_clients')
    if denied:
        return denied
    recherche = (recherche or '').strip()
    if not recherche:
        raise ValueError("Le parametre 'recherche' (nom, prenom, matricule ou email) est requis.")

    clients = Client.query.filter(
        or_(
            Client.nom.ilike(f'%{recherche}%'),
            Client.prenom.ilike(f'%{recherche}%'),
            Client.matricule.ilike(f'%{recherche}%'),
            Client.email.ilike(f'%{recherche}%')
        )
    ).limit(5).all()

    if not clients:
        return {'trouve': False, 'message': f"Aucun client ne correspond a '{recherche}'."}

    return {
        'trouve': True,
        'resultats': [
            {'client': c.nom_complet, 'matricule': c.matricule, 'solde': _round2(c.solde)}
            for c in clients
        ],
    }


def tool_nombre_clients(user):
    denied = _check_access(user, 'gestion_clients')
    if denied:
        return denied
    total = db.session.query(func.count(Client.id)).scalar() or 0
    return {'nombre_clients': int(total)}


def tool_liste_clients(user, recherche=None, limite=50):
    denied = _check_access(user, 'gestion_clients')
    if denied:
        return denied
    limite = max(1, min(int(limite or 50), 200))
    query = Client.query
    recherche = (recherche or '').strip()
    if recherche:
        query = query.filter(
            or_(
                Client.nom.ilike(f'%{recherche}%'),
                Client.prenom.ilike(f'%{recherche}%'),
                Client.matricule.ilike(f'%{recherche}%'),
                Client.email.ilike(f'%{recherche}%')
            )
        )

    total = query.count()
    clients = query.order_by(Client.nom.asc()).limit(limite).all()
    return {
        'nombre_total_correspondant': total,
        'clients': [
            {
                'client': c.nom_complet,
                'matricule': c.matricule,
                'groupe': c.groupe.nom if c.groupe else None,
                'solde': _round2(c.solde),
            }
            for c in clients
        ],
    }


def tool_nombre_groupes_clients(user):
    denied = _check_access(user, 'gestion_groupes_clients')
    if denied:
        return denied
    total = db.session.query(func.count(GroupeClient.id)).scalar() or 0
    return {'nombre_groupes_clients': int(total)}


def tool_liste_groupes_clients(user):
    denied = _check_access(user, 'gestion_groupes_clients')
    if denied:
        return denied
    rows = db.session.query(
        GroupeClient.nom, GroupeClient.solde, GroupeClient.pourcentage_absorption,
        func.count(Client.id)
    ).outerjoin(Client, Client.groupe_id == GroupeClient.id).group_by(
        GroupeClient.id, GroupeClient.nom, GroupeClient.solde, GroupeClient.pourcentage_absorption
    ).order_by(GroupeClient.nom.asc()).all()

    return {
        'nombre_groupes': len(rows),
        'groupes': [
            {
                'groupe': nom,
                'solde': _round2(solde),
                'pourcentage_absorption': _round2(pct),
                'nombre_clients': int(count),
            }
            for nom, solde, pct, count in rows
        ],
    }


def tool_solde_groupe_client(user, recherche):
    denied = _check_access(user, 'gestion_groupes_clients')
    if denied:
        return denied
    recherche = (recherche or '').strip()
    if not recherche:
        raise ValueError("Le parametre 'recherche' (nom du groupe) est requis.")

    groupes = GroupeClient.query.filter(GroupeClient.nom.ilike(f'%{recherche}%')).limit(5).all()
    if not groupes:
        return {'trouve': False, 'message': f"Aucun groupe client ne correspond a '{recherche}'."}

    return {
        'trouve': True,
        'resultats': [
            {
                'groupe': g.nom,
                'solde': _round2(g.solde),
                'pourcentage_absorption': _round2(g.pourcentage_absorption),
                'nombre_clients': len(g.clients),
            }
            for g in groupes
        ],
    }


def tool_clients_par_groupe(user, nom_groupe):
    denied = _check_access(user, 'gestion_groupes_clients')
    if denied:
        return denied
    nom_groupe = (nom_groupe or '').strip()
    if not nom_groupe:
        raise ValueError("Le parametre 'nom_groupe' est requis.")

    groupe = GroupeClient.query.filter(GroupeClient.nom.ilike(f'%{nom_groupe}%')).first()
    if not groupe:
        return {'trouve': False, 'message': f"Aucun groupe client ne correspond a '{nom_groupe}'."}

    clients = Client.query.filter_by(groupe_id=groupe.id).order_by(Client.nom.asc()).all()
    return {
        'trouve': True,
        'groupe': groupe.nom,
        'nombre_clients': len(clients),
        'clients': [
            {'client': c.nom_complet, 'matricule': c.matricule, 'solde': _round2(c.solde)}
            for c in clients
        ],
    }


def tool_clients_sans_groupe(user, limite=15):
    denied = _check_access(user, 'gestion_clients')
    if denied:
        return denied
    limite = max(1, min(int(limite or 15), 50))
    total = db.session.query(func.count(Client.id)).filter(Client.groupe_id.is_(None)).scalar() or 0
    clients = Client.query.filter(Client.groupe_id.is_(None)).order_by(Client.nom.asc()).limit(limite).all()
    return {
        'nombre_total_sans_groupe': int(total),
        'apercu': [{'client': c.nom_complet, 'matricule': c.matricule} for c in clients],
    }


def tool_top_clients_solde(user, limite=10):
    denied = _check_access(user, 'gestion_clients')
    if denied:
        return denied
    limite = max(1, min(int(limite or 10), 50))
    clients = Client.query.order_by(Client.solde.desc()).limit(limite).all()
    return {
        'classement': [
            {'client': c.nom_complet, 'matricule': c.matricule, 'solde': _round2(c.solde)}
            for c in clients
        ],
    }


def tool_solde_total_clients_et_groupes(user):
    denied = _check_access(user, 'gestion_clients')
    if denied:
        return denied
    total_clients = db.session.query(func.coalesce(func.sum(Client.solde), 0.0)).scalar() or 0
    total_groupes = db.session.query(func.coalesce(func.sum(GroupeClient.solde), 0.0)).scalar() or 0
    return {
        'solde_total_clients': _round2(total_clients),
        'solde_total_groupes_clients': _round2(total_groupes),
        'solde_total_general': _round2(total_clients + total_groupes),
    }


# ---------------------------------------------------------------------------
# Inventaires
# ---------------------------------------------------------------------------

def _inventaire_personne_label(u):
    return f'{u.prenom} {u.nom}'.strip() if u else None


def tool_liste_inventaires(user, limite=10):
    """ReflexPharma ne planifie pas d'inventaires futurs : il n'existe qu'un inventaire
    'en_cours' au plus (le cas echeant) et l'historique des inventaires passes (valide/annule)."""
    denied = _check_access(user, 'gestion_inventaire')
    if denied:
        return denied
    limite = max(1, min(int(limite or 10), 50))
    inventaires = Inventaire.query.order_by(Inventaire.created_at.desc()).limit(limite).all()

    resultats = []
    for inv in inventaires:
        lignes = inv.lignes
        resultats.append({
            'titre': inv.titre,
            'statut': inv.statut,
            'cree_le': inv.created_at.strftime('%Y-%m-%d %H:%M') if inv.created_at else None,
            'cree_par': _inventaire_personne_label(inv.created_by),
            'valide_le': inv.validated_at.strftime('%Y-%m-%d %H:%M') if inv.validated_at else None,
            'valide_par': _inventaire_personne_label(inv.validated_by),
            'nombre_produits': len(lignes),
            'nombre_produits_comptes': sum(1 for l in lignes if l.is_scanned),
            'nombre_ecarts': sum(1 for l in lignes if l.a_decalage),
        })

    return {
        'nombre_inventaires': len(resultats),
        'inventaires': resultats,
    }


def tool_detail_inventaire(user, recherche=None):
    """Sans recherche : renvoie l'inventaire actuellement en cours s'il y en a un, sinon le
    plus recent (donc le 'dernier inventaire')."""
    denied = _check_access(user, 'gestion_inventaire')
    if denied:
        return denied
    recherche = (recherche or '').strip()
    if recherche:
        inv = Inventaire.query.filter(Inventaire.titre.ilike(f'%{recherche}%')).order_by(Inventaire.created_at.desc()).first()
    else:
        inv = Inventaire.query.filter_by(statut='en_cours').order_by(Inventaire.created_at.desc()).first()
        if not inv:
            inv = Inventaire.query.order_by(Inventaire.created_at.desc()).first()

    if not inv:
        return {'trouve': False, 'message': "Aucun inventaire trouve."}

    lignes = inv.lignes
    ecarts = [l for l in lignes if l.a_decalage]

    return {
        'trouve': True,
        'titre': inv.titre,
        'statut': inv.statut,
        'cree_le': inv.created_at.strftime('%Y-%m-%d %H:%M') if inv.created_at else None,
        'cree_par': _inventaire_personne_label(inv.created_by),
        'valide_le': inv.validated_at.strftime('%Y-%m-%d %H:%M') if inv.validated_at else None,
        'valide_par': _inventaire_personne_label(inv.validated_by),
        'nombre_produits_total': len(lignes),
        'nombre_produits_comptes': sum(1 for l in lignes if l.is_scanned),
        'nombre_ecarts': len(ecarts),
        'ecarts': [
            {
                'produit': l.produit.nom if l.produit else None,
                'code_suivi': l.code_suivi,
                'quantite_avant': l.total_avant,
                'quantite_apres': l.total_apres,
                'constate_par': _inventaire_personne_label(l.constate_by),
            }
            for l in ecarts[:30]
        ],
        'ecarts_tronques': len(ecarts) > 30,
    }


# ---------------------------------------------------------------------------
# Impots / declaration des taxes
# ---------------------------------------------------------------------------

def _ventes_periode_fiscale(start_dt, end_dt):
    """Ventes validees sur un intervalle : la meme base que le module Impots de l'application."""
    return Vente.query.filter(
        Vente.statut == 'validee',
        Vente.created_at >= start_dt,
        Vente.created_at <= end_dt
    ).order_by(Vente.created_at.asc()).all()


def _totaux_fiscaux(ventes):
    """HT / TVA effective / benefice / TTC agreges en SQL sur les lignes de vente,
    comme compute_ventes_totals_reels() cote vues (TVA reelle hors marge coefficient)."""
    numeros = [v.numero_vente for v in ventes]
    tva_reelle, benefice = 0.0, 0.0
    if numeros:
        tva_expr = VenteLigne.total_ht * (VenteLigne.tva_pourcentage / 100.0)
        benefice_expr = func.greatest(VenteLigne.total_ttc - VenteLigne.total_ht - tva_expr, 0.0)
        tva_sum, benefice_sum = db.session.query(
            func.coalesce(func.sum(tva_expr), 0.0),
            func.coalesce(func.sum(benefice_expr), 0.0)
        ).filter(VenteLigne.numero_vente.in_(numeros)).one()
        tva_reelle, benefice = float(tva_sum or 0), float(benefice_sum or 0)
    return {
        'nombre_ventes': len(ventes),
        'total_ht': _round2(sum(v.total_ht or 0 for v in ventes)),
        'total_tva': _round2(tva_reelle),
        'total_benefice': _round2(benefice),
        'total_ttc': _round2(sum(v.total_ttc or 0 for v in ventes)),
    }


def _repartition_tva(ventes):
    numeros = [v.numero_vente for v in ventes]
    if not numeros:
        return []
    tva_expr = VenteLigne.total_ht * (VenteLigne.tva_pourcentage / 100.0)
    rows = db.session.query(
        VenteLigne.tva_pourcentage,
        func.coalesce(func.sum(VenteLigne.total_ht), 0.0),
        func.coalesce(func.sum(tva_expr), 0.0)
    ).filter(
        VenteLigne.numero_vente.in_(numeros)
    ).group_by(VenteLigne.tva_pourcentage).order_by(VenteLigne.tva_pourcentage).all()
    return [
        {'taux_tva_pourcentage': _round2(taux), 'base_ht': _round2(ht), 'montant_tva': _round2(tva)}
        for taux, ht, tva in rows
    ]


def _declaration_bornes_dt(declaration):
    start_dt = datetime.combine(declaration.periode_debut, datetime.min.time())
    end_dt = datetime.combine(declaration.periode_fin, datetime.max.time())
    return start_dt, end_dt


def _serialize_declaration(declaration):
    if declaration.est_declaree:
        totaux = {
            'nombre_ventes': declaration.nb_ventes,
            'total_ht': _round2(declaration.total_ht),
            'total_tva': _round2(declaration.total_tva),
            'total_benefice': _round2(declaration.total_benefice),
            'total_ttc': _round2(declaration.total_ttc),
        }
    else:
        start_dt, end_dt = _declaration_bornes_dt(declaration)
        totaux = _totaux_fiscaux(_ventes_periode_fiscale(start_dt, end_dt))
    return {
        'reference': declaration.reference,
        'periode_debut': declaration.periode_debut.strftime('%Y-%m-%d'),
        'periode_fin': declaration.periode_fin.strftime('%Y-%m-%d'),
        'statut': 'declaree' if declaration.est_declaree else 'en_preparation',
        'note': declaration.note,
        'totaux_geles_a_la_declaration': declaration.est_declaree,
        **totaux,
        'creee_le': declaration.created_at.strftime('%Y-%m-%d %H:%M') if declaration.created_at else None,
        'creee_par': _inventaire_personne_label(declaration.created_by),
        'declaree_le': declaration.declared_at.strftime('%Y-%m-%d %H:%M') if declaration.declared_at else None,
        'declaree_par': _inventaire_personne_label(declaration.declared_by),
    }


def tool_liste_declarations_impots(user, statut=None, limite=10):
    denied = _check_access(user, 'module_impots')
    if denied:
        return denied
    limite = max(1, min(int(limite or 10), 50))
    query = DeclarationImpot.query.order_by(DeclarationImpot.periode_debut.desc())
    statut = (statut or '').strip()
    if statut:
        if statut not in ('declaree', 'en_preparation'):
            raise ValueError("Statut inconnu. Valeurs valides : declaree, en_preparation.")
        query = query.filter(DeclarationImpot.statut == statut)
    declarations = query.limit(limite).all()
    return {
        'nombre_declarations': len(declarations),
        'declarations': [_serialize_declaration(d) for d in declarations],
    }


def tool_detail_declaration_impot(user, recherche=None):
    """Sans recherche : renvoie la declaration la plus recente (par periode). La recherche
    accepte une reference (ex: IMP-123456, partielle) ou une date AAAA-MM-JJ contenue
    dans la periode."""
    denied = _check_access(user, 'module_impots')
    if denied:
        return denied
    recherche = (recherche or '').strip()
    declaration = None
    if recherche:
        try:
            jour = datetime.strptime(recherche, '%Y-%m-%d').date()
            declaration = DeclarationImpot.query.filter(
                DeclarationImpot.periode_debut <= jour,
                DeclarationImpot.periode_fin >= jour
            ).order_by(DeclarationImpot.periode_debut.desc()).first()
        except ValueError:
            declaration = DeclarationImpot.query.filter(
                DeclarationImpot.reference.ilike(f'%{recherche}%')
            ).order_by(DeclarationImpot.periode_debut.desc()).first()
    else:
        declaration = DeclarationImpot.query.order_by(DeclarationImpot.periode_debut.desc()).first()

    if not declaration:
        return {'trouve': False, 'message': "Aucune declaration d'impots trouvee."}

    start_dt, end_dt = _declaration_bornes_dt(declaration)
    ventes = _ventes_periode_fiscale(start_dt, end_dt)
    return {
        'trouve': True,
        **_serialize_declaration(declaration),
        'repartition_par_taux_tva': _repartition_tva(ventes),
        'ventes': [
            {
                'numero_vente': v.numero_vente,
                'date': v.created_at.strftime('%Y-%m-%d %H:%M') if v.created_at else None,
                'client': v.client_label,
                'total_ht': _round2(v.total_ht),
                'total_ttc': _round2(v.total_ttc),
            }
            for v in ventes[:30]
        ],
        'ventes_tronquees': len(ventes) > 30,
    }


def tool_taxes_a_declarer_periode(user, periode='mois_dernier', date_debut=None, date_fin=None):
    """Montants fiscaux (HT, TVA par taux, TTC) sur une periode quelconque + verification
    de la couverture de cette periode par les declarations existantes."""
    denied = _check_access(user, 'module_impots')
    if denied:
        return denied
    start_dt, end_dt, label = _resolve_periode(periode, date_debut, date_fin)
    ventes = _ventes_periode_fiscale(start_dt, end_dt)

    declarations = DeclarationImpot.query.filter(
        DeclarationImpot.periode_debut <= end_dt.date(),
        DeclarationImpot.periode_fin >= start_dt.date()
    ).order_by(DeclarationImpot.periode_debut.asc()).all()

    periode_couverte_et_declaree = any(
        d.est_declaree and d.periode_debut <= start_dt.date() and d.periode_fin >= end_dt.date()
        for d in declarations
    )

    return {
        'periode': label,
        'date_debut': start_dt.strftime('%Y-%m-%d'),
        'date_fin': end_dt.strftime('%Y-%m-%d'),
        **_totaux_fiscaux(ventes),
        'repartition_par_taux_tva': _repartition_tva(ventes),
        'periode_entierement_couverte_par_une_declaration_marquee_declaree': periode_couverte_et_declaree,
        'declarations_chevauchant_la_periode': [
            {
                'reference': d.reference,
                'periode_debut': d.periode_debut.strftime('%Y-%m-%d'),
                'periode_fin': d.periode_fin.strftime('%Y-%m-%d'),
                'statut': 'declaree' if d.est_declaree else 'en_preparation',
                'declaree_le': d.declared_at.strftime('%Y-%m-%d %H:%M') if d.declared_at else None,
                'declaree_par': _inventaire_personne_label(d.declared_by),
            }
            for d in declarations
        ],
    }


# ---------------------------------------------------------------------------
# Employes / equipe / permissions
# ---------------------------------------------------------------------------

def tool_nombre_employes(user):
    denied = _check_access(user, 'gestion_employes')
    if denied:
        return denied
    total = db.session.query(func.count(User.id)).scalar() or 0
    actifs = db.session.query(func.count(User.id)).filter(User.is_active.is_(True)).scalar() or 0
    par_role = db.session.query(User.role, func.count(User.id)).group_by(User.role).all()
    return {
        'nombre_total_employes': int(total),
        'nombre_employes_actifs': int(actifs),
        'nombre_employes_inactifs': int(total) - int(actifs),
        'repartition_par_role': {role or 'non precise': int(count) for role, count in par_role},
    }


def _employe_ancienne_annees(u):
    if not u.date_prise_poste:
        return None
    jours = (date.today() - u.date_prise_poste).days
    return round(jours / 365.25, 1)


def tool_liste_employes(user, statut='actifs'):
    denied = _check_access(user, 'gestion_employes')
    if denied:
        return denied
    query = User.query
    if statut == 'actifs':
        query = query.filter(User.is_active.is_(True))
    elif statut == 'inactifs':
        query = query.filter(User.is_active.is_(False))
    # 'tous' -> pas de filtre

    users = query.order_by(User.nom.asc()).all()
    return {
        'statut_filtre': statut,
        'nombre': len(users),
        'employes': [
            {
                'employe': f'{u.prenom} {u.nom}'.strip(),
                'role': u.role,
                'poste': u.poste,
                'actif': bool(u.is_active),
                'anciennete_annees': _employe_ancienne_annees(u),
            }
            for u in users
        ],
    }


def tool_employes_par_poste(user):
    denied = _check_access(user, 'gestion_employes')
    if denied:
        return denied
    rows = db.session.query(User.poste, func.count(User.id)).filter(
        User.is_active.is_(True)
    ).group_by(User.poste).order_by(func.count(User.id).desc()).all()
    return {
        'repartition': [
            {'poste': poste or 'Sans poste attribué', 'nombre_employes': int(count)}
            for poste, count in rows
        ],
    }


def tool_postes_disponibles(user):
    denied = _check_access(user, 'gestion_postes')
    if denied:
        return denied
    postes = Poste.query.order_by(Poste.nom.asc()).all()
    return {
        'nombre_postes': len(postes),
        'postes': [{'poste': p.nom, 'description': p.description} for p in postes],
    }


def tool_acces_module(user, module):
    denied = _check_access(user, 'gestion_employes')
    if denied:
        return denied
    module = (module or '').strip()
    if module not in FEATURES:
        raise ValueError(
            f"Module inconnu: '{module}'. Modules valides: {', '.join(FEATURES.keys())}"
        )

    users = User.query.filter_by(is_active=True).order_by(User.nom.asc()).all()
    avec_acces, sans_acces = [], []
    for u in users:
        entry = {'employe': f'{u.prenom} {u.nom}'.strip(), 'role': u.role, 'poste': u.poste}
        (avec_acces if u.has_permission(module) else sans_acces).append(entry)

    return {
        'module': module,
        'module_label': FEATURES.get(module, module),
        'nombre_avec_acces': len(avec_acces),
        'employes_avec_acces': avec_acces,
        'nombre_sans_acces': len(sans_acces),
        'employes_sans_acces': sans_acces,
    }


def tool_modules_disponibles(user):
    """Liste les modules/fonctionnalites dont l'acces peut etre accorde ou restreint.
    Pas de restriction : ne revele aucune donnee metier, juste les cles/labels."""
    return {'modules': [{'cle': cle, 'nom': nom} for cle, nom in FEATURES.items()]}


def tool_mes_modules_accessibles(user):
    """Modules auxquels LA PERSONNE QUI POSE LA QUESTION a reellement acces (pas la
    liste de tous les modules existants). Utilise le has_permission() reel de
    l'application (pas le bypass admin/superadmin propre a l'assistant IA), pour
    refleter exactement ce que cette personne voit dans le logiciel. Aucune
    restriction : on peut toujours consulter ses propres acces."""
    if user is None:
        return {'error': "Utilisateur inconnu."}

    accessibles = [
        {'cle': cle, 'nom': nom}
        for cle, nom in FEATURES.items()
        if user.has_permission(cle)
    ]

    return {
        'employe': f'{user.prenom} {user.nom}'.strip(),
        'role': user.role,
        'nombre_modules_accessibles': len(accessibles),
        'modules_accessibles': accessibles,
    }


# ---------------------------------------------------------------------------
# Generation de rapport PDF telechargeable depuis le chat
# ---------------------------------------------------------------------------

REPORTS_DIR = os.path.join(
    os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__)))),
    'ai_reports'
)
REPORT_FILENAME_RE = re.compile(r'^[0-9a-f]{32}__[A-Za-z0-9_-]+\.(pdf|xlsx)$')


def _cleanup_old_reports(max_age_hours=48):
    """Best-effort : supprime les rapports generes il y a plus de 48h pour ne pas accumuler de fichiers."""
    try:
        if not os.path.isdir(REPORTS_DIR):
            return
        cutoff = datetime.now().timestamp() - max_age_hours * 3600
        for fname in os.listdir(REPORTS_DIR):
            fpath = os.path.join(REPORTS_DIR, fname)
            if os.path.isfile(fpath) and os.path.getmtime(fpath) < cutoff:
                os.remove(fpath)
    except OSError:
        pass


def _build_pdf_report(filepath, titre, sections):
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import cm
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle

    styles = getSampleStyleSheet()
    title_style = ParagraphStyle('ReportTitle', parent=styles['Title'], textColor=colors.HexColor('#2c3e50'), fontSize=18, spaceAfter=4)
    meta_style = ParagraphStyle('ReportMeta', parent=styles['Normal'], textColor=colors.HexColor('#7a8896'), fontSize=9, spaceAfter=16)
    section_style = ParagraphStyle('SectionTitle', parent=styles['Heading2'], textColor=colors.HexColor('#3498db'), fontSize=13, spaceBefore=14, spaceAfter=6)
    body_style = ParagraphStyle('Body', parent=styles['Normal'], fontSize=10, leading=14)
    header_cell_style = ParagraphStyle('HeaderCell', parent=styles['Normal'], fontSize=9, leading=11, textColor=colors.white, fontName='Helvetica-Bold')
    cell_style = ParagraphStyle('Cell', parent=styles['Normal'], fontSize=9, leading=11)

    doc = SimpleDocTemplate(filepath, pagesize=A4, topMargin=2 * cm, bottomMargin=2 * cm, leftMargin=2 * cm, rightMargin=2 * cm)
    story = [
        Paragraph(xml_escape(titre), title_style),
        Paragraph(f"Généré le {datetime.now().strftime('%d/%m/%Y à %H:%M')} — ReflexPharma", meta_style),
    ]

    for section in sections or []:
        section_title = (section.get('titre_section') or '').strip()
        if section_title:
            story.append(Paragraph(xml_escape(section_title), section_style))

        texte = (section.get('texte') or '').strip()
        if texte:
            for line in texte.split('\n'):
                if line.strip():
                    story.append(Paragraph(xml_escape(line.strip()), body_style))
            story.append(Spacer(1, 6))

        tableau = section.get('tableau') or {}
        colonnes = tableau.get('colonnes') or []
        lignes = tableau.get('lignes') or []
        if colonnes and lignes:
            header = [Paragraph(xml_escape(str(c)), header_cell_style) for c in colonnes]
            body_rows = [
                [Paragraph(xml_escape(str(cell)) if cell is not None else '', cell_style) for cell in row]
                for row in lignes
            ]
            # Le tableau occupe toute la largeur utile de la page (premiere colonne
            # plus large : c'est generalement le libelle)
            largeur_utile = A4[0] - 4 * cm
            nb_colonnes = len(colonnes)
            if nb_colonnes == 1:
                col_widths = [largeur_utile]
            else:
                premiere = largeur_utile * (0.35 if nb_colonnes <= 4 else 0.28)
                autres = (largeur_utile - premiere) / (nb_colonnes - 1)
                col_widths = [premiere] + [autres] * (nb_colonnes - 1)
            table = Table([header] + body_rows, colWidths=col_widths, hAlign='LEFT', repeatRows=1)
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2c3e50')),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f7fafd')]),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#e0e4ea')),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                ('LEFTPADDING', (0, 0), (-1, -1), 6),
                ('RIGHTPADDING', (0, 0), (-1, -1), 6),
                ('TOPPADDING', (0, 0), (-1, -1), 4),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
            ]))
            story.append(table)
            story.append(Spacer(1, 10))

    doc.build(story)


def _build_excel_report(filepath, titre, sections):
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = 'Rapport'
    header_fill = PatternFill(start_color='2C3E50', end_color='2C3E50', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF')

    row = 1
    ws.cell(row=row, column=1, value=titre).font = Font(bold=True, size=14, color='2C3E50')
    row += 1
    ws.cell(row=row, column=1, value=f"Généré le {datetime.now().strftime('%d/%m/%Y à %H:%M')} — ReflexPharma").font = Font(size=9, color='7A8896')
    row += 2

    for section in sections or []:
        section_title = (section.get('titre_section') or '').strip()
        if section_title:
            ws.cell(row=row, column=1, value=section_title).font = Font(bold=True, size=12, color='3498DB')
            row += 1

        texte = (section.get('texte') or '').strip()
        if texte:
            for line in texte.split('\n'):
                if line.strip():
                    ws.cell(row=row, column=1, value=line.strip())
                    row += 1

        tableau = section.get('tableau') or {}
        colonnes = tableau.get('colonnes') or []
        lignes = tableau.get('lignes') or []
        if colonnes and lignes:
            for ci, col in enumerate(colonnes, start=1):
                cell = ws.cell(row=row, column=ci, value=str(col))
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center')
            row += 1
            for ligne in lignes:
                for ci, valeur in enumerate(ligne, start=1):
                    if isinstance(valeur, (int, float)) and not isinstance(valeur, bool):
                        ws.cell(row=row, column=ci, value=valeur)
                    else:
                        ws.cell(row=row, column=ci, value='' if valeur is None else str(valeur))
                row += 1
        row += 1

    for col_idx in range(1, (ws.max_column or 1) + 1):
        longueur = max(
            (len(str(ws.cell(row=r, column=col_idx).value or '')) for r in range(1, ws.max_row + 1)),
            default=0
        )
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max(12, longueur + 2), 50)

    wb.save(filepath)


def _preparer_fichier_rapport(titre, extension):
    """Chemin + nom public d'un nouveau fichier de rapport (token non devinable)."""
    os.makedirs(REPORTS_DIR, exist_ok=True)
    _cleanup_old_reports()
    token = uuid.uuid4().hex
    ascii_titre = unicodedata.normalize('NFKD', titre).encode('ascii', 'ignore').decode('ascii')
    safe_titre = re.sub(r'[^A-Za-z0-9_-]+', '_', ascii_titre).strip('_') or 'rapport'
    filename = f'{token}__{safe_titre}.{extension}'
    return os.path.join(REPORTS_DIR, filename), filename, f'{safe_titre}.{extension}'


def _valider_sections_rapport(titre, sections):
    titre = (titre or 'Rapport ReflexPharma').strip() or 'Rapport ReflexPharma'
    if not isinstance(sections, list) or not sections:
        raise ValueError(
            "Le parametre 'sections' est requis et doit contenir au moins une section "
            "(avec un titre_section et du texte et/ou un tableau)."
        )
    return titre


def tool_generer_rapport_pdf(user, titre, sections):
    """Pas de restriction propre a cet outil : il ne fait que mettre en forme des
    donnees deja obtenues (et donc deja verifiees) via d'autres outils."""
    titre = _valider_sections_rapport(titre, sections)
    filepath, filename, nom_public = _preparer_fichier_rapport(titre, 'pdf')
    _build_pdf_report(filepath, titre, sections)
    return {
        'pdf_genere': True,
        'titre': titre,
        'nom_fichier': nom_public,
        'url_telechargement': url_for('admin.assistant_download_report', filename=filename),
    }


def tool_generer_rapport_excel(user, titre, sections):
    """Comme tool_generer_rapport_pdf mais produit un classeur Excel (.xlsx) : meme
    structure de sections, rendue dans une feuille unique (titres, textes, tableaux)."""
    titre = _valider_sections_rapport(titre, sections)
    filepath, filename, nom_public = _preparer_fichier_rapport(titre, 'xlsx')
    _build_excel_report(filepath, titre, sections)
    return {
        'excel_genere': True,
        'titre': titre,
        'nom_fichier': nom_public,
        'url_telechargement': url_for('admin.assistant_download_report', filename=filename),
    }


# ---------------------------------------------------------------------------
# Registre des outils + dispatch (format compatible Mistral/OpenAI tool-calling)
# ---------------------------------------------------------------------------

TOOL_FUNCTIONS = {
    'chiffre_affaires_periode': tool_chiffre_affaires_periode,
    'comparer_ca_mois_precedent': tool_comparer_ca_mois_precedent,
    'prevision_chiffre_affaires': tool_prevision_chiffre_affaires,
    'employe_du_mois': tool_employe_du_mois,
    'nombre_produits': tool_nombre_produits,
    'liste_produits': tool_liste_produits,
    'liste_fournisseurs': tool_liste_fournisseurs,
    'liste_groupes_fournisseurs': tool_liste_groupes_fournisseurs,
    'stock_produit': tool_stock_produit,
    'produits_stock_faible': tool_produits_stock_faible,
    'produits_peremption_proche': tool_produits_peremption_proche,
    'top_produits_vendus': tool_top_produits_vendus,
    'sorties_stock_periode': tool_sorties_stock_periode,
    'dernieres_sorties_stock': tool_dernieres_sorties_stock,
    'sorties_stock_produit': tool_sorties_stock_produit,
    'top_clients': tool_top_clients,
    'solde_client': tool_solde_client,
    'nombre_clients': tool_nombre_clients,
    'liste_clients': tool_liste_clients,
    'nombre_groupes_clients': tool_nombre_groupes_clients,
    'liste_groupes_clients': tool_liste_groupes_clients,
    'solde_groupe_client': tool_solde_groupe_client,
    'clients_par_groupe': tool_clients_par_groupe,
    'clients_sans_groupe': tool_clients_sans_groupe,
    'top_clients_solde': tool_top_clients_solde,
    'solde_total_clients_et_groupes': tool_solde_total_clients_et_groupes,
    'liste_commandes': tool_liste_commandes,
    'detail_commande': tool_detail_commande,
    'commandes_produit': tool_commandes_produit,
    'stats_commandes': tool_stats_commandes,
    'generer_bon_commande_pdf': tool_generer_bon_commande_pdf,
    'resume_finance': tool_resume_finance,
    'operations_financieres': tool_operations_financieres,
    'generer_export_operations_financieres_pdf': tool_generer_export_operations_financieres_pdf,
    'generer_export_operations_financieres_excel': tool_generer_export_operations_financieres_excel,
    'liste_inventaires': tool_liste_inventaires,
    'detail_inventaire': tool_detail_inventaire,
    'liste_declarations_impots': tool_liste_declarations_impots,
    'detail_declaration_impot': tool_detail_declaration_impot,
    'taxes_a_declarer_periode': tool_taxes_a_declarer_periode,
    'nombre_employes': tool_nombre_employes,
    'liste_employes': tool_liste_employes,
    'employes_par_poste': tool_employes_par_poste,
    'postes_disponibles': tool_postes_disponibles,
    'acces_module': tool_acces_module,
    'modules_disponibles': tool_modules_disponibles,
    'mes_modules_accessibles': tool_mes_modules_accessibles,
    'generer_rapport_pdf': tool_generer_rapport_pdf,
    'generer_rapport_excel': tool_generer_rapport_excel,
}


def call_ai_tool(name, arguments, user=None):
    fn = TOOL_FUNCTIONS.get(name)
    if not fn:
        return {
            'error': (
                f"Outil '{name}' inexistant. N'invente jamais un nom d'outil hors de la liste fournie : "
                "reponds directement en texte si aucun outil disponible ne correspond a la question."
            )
        }
    try:
        return fn(user, **(arguments or {}))
    except TypeError as exc:
        return {'error': f"Parametres invalides pour {name}: {exc}"}
    except ValueError as exc:
        return {'error': str(exc)}
    except Exception as exc:  # ne jamais faire planter la conversation
        return {'error': f"Erreur interne lors de l'appel a {name}: {exc}"}


AI_TOOLS = [
    {
        'type': 'function',
        'function': {
            'name': 'chiffre_affaires_periode',
            'description': (
                "Donne le nombre de ventes, le chiffre d'affaires (HT et TTC) et le panier moyen "
                "sur une periode donnee (aujourd'hui, hier, cette semaine, ce mois, une periode "
                "personnalisee, etc.)."
            ),
            'parameters': {
                'type': 'object',
                'properties': {
                    'periode': {'type': 'string', 'enum': PERIODES_VALIDES, 'description': _PERIODE_DESC},
                    'date_debut': {'type': 'string', 'description': "Date de debut AAAA-MM-JJ, uniquement si periode='personnalise'."},
                    'date_fin': {'type': 'string', 'description': "Date de fin AAAA-MM-JJ, uniquement si periode='personnalise'."},
                },
                'required': ['periode'],
            },
        },
    },
    {
        'type': 'function',
        'function': {
            'name': 'comparer_ca_mois_precedent',
            'description': (
                "Compare le chiffre d'affaires du mois en cours avec celui du mois precedent et "
                "calcule la variation en montant et en pourcentage. A utiliser pour toute question "
                "du type 'comment se compare le CA / les ventes par rapport au mois dernier'."
            ),
            'parameters': {'type': 'object', 'properties': {}},
        },
    },
    {
        'type': 'function',
        'function': {
            'name': 'prevision_chiffre_affaires',
            'description': (
                "Calcule une prevision du chiffre d'affaires futur par regression lineaire sur les "
                "30 derniers jours de ventes quotidiennes, et projette la tendance sur un horizon "
                "donne (en jours, defaut 7). A utiliser pour toute question sur les previsions, "
                "tendances ou projections de ventes futures."
            ),
            'parameters': {
                'type': 'object',
                'properties': {
                    'horizon_jours': {'type': 'integer', 'description': "Nombre de jours a projeter dans le futur (defaut 7, max 60)."},
                },
            },
        },
    },
    {
        'type': 'function',
        'function': {
            'name': 'employe_du_mois',
            'description': (
                "Classe les employes par chiffre d'affaires genere sur le mois en cours ou le mois "
                "precedent, pour repondre a 'qui est l'employe du mois' ou identifier les meilleurs "
                "vendeurs."
            ),
            'parameters': {
                'type': 'object',
                'properties': {
                    'periode': {'type': 'string', 'enum': ['ce_mois', 'mois_dernier'], 'description': "ce_mois (defaut) ou mois_dernier."},
                },
            },
        },
    },
    {
        'type': 'function',
        'function': {
            'name': 'nombre_produits',
            'description': "Donne le nombre total de produits references dans le catalogue de la pharmacie.",
            'parameters': {'type': 'object', 'properties': {}},
        },
    },
    {
        'type': 'function',
        'function': {
            'name': 'liste_produits',
            'description': (
                "Liste les produits du CATALOGUE (fiche produit : nom, code, fournisseur, rayon, "
                "famille, conditionnement), avec une recherche optionnelle. A utiliser pour 'quels sont "
                "nos produits', 'liste des produits', pour retrouver le nom d'un ou plusieurs produits "
                "mentionnes precedemment (par exemple juste apres avoir demande le nombre de produits), "
                "ou pour une question sur le 'conditionnement' d'un produit — le conditionnement est la "
                "STRUCTURE D'EMBALLAGE (combien de niveaux : unite seule / unite+sous-unite / "
                "unite+sous-unite+sous-sous-unite), une notion totalement differente de la quantite "
                "actuellement en stock. Ne contient PAS de quantite en stock physique — pour ca, utilise "
                "stock_produit ou produits_stock_faible."
            ),
            'parameters': {
                'type': 'object',
                'properties': {
                    'recherche': {'type': 'string', 'description': "Filtre optionnel sur nom ou code produit (recherche partielle, texte uniquement — pas de prix/quantite). Laisser vide pour lister tout le catalogue et chercher toi-meme dans les resultats si le critere est une valeur numerique."},
                    'limite': {'type': 'integer', 'description': "Nombre maximum de produits a retourner (defaut 50, max 200)."},
                },
            },
        },
    },
    {
        'type': 'function',
        'function': {
            'name': 'liste_fournisseurs',
            'description': (
                "Liste les fournisseurs avec leur coefficient et leur taux de TVA effectifs (leur "
                "propre valeur si personnalisee, sinon celle heritee de leur groupe fournisseur), "
                "ainsi que le groupe fournisseur auquel ils appartiennent. A utiliser pour toute "
                "question sur le coefficient ou la TVA 'par fournisseur'."
            ),
            'parameters': {
                'type': 'object',
                'properties': {
                    'recherche': {'type': 'string', 'description': "Filtre optionnel sur nom ou prefixe du fournisseur (recherche partielle, texte uniquement). Ne filtre PAS sur le coefficient ou la TVA (des valeurs numeriques) : pour trouver un fournisseur par son coefficient/TVA, laisse ce parametre vide pour tout recuperer, puis identifie toi-meme le bon fournisseur dans les resultats retournes."},
                    'limite': {'type': 'integer', 'description': "Nombre maximum de fournisseurs a retourner (defaut 50, max 200)."},
                },
            },
        },
    },
    {
        'type': 'function',
        'function': {
            'name': 'liste_groupes_fournisseurs',
            'description': (
                "Liste les groupes fournisseurs avec leur coefficient par defaut et leur taux de TVA "
                "par defaut (valeurs heritees par les fournisseurs du groupe qui n'ont pas de valeur "
                "personnalisee), et le nombre de fournisseurs dans chaque groupe. A utiliser pour toute "
                "question sur le coefficient ou la TVA 'par groupe fournisseur'."
            ),
            'parameters': {
                'type': 'object',
                'properties': {
                    'recherche': {'type': 'string', 'description': "Filtre optionnel sur le nom du groupe (recherche partielle, texte uniquement). Ne filtre PAS sur le coefficient ou la TVA : pour trouver un groupe par ces valeurs, laisse ce parametre vide et identifie toi-meme le bon groupe dans les resultats."},
                },
            },
        },
    },
    {
        'type': 'function',
        'function': {
            'name': 'stock_produit',
            'description': (
                "Cherche un ou plusieurs produits par nom ou code produit (recherche partielle) et "
                "donne leur quantite en stock actuelle (tous lots confondus, avec le conditionnement du "
                "produit pour interpreter correctement les quantites unites/sous-unites/sous-sous-unites) "
                "ainsi que la prochaine date de peremption. A utiliser pour toute question 'quelle est la "
                "quantite en stock de tel produit' — mais PAS pour une simple question sur le "
                "conditionnement seul (utilise plutot liste_produits dans ce cas)."
            ),
            'parameters': {
                'type': 'object',
                'properties': {
                    'nom_produit': {'type': 'string', 'description': "Nom ou code du produit recherche (recherche partielle, insensible a la casse)."},
                },
                'required': ['nom_produit'],
            },
        },
    },
    {
        'type': 'function',
        'function': {
            'name': 'produits_stock_faible',
            'description': (
                "Liste les produits dont le stock total est faible ou nul (rupture de stock), en "
                "dessous d'un seuil donne (defaut 10 unites). Utile pour detecter les produits a "
                "reapprovisionner d'urgence."
            ),
            'parameters': {
                'type': 'object',
                'properties': {
                    'seuil': {'type': 'integer', 'description': "Seuil en dessous duquel un produit est considere en stock faible (defaut 10). Utiliser 0 pour ne lister que les ruptures totales."},
                    'limite': {'type': 'integer', 'description': "Nombre maximum de produits a retourner (defaut 15)."},
                },
            },
        },
    },
    {
        'type': 'function',
        'function': {
            'name': 'produits_peremption_proche',
            'description': (
                "Liste les lots de produits dont la date de peremption arrive dans les N prochains "
                "jours (defaut 30), tries par date de peremption la plus proche."
            ),
            'parameters': {
                'type': 'object',
                'properties': {
                    'nb_jours': {'type': 'integer', 'description': "Horizon en jours (defaut 30)."},
                    'limite': {'type': 'integer', 'description': "Nombre maximum de lots a retourner (defaut 15)."},
                },
            },
        },
    },
    {
        'type': 'function',
        'function': {
            'name': 'top_produits_vendus',
            'description': "Classe les produits les plus vendus sur une periode, par chiffre d'affaires ou par quantite.",
            'parameters': {
                'type': 'object',
                'properties': {
                    'periode': {'type': 'string', 'enum': PERIODES_VALIDES, 'description': _PERIODE_DESC},
                    'date_debut': {'type': 'string', 'description': "AAAA-MM-JJ, si periode='personnalise'."},
                    'date_fin': {'type': 'string', 'description': "AAAA-MM-JJ, si periode='personnalise'."},
                    'critere': {'type': 'string', 'enum': ['chiffre_affaires', 'quantite'], 'description': "Trier par chiffre_affaires (defaut) ou par quantite vendue."},
                    'limite': {'type': 'integer', 'description': "Nombre de produits a retourner (defaut 5)."},
                },
                'required': ['periode'],
            },
        },
    },
    {
        'type': 'function',
        'function': {
            'name': 'sorties_stock_periode',
            'description': (
                "Donne un resume des sorties de stock (ventes, pertes, casse, corrections, transferts...) "
                "sur une periode : nombre de sorties, quantite totale, valeur HT/TTC, et repartition par "
                "raison. A utiliser pour toute question du type 'sorties de stock du jour/de la semaine/du mois'."
            ),
            'parameters': {
                'type': 'object',
                'properties': {
                    'periode': {'type': 'string', 'enum': PERIODES_VALIDES, 'description': _PERIODE_DESC},
                    'date_debut': {'type': 'string', 'description': "AAAA-MM-JJ, si periode='personnalise'."},
                    'date_fin': {'type': 'string', 'description': "AAAA-MM-JJ, si periode='personnalise'."},
                },
                'required': ['periode'],
            },
        },
    },
    {
        'type': 'function',
        'function': {
            'name': 'dernieres_sorties_stock',
            'description': (
                "Liste le detail des sorties de stock sur une periode (produit, quantite, raison, qui l'a "
                "effectuee, valeur, date/heure), les plus recentes en premier. A utiliser quand on veut le "
                "detail ligne par ligne plutot qu'un simple total (ex: 'quelles sont les sorties de stock "
                "d'aujourd'hui', 'montre-moi les dernieres sorties')."
            ),
            'parameters': {
                'type': 'object',
                'properties': {
                    'periode': {'type': 'string', 'enum': PERIODES_VALIDES, 'description': _PERIODE_DESC},
                    'date_debut': {'type': 'string', 'description': "AAAA-MM-JJ, si periode='personnalise'."},
                    'date_fin': {'type': 'string', 'description': "AAAA-MM-JJ, si periode='personnalise'."},
                    'limite': {'type': 'integer', 'description': "Nombre maximum de sorties a retourner (defaut 10, max 50)."},
                },
                'required': ['periode'],
            },
        },
    },
    {
        'type': 'function',
        'function': {
            'name': 'sorties_stock_produit',
            'description': "Donne le total des sorties de stock (quantite et valeur) pour un produit specifique sur une periode.",
            'parameters': {
                'type': 'object',
                'properties': {
                    'nom_produit': {'type': 'string', 'description': "Nom ou code du produit recherche (recherche partielle)."},
                    'periode': {'type': 'string', 'enum': PERIODES_VALIDES, 'description': _PERIODE_DESC},
                    'date_debut': {'type': 'string', 'description': "AAAA-MM-JJ, si periode='personnalise'."},
                    'date_fin': {'type': 'string', 'description': "AAAA-MM-JJ, si periode='personnalise'."},
                },
                'required': ['nom_produit', 'periode'],
            },
        },
    },
    {
        'type': 'function',
        'function': {
            'name': 'top_clients',
            'description': "Classe les clients ayant depense le plus sur une periode donnee.",
            'parameters': {
                'type': 'object',
                'properties': {
                    'periode': {'type': 'string', 'enum': PERIODES_VALIDES, 'description': _PERIODE_DESC},
                    'date_debut': {'type': 'string', 'description': "AAAA-MM-JJ, si periode='personnalise'."},
                    'date_fin': {'type': 'string', 'description': "AAAA-MM-JJ, si periode='personnalise'."},
                    'limite': {'type': 'integer', 'description': "Nombre de clients a retourner (defaut 5)."},
                },
                'required': ['periode'],
            },
        },
    },
    {
        'type': 'function',
        'function': {
            'name': 'solde_client',
            'description': "Cherche un client par nom, prenom, matricule ou email et donne son solde de compte actuel.",
            'parameters': {
                'type': 'object',
                'properties': {
                    'recherche': {'type': 'string', 'description': "Nom, prenom, matricule ou email (recherche partielle)."},
                },
                'required': ['recherche'],
            },
        },
    },
    {
        'type': 'function',
        'function': {
            'name': 'nombre_clients',
            'description': "Donne le nombre total de clients enregistres.",
            'parameters': {'type': 'object', 'properties': {}},
        },
    },
    {
        'type': 'function',
        'function': {
            'name': 'liste_clients',
            'description': (
                "Liste les clients (nom, matricule, groupe, solde), avec une recherche optionnelle. "
                "A utiliser pour 'quels sont nos clients', 'liste des clients', ou pour retrouver le "
                "nom d'un ou plusieurs clients mentionnes precedemment dans la conversation (par "
                "exemple juste apres avoir demande le nombre de clients)."
            ),
            'parameters': {
                'type': 'object',
                'properties': {
                    'recherche': {'type': 'string', 'description': "Filtre optionnel sur nom, prenom, matricule ou email (recherche partielle, texte uniquement). Ne filtre PAS sur le solde (numerique) : pour trouver un client par son solde, laisse ce parametre vide et identifie toi-meme le bon client dans les resultats (ou utilise top_clients_solde)."},
                    'limite': {'type': 'integer', 'description': "Nombre maximum de clients a retourner (defaut 50, max 200)."},
                },
            },
        },
    },
    {
        'type': 'function',
        'function': {
            'name': 'nombre_groupes_clients',
            'description': "Donne le nombre total de groupes clients enregistres (ex: mutuelles, entreprises partenaires...).",
            'parameters': {'type': 'object', 'properties': {}},
        },
    },
    {
        'type': 'function',
        'function': {
            'name': 'liste_groupes_clients',
            'description': (
                "Liste tous les groupes clients avec leur solde, leur pourcentage d'absorption et le "
                "nombre de clients qui en font partie."
            ),
            'parameters': {'type': 'object', 'properties': {}},
        },
    },
    {
        'type': 'function',
        'function': {
            'name': 'solde_groupe_client',
            'description': "Cherche un groupe client par nom (recherche partielle) et donne son solde et son pourcentage d'absorption.",
            'parameters': {
                'type': 'object',
                'properties': {
                    'recherche': {'type': 'string', 'description': "Nom du groupe client recherche (recherche partielle)."},
                },
                'required': ['recherche'],
            },
        },
    },
    {
        'type': 'function',
        'function': {
            'name': 'clients_par_groupe',
            'description': "Liste les clients membres d'un groupe client donne, avec leur solde individuel.",
            'parameters': {
                'type': 'object',
                'properties': {
                    'nom_groupe': {'type': 'string', 'description': "Nom du groupe client (recherche partielle)."},
                },
                'required': ['nom_groupe'],
            },
        },
    },
    {
        'type': 'function',
        'function': {
            'name': 'clients_sans_groupe',
            'description': "Donne le nombre de clients qui ne sont rattaches a aucun groupe, et un apercu de leur liste.",
            'parameters': {
                'type': 'object',
                'properties': {
                    'limite': {'type': 'integer', 'description': "Nombre de clients a lister dans l'apercu (defaut 15)."},
                },
            },
        },
    },
    {
        'type': 'function',
        'function': {
            'name': 'top_clients_solde',
            'description': (
                "Classe les clients ayant le solde de compte (credit) le plus eleve, tous mouvements "
                "confondus (pas seulement leurs achats sur une periode)."
            ),
            'parameters': {
                'type': 'object',
                'properties': {
                    'limite': {'type': 'integer', 'description': "Nombre de clients a retourner (defaut 10)."},
                },
            },
        },
    },
    {
        'type': 'function',
        'function': {
            'name': 'solde_total_clients_et_groupes',
            'description': "Donne la somme totale des soldes de tous les clients et de tous les groupes clients.",
            'parameters': {'type': 'object', 'properties': {}},
        },
    },
    {
        'type': 'function',
        'function': {
            'name': 'liste_commandes',
            'description': (
                "Liste les commandes FOURNISSEURS (reapprovisionnement de la pharmacie, numeros "
                "CMD-...) : numero, fournisseur, statut (en_cours = passee mais pas encore "
                "receptionnee, livree, annulee), qui l'a creee et quand, quantites et montant HT "
                "commandes, et pour les livrees les ecarts de livraison (manquants). Une commande "
                "marquee 'est_une_relance_de' est une re-commande des produits manquants d'une "
                "commande d'origine. A utiliser pour 'quelles commandes sont en cours', 'liste des "
                "commandes', 'les commandes de tel fournisseur', 'a-t-on recu nos commandes'. "
                "Ne PAS confondre avec les ventes aux clients."
            ),
            'parameters': {
                'type': 'object',
                'properties': {
                    'statut': {'type': 'string', 'enum': list(_STATUTS_COMMANDE), 'description': "Filtre optionnel sur le statut. Laisser vide pour toutes les commandes."},
                    'fournisseur': {'type': 'string', 'description': "Filtre optionnel sur le nom du fournisseur (recherche partielle, texte uniquement)."},
                    'periode': {'type': 'string', 'enum': PERIODES_VALIDES, 'description': "Filtre optionnel : " + _PERIODE_DESC + " Laisser vide pour toutes les periodes."},
                    'date_debut': {'type': 'string', 'description': "AAAA-MM-JJ, si periode='personnalise'."},
                    'date_fin': {'type': 'string', 'description': "AAAA-MM-JJ, si periode='personnalise'."},
                    'limite': {'type': 'integer', 'description': "Nombre maximum de commandes a retourner (defaut 10, max 50)."},
                },
            },
        },
    },
    {
        'type': 'function',
        'function': {
            'name': 'detail_commande',
            'description': (
                "Donne le detail complet d'une commande fournisseur : chaque ligne produit avec la "
                "quantite commandee, la quantite livree et l'ecart (negatif = manquant a la "
                "livraison), les montants HT, le stock au moment de la commande, ainsi que les "
                "eventuelles relances creees a partir de cette commande. Sans parametre 'recherche', "
                "renvoie la commande la plus recente (donc 'la derniere commande'). A utiliser pour "
                "'que contient la commande CMD-..., 'quels produits manquaient a la livraison', "
                "'detail de ma derniere commande'."
            ),
            'parameters': {
                'type': 'object',
                'properties': {
                    'recherche': {'type': 'string', 'description': "Numero de commande (ex: CMD-20260717-001, recherche partielle) OU nom du fournisseur (renvoie alors sa commande la plus recente). Laisser vide pour la derniere commande toutes confondues."},
                },
            },
        },
    },
    {
        'type': 'function',
        'function': {
            'name': 'commandes_produit',
            'description': (
                "Historique des commandes fournisseurs contenant un produit donne (recherche par nom "
                "ou code produit), de la plus recente a la plus ancienne : numero, date, fournisseur, "
                "statut, quantite commandee/livree et ecart. A utiliser pour 'quand a-t-on commande "
                "tel produit pour la derniere fois', 'combien de tel produit a-t-on commande', 'ce "
                "produit est-il deja en commande'."
            ),
            'parameters': {
                'type': 'object',
                'properties': {
                    'nom_produit': {'type': 'string', 'description': "Nom ou code du produit recherche (recherche partielle, insensible a la casse)."},
                    'limite': {'type': 'integer', 'description': "Nombre maximum de lignes a retourner (defaut 10, max 50)."},
                },
                'required': ['nom_produit'],
            },
        },
    },
    {
        'type': 'function',
        'function': {
            'name': 'stats_commandes',
            'description': (
                "Resume des commandes fournisseurs sur une periode : nombre de commandes par statut "
                "(en_cours/livree/annulee), montant total HT commande et livre, nombre de commandes "
                "livrees avec ecart, total d'unites manquantes a la livraison, et top fournisseurs "
                "par montant commande. A utiliser pour 'combien avons-nous commande ce mois', "
                "'montant des commandes en cours', 'quels fournisseurs commande-t-on le plus'."
            ),
            'parameters': {
                'type': 'object',
                'properties': {
                    'periode': {'type': 'string', 'enum': PERIODES_VALIDES, 'description': _PERIODE_DESC},
                    'date_debut': {'type': 'string', 'description': "AAAA-MM-JJ, si periode='personnalise'."},
                    'date_fin': {'type': 'string', 'description': "AAAA-MM-JJ, si periode='personnalise'."},
                },
                'required': ['periode'],
            },
        },
    },
    {
        'type': 'function',
        'function': {
            'name': 'generer_bon_commande_pdf',
            'description': (
                "Genere le bon de commande PDF OFFICIEL d'une commande fournisseur, strictement "
                "identique a celui que produit le module Commandes de l'application (memes styles, "
                "meme mise en page, tableau des produits sur toute la largeur). A utiliser DES QU'ON "
                "demande 'le bon de commande', 'le PDF de la commande CMD-...', 'imprime ma derniere "
                "commande' — n'utilise JAMAIS generer_rapport_pdf pour un bon de commande. Sans "
                "parametre 'recherche', prend la commande la plus recente. Apres l'appel, un bouton de "
                "telechargement s'affiche automatiquement sous ton message : ne redonne pas l'URL."
            ),
            'parameters': {
                'type': 'object',
                'properties': {
                    'recherche': {'type': 'string', 'description': "Numero de commande (ex: CMD-20260717-001, recherche partielle) OU nom du fournisseur (prend alors sa commande la plus recente). Laisser vide pour la derniere commande toutes confondues."},
                },
            },
        },
    },
    {
        'type': 'function',
        'function': {
            'name': 'resume_finance',
            'description': (
                "Donne, pour une periode donnee, le chiffre d'affaires (HT/TTC) et le benefice, ainsi "
                "que le solde de tresorerie ACTUEL (toujours global, jamais limite a la periode : c'est "
                "un cumul du benefice depuis toujours, ajuste par les encaissements/decaissements "
                "manuels). A utiliser pour 'quel est le CA/benefice de ce mois', 'quel est le solde "
                "actuel/de caisse', 'combien a-t-on de tresorerie'. Le module Finance est distinct du "
                "solde des CLIENTS (credit compte client) : ne pas confondre avec solde_client."
            ),
            'parameters': {
                'type': 'object',
                'properties': {
                    'periode': {'type': 'string', 'enum': PERIODES_VALIDES, 'description': _PERIODE_DESC},
                    'date_debut': {'type': 'string', 'description': "AAAA-MM-JJ, si periode='personnalise'."},
                    'date_fin': {'type': 'string', 'description': "AAAA-MM-JJ, si periode='personnalise'."},
                },
                'required': ['periode'],
            },
        },
    },
    {
        'type': 'function',
        'function': {
            'name': 'operations_financieres',
            'description': (
                "Liste les operations financieres manuelles (encaissements et decaissements de caisse) "
                "avec leur raison, du module Finance : montant, raison, note, qui l'a enregistree, "
                "date. A utiliser pour 'quels encaissements/decaissements ont ete faits', 'pourquoi le "
                "solde a change', 'historique des mouvements de caisse'. Sans periode, renvoie tout "
                "l'historique."
            ),
            'parameters': {
                'type': 'object',
                'properties': {
                    'type_operation': {'type': 'string', 'enum': ['encaissement', 'decaissement'], 'description': "Filtre optionnel sur le type d'operation. Laisser vide pour les deux."},
                    'periode': {'type': 'string', 'enum': PERIODES_VALIDES, 'description': "Filtre optionnel : " + _PERIODE_DESC + " Laisser vide pour tout l'historique."},
                    'date_debut': {'type': 'string', 'description': "AAAA-MM-JJ, si periode='personnalise'."},
                    'date_fin': {'type': 'string', 'description': "AAAA-MM-JJ, si periode='personnalise'."},
                    'limite': {'type': 'integer', 'description': "Nombre maximum d'operations a retourner (defaut 20, max 100)."},
                },
            },
        },
    },
    {
        'type': 'function',
        'function': {
            'name': 'generer_export_operations_financieres_pdf',
            'description': (
                "Genere l'export PDF OFFICIEL des operations financieres (encaissements/decaissements), "
                "strictement identique a celui du module Finance de l'application (memes styles, meme "
                "recapitulatif, tableau pleine largeur). A utiliser DES QU'ON demande 'exporte les "
                "operations financieres en PDF', 'le PDF des encaissements/decaissements' — pour un "
                "Excel, utilise generer_export_operations_financieres_excel a la place. Sans periode, "
                "exporte tout l'historique. Apres l'appel, un bouton de telechargement s'affiche "
                "automatiquement sous ton message : ne redonne pas l'URL."
            ),
            'parameters': {
                'type': 'object',
                'properties': {
                    'periode': {'type': 'string', 'enum': PERIODES_VALIDES, 'description': "Periode a exporter (optionnelle). " + _PERIODE_DESC + " Laisser vide pour tout l'historique."},
                    'date_debut': {'type': 'string', 'description': "AAAA-MM-JJ, si periode='personnalise'."},
                    'date_fin': {'type': 'string', 'description': "AAAA-MM-JJ, si periode='personnalise'."},
                },
            },
        },
    },
    {
        'type': 'function',
        'function': {
            'name': 'generer_export_operations_financieres_excel',
            'description': (
                "Genere l'export Excel (.xlsx) OFFICIEL des operations financieres (encaissements/"
                "decaissements), strictement identique a celui du module Finance de l'application. A "
                "utiliser DES QU'ON demande un fichier Excel/tableur des operations financieres — pour "
                "un PDF, utilise generer_export_operations_financieres_pdf a la place. Sans periode, "
                "exporte tout l'historique."
            ),
            'parameters': {
                'type': 'object',
                'properties': {
                    'periode': {'type': 'string', 'enum': PERIODES_VALIDES, 'description': "Periode a exporter (optionnelle). " + _PERIODE_DESC + " Laisser vide pour tout l'historique."},
                    'date_debut': {'type': 'string', 'description': "AAAA-MM-JJ, si periode='personnalise'."},
                    'date_fin': {'type': 'string', 'description': "AAAA-MM-JJ, si periode='personnalise'."},
                },
            },
        },
    },
    {
        'type': 'function',
        'function': {
            'name': 'liste_inventaires',
            'description': (
                "Liste l'historique des inventaires (titre, statut en_cours/valide/annule, date et "
                "auteur de creation, date et auteur de validation, nombre de produits et d'ecarts), du "
                "plus recent au plus ancien. A utiliser pour 'quand est le dernier inventaire', 'liste "
                "des inventaires', 'historique des inventaires'. IMPORTANT : ReflexPharma ne planifie "
                "aucun inventaire dans le futur — il n'existe que des inventaires deja crees (au plus un "
                "'en_cours' a la fois) et l'historique des inventaires passes. Si on te demande les "
                "inventaires 'a venir/futurs', explique qu'il n'y a pas de planification, et precise s'il "
                "y a un inventaire actuellement en_cours (donc pas encore termine)."
            ),
            'parameters': {
                'type': 'object',
                'properties': {
                    'limite': {'type': 'integer', 'description': "Nombre maximum d'inventaires a retourner (defaut 10, max 50)."},
                },
            },
        },
    },
    {
        'type': 'function',
        'function': {
            'name': 'detail_inventaire',
            'description': (
                "Donne le detail d'un inventaire : statut, dates, qui l'a cree/valide, nombre de produits "
                "comptes, et la liste des ecarts constates (produit, quantite avant/apres, qui a compte). "
                "Sans parametre 'recherche', renvoie l'inventaire en cours s'il y en a un, sinon le plus "
                "recent (donc 'le dernier inventaire'). Pour generer un PDF de ce rapport (comme l'export "
                "manuel), appelle d'abord cet outil puis generer_rapport_pdf avec les donnees obtenues."
            ),
            'parameters': {
                'type': 'object',
                'properties': {
                    'recherche': {'type': 'string', 'description': "Titre de l'inventaire recherche (recherche partielle, texte uniquement). Laisser vide pour l'inventaire en cours / le plus recent."},
                },
            },
        },
    },
    {
        'type': 'function',
        'function': {
            'name': 'liste_declarations_impots',
            'description': (
                "Liste les periodes de declaration des taxes du module Impots (reference IMP-xxxxxx, "
                "periode debut/fin, statut declaree ou en_preparation, totaux HT/TVA/benefice/TTC, qui a "
                "declare et quand), de la plus recente a la plus ancienne. A utiliser pour 'liste des "
                "declarations d'impots', 'quelles periodes ont ete declarees', 'quand a-t-on declare "
                "les taxes pour la derniere fois', 'y a-t-il une declaration en preparation'. Les totaux "
                "d'une periode declaree sont geles au moment de la declaration ; ceux d'une periode en "
                "preparation sont recalcules en direct."
            ),
            'parameters': {
                'type': 'object',
                'properties': {
                    'statut': {'type': 'string', 'enum': ['declaree', 'en_preparation'], 'description': "Filtre optionnel sur le statut. Laisser vide pour toutes les declarations."},
                    'limite': {'type': 'integer', 'description': "Nombre maximum de declarations a retourner (defaut 10, max 50)."},
                },
            },
        },
    },
    {
        'type': 'function',
        'function': {
            'name': 'detail_declaration_impot',
            'description': (
                "Donne le detail complet d'une declaration d'impots : periode, statut, totaux HT/TVA/"
                "benefice/TTC, repartition de la TVA par taux (l'information cle pour remplir la "
                "declaration fiscale), auteur, date de declaration, et la liste des ventes incluses. "
                "Sans parametre 'recherche', renvoie la declaration la plus recente. Pour generer un PDF "
                "de ce rapport, appelle d'abord cet outil puis generer_rapport_pdf avec les donnees "
                "obtenues (l'export PDF officiel reste disponible dans le module Impots de l'application)."
            ),
            'parameters': {
                'type': 'object',
                'properties': {
                    'recherche': {'type': 'string', 'description': "Reference de la declaration (ex: IMP-123456, recherche partielle) OU une date AAAA-MM-JJ contenue dans la periode recherchee. Laisser vide pour la declaration la plus recente."},
                },
            },
        },
    },
    {
        'type': 'function',
        'function': {
            'name': 'taxes_a_declarer_periode',
            'description': (
                "Calcule les montants fiscaux d'une periode quelconque (meme sans declaration creee) : "
                "total HT, TVA effective, benefice, TTC et repartition de la TVA par taux, sur les ventes "
                "validees. Indique aussi si cette periode est deja couverte par une declaration marquee "
                "declaree, et liste les declarations qui la chevauchent. A utiliser pour 'combien de TVA "
                "dois-je declarer pour juin', 'est-ce que les impots de ce trimestre ont ete declares', "
                "'montant des taxes a declarer ce mois-ci'. Pour creer ou marquer une declaration, "
                "l'utilisateur doit passer par le module Impots de l'application (toi tu ne peux que "
                "consulter)."
            ),
            'parameters': {
                'type': 'object',
                'properties': {
                    'periode': {'type': 'string', 'enum': PERIODES_VALIDES, 'description': _PERIODE_DESC},
                    'date_debut': {'type': 'string', 'description': "Date de debut AAAA-MM-JJ, uniquement si periode='personnalise'."},
                    'date_fin': {'type': 'string', 'description': "Date de fin AAAA-MM-JJ, uniquement si periode='personnalise'."},
                },
                'required': ['periode'],
            },
        },
    },
    {
        'type': 'function',
        'function': {
            'name': 'nombre_employes',
            'description': (
                "Donne le nombre total d'employes (actifs/inactifs) et leur repartition par role "
                "(superadmin, admin, employee). A utiliser pour 'combien d'employes avons-nous'."
            ),
            'parameters': {'type': 'object', 'properties': {}},
        },
    },
    {
        'type': 'function',
        'function': {
            'name': 'liste_employes',
            'description': (
                "Liste les employes (nom, role, poste, anciennete en annees). A utiliser pour "
                "'qui travaille ici', 'liste des employes', 'qui a le plus d'anciennete', etc."
            ),
            'parameters': {
                'type': 'object',
                'properties': {
                    'statut': {'type': 'string', 'enum': ['actifs', 'inactifs', 'tous'], 'description': "Filtrer par statut du compte (defaut actifs)."},
                },
            },
        },
    },
    {
        'type': 'function',
        'function': {
            'name': 'employes_par_poste',
            'description': "Donne le nombre d'employes actifs par poste/metier (ex: pharmacien, caissier...).",
            'parameters': {'type': 'object', 'properties': {}},
        },
    },
    {
        'type': 'function',
        'function': {
            'name': 'postes_disponibles',
            'description': "Liste les postes/metiers configures dans la pharmacie (nom et description).",
            'parameters': {'type': 'object', 'properties': {}},
        },
    },
    {
        'type': 'function',
        'function': {
            'name': 'acces_module',
            'description': (
                "Donne la liste des employes ayant (ou n'ayant pas) acces a un module/fonctionnalite "
                "donne du logiciel. A utiliser pour toute question du type 'qui a acces a tel module', "
                "'qui peut gerer les ventes/le stock/les clients', etc. Si tu ne connais pas la cle exacte "
                "du module, appelle d'abord l'outil modules_disponibles."
            ),
            'parameters': {
                'type': 'object',
                'properties': {
                    'module': {'type': 'string', 'description': _MODULE_DESC},
                },
                'required': ['module'],
            },
        },
    },
    {
        'type': 'function',
        'function': {
            'name': 'modules_disponibles',
            'description': (
                "Liste TOUTES les cles de modules/fonctionnalites existantes dans le logiciel (utile "
                "avant d'appeler acces_module si le module demande n'est pas evident). N'utilise JAMAIS "
                "cet outil pour repondre a 'a quels modules ai-je acces' — cela listerait tout le "
                "catalogue au lieu des acces reels de la personne : utilise mes_modules_accessibles "
                "pour cette question-la."
            ),
            'parameters': {'type': 'object', 'properties': {}},
        },
    },
    {
        'type': 'function',
        'function': {
            'name': 'mes_modules_accessibles',
            'description': (
                "Donne la liste des modules auxquels LA PERSONNE QUI POSE LA QUESTION a reellement "
                "acces dans ReflexPharma — PAS la liste de tous les modules existants dans le logiciel. "
                "A utiliser pour toute question du type 'a quels modules ai-je acces', 'qu'est-ce que je "
                "peux faire/voir ici', 'quelles sont mes permissions'. Ne liste que ce que cette personne "
                "a le droit de faire, jamais le catalogue complet des modules possibles."
            ),
            'parameters': {'type': 'object', 'properties': {}},
        },
    },
    {
        'type': 'function',
        'function': {
            'name': 'generer_rapport_pdf',
            'description': (
                "Genere un rapport PDF telechargeable a partir de donnees deja obtenues via d'autres "
                "outils, et le rend disponible en telechargement dans le chat. N'utilise cet outil QUE "
                "si l'utilisateur demande explicitement un PDF, un rapport, un document ou une "
                "impression des resultats — jamais spontanement. Si l'utilisateur demande un fichier "
                "Excel, un tableur ou un .xlsx, utilise generer_rapport_excel a la place (memes "
                "parametres). Pour le bon de commande d'une commande fournisseur, utilise "
                "generer_bon_commande_pdf (document officiel de l'application), PAS cet outil. Si le "
                "format n'est pas precise pour un 'export' ou un 'rapport', choisis le PDF. Appelle "
                "d'abord le(s) outil(s) necessaires pour obtenir les donnees demandees, PUIS appelle "
                "cet outil avec un titre et le contenu structure (texte et/ou tableaux) a inclure dans "
                "le document."
            ),
            'parameters': {
                'type': 'object',
                'properties': {
                    'titre': {'type': 'string', 'description': "Titre du rapport (ex: 'Ventes du jour - 13/07/2026', 'Sorties de stock - Juillet 2026')."},
                    'sections': {
                        'type': 'array',
                        'description': "Sections composant le rapport, dans l'ordre d'affichage.",
                        'items': {
                            'type': 'object',
                            'properties': {
                                'titre_section': {'type': 'string', 'description': "Titre de la section."},
                                'texte': {'type': 'string', 'description': "Texte libre de la section (optionnel), une phrase par ligne."},
                                'tableau': {
                                    'type': 'object',
                                    'description': "Tableau optionnel pour cette section.",
                                    'properties': {
                                        'colonnes': {'type': 'array', 'items': {'type': 'string'}, 'description': "Noms des colonnes."},
                                        'lignes': {
                                            'type': 'array',
                                            'items': {'type': 'array', 'items': {'type': 'string'}},
                                            'description': "Lignes du tableau, chaque ligne etant une liste de valeurs (une par colonne, converties en texte).",
                                        },
                                    },
                                },
                            },
                            'required': ['titre_section'],
                        },
                    },
                },
                'required': ['titre', 'sections'],
            },
        },
    },
    {
        'type': 'function',
        'function': {
            'name': 'generer_rapport_excel',
            'description': (
                "Genere un classeur Excel (.xlsx) telechargeable a partir de donnees deja obtenues "
                "via d'autres outils, et le rend disponible en telechargement dans le chat. N'utilise "
                "cet outil QUE si l'utilisateur demande explicitement un fichier Excel, un tableur, un "
                ".xlsx ou un export 'pour travailler les donnees' — jamais spontanement. Pour un PDF, "
                "un document ou une impression, utilise generer_rapport_pdf a la place. Appelle "
                "d'abord le(s) outil(s) necessaires pour obtenir les donnees demandees, PUIS appelle "
                "cet outil avec un titre et le contenu structure : privilégie les tableaux (colonnes + "
                "lignes), c'est ce qui rend un fichier Excel utile."
            ),
            'parameters': {
                'type': 'object',
                'properties': {
                    'titre': {'type': 'string', 'description': "Titre du classeur (ex: 'Commandes en cours - 17/07/2026')."},
                    'sections': {
                        'type': 'array',
                        'description': "Sections composant le classeur, dans l'ordre d'affichage (rendues les unes sous les autres dans la feuille).",
                        'items': {
                            'type': 'object',
                            'properties': {
                                'titre_section': {'type': 'string', 'description': "Titre de la section."},
                                'texte': {'type': 'string', 'description': "Texte libre de la section (optionnel), une phrase par ligne."},
                                'tableau': {
                                    'type': 'object',
                                    'description': "Tableau optionnel pour cette section (fortement recommande pour un Excel).",
                                    'properties': {
                                        'colonnes': {'type': 'array', 'items': {'type': 'string'}, 'description': "Noms des colonnes."},
                                        'lignes': {
                                            'type': 'array',
                                            'items': {'type': 'array', 'items': {'type': 'string'}},
                                            'description': "Lignes du tableau, chaque ligne etant une liste de valeurs (une par colonne).",
                                        },
                                    },
                                },
                            },
                            'required': ['titre_section'],
                        },
                    },
                },
                'required': ['titre', 'sections'],
            },
        },
    },
]
