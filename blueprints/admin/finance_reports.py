"""Logique et exports du module Finance, partagés entre les vues (blueprints/admin/views.py)
et l'assistant IA (blueprints/admin/ai_tools.py) : les deux passent par les mêmes fonctions,
ce qui garantit des calculs et des documents strictement identiques.
"""

from datetime import datetime

from extensions import db
from models.vente import Vente, VenteLigne
from models.finance import OperationFinanciere
from utils.currencies import devise_active


def compute_benefice_total_all_time():
    """Bénéfice cumulé (marge coefficient, hors TVA effective) sur TOUTES les ventes
    validées depuis toujours, calculé en une requête SQL directement sur les lignes
    de vente — jamais stocké nulle part, jamais modifié par les opérations
    financières (encaissements/décaissements)."""
    tva_expr = VenteLigne.total_ht * (VenteLigne.tva_pourcentage / 100.0)
    benefice_expr = db.func.greatest(VenteLigne.total_ttc - VenteLigne.total_ht - tva_expr, 0.0)
    total = db.session.query(db.func.coalesce(db.func.sum(benefice_expr), 0.0)).join(
        Vente, Vente.numero_vente == VenteLigne.numero_vente
    ).filter(Vente.statut == 'validee').scalar()
    return float(total or 0)


def compute_totaux_operations(start_dt=None, end_dt=None):
    """Total des encaissements et décaissements, optionnellement borné dans le temps."""
    query = db.session.query(
        db.func.coalesce(db.func.sum(
            db.case((OperationFinanciere.type == 'encaissement', OperationFinanciere.montant), else_=0.0)
        ), 0.0),
        db.func.coalesce(db.func.sum(
            db.case((OperationFinanciere.type == 'decaissement', OperationFinanciere.montant), else_=0.0)
        ), 0.0),
    )
    if start_dt is not None:
        query = query.filter(OperationFinanciere.created_at >= start_dt)
    if end_dt is not None:
        query = query.filter(OperationFinanciere.created_at <= end_dt)
    encaissements, decaissements = query.one()
    return float(encaissements or 0), float(decaissements or 0)


def compute_solde_actuel():
    """Solde de trésorerie actuel = bénéfice cumulé depuis toujours + encaissements
    manuels - décaissements manuels (toutes périodes confondues). Ces opérations
    n'écrivent jamais dans Vente/VenteLigne : le bénéfice affiché partout ailleurs
    dans l'application (stats, impôts...) reste une donnée pure déduite des ventes,
    quel que soit l'état du solde."""
    encaissements, decaissements = compute_totaux_operations()
    return compute_benefice_total_all_time() + encaissements - decaissements


def query_operations_financieres(start_dt=None, end_dt=None, type_filtre=None):
    """Opérations financières filtrées par intervalle de dates et/ou type, les plus
    récentes en premier."""
    query = OperationFinanciere.query
    if start_dt is not None:
        query = query.filter(OperationFinanciere.created_at >= start_dt)
    if end_dt is not None:
        query = query.filter(OperationFinanciere.created_at <= end_dt)
    if type_filtre in ('encaissement', 'decaissement'):
        query = query.filter(OperationFinanciere.type == type_filtre)
    return query.order_by(OperationFinanciere.created_at.desc()).all()


def label_periode_dates(date_from=None, date_to=None):
    """Libellé de période à partir de deux dates (module Finance : filtres date_from/date_to)."""
    if date_from and date_to:
        return f"du {date_from.strftime('%d/%m/%Y')} au {date_to.strftime('%d/%m/%Y')}"
    if date_from:
        return f"depuis le {date_from.strftime('%d/%m/%Y')}"
    if date_to:
        return f"jusqu'au {date_to.strftime('%d/%m/%Y')}"
    return "toutes périodes"


def build_operations_financieres_pdf(target, operations, periode_label, tire_par, pharmacy_name, solde_actuel=None):
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

    doc = SimpleDocTemplate(target, pagesize=A4, topMargin=24, bottomMargin=24, leftMargin=24, rightMargin=24)
    styles = getSampleStyleSheet()
    styles.add(ParagraphStyle('Small', parent=styles['Normal'], fontSize=8, leading=10))
    # Le tableau occupe presque toute la largeur utile de la page A4
    largeur_utile = A4[0] - doc.leftMargin - doc.rightMargin

    devise = devise_active()
    total_encaissements = sum(o.montant or 0 for o in operations if o.type == 'encaissement')
    total_decaissements = sum(o.montant or 0 for o in operations if o.type == 'decaissement')

    elements = [
        Paragraph(f'Opérations financières - {pharmacy_name}', styles['Title']),
        Paragraph(
            f'Période : {periode_label} | Date du tirage : {datetime.now().strftime("%d/%m/%Y %H:%M")} | '
            f'Tiré par : {tire_par}',
            styles['Small']),
        Spacer(1, 8)
    ]

    summary = [
        ["Nombre d'opérations", str(len(operations)), 'Total encaissements', f'{total_encaissements:.2f} {devise}'],
        ['Total décaissements', f'{total_decaissements:.2f} {devise}', 'Solde actuel',
         f'{solde_actuel:.2f} {devise}' if solde_actuel is not None else '-'],
    ]
    elements.append(Table(summary, colWidths=[largeur_utile * p for p in (0.25, 0.25, 0.25, 0.25)], style=TableStyle([
        ('GRID', (0, 0), (-1, -1), 0.25, colors.HexColor('#D1D5DB')),
        ('BACKGROUND', (0, 0), (-1, -1), colors.HexColor('#F9FAFB')),
        ('FONTSIZE', (0, 0), (-1, -1), 8),
        ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
        ('FONTNAME', (2, 0), (2, -1), 'Helvetica-Bold'),
    ])))
    elements.append(Spacer(1, 12))

    cell_style = ParagraphStyle('CelluleFinance', parent=styles['Normal'], fontSize=8, leading=10)
    colWidths = [largeur_utile * p for p in (0.13, 0.12, 0.32, 0.11, 0.16, 0.16)]
    data = [['Date', 'Type', 'Raison', 'Montant', 'Enregistré par', 'Note']]
    for o in operations:
        data.append([
            o.created_at.strftime('%d/%m/%Y %H:%M') if o.created_at else '',
            'Encaissement' if o.type == 'encaissement' else 'Décaissement',
            Paragraph(o.raison or '', cell_style),
            f'{(o.montant or 0):.2f} {devise}',
            o.created_by_nom or '',
            Paragraph(o.note or '', cell_style),
        ])
    if not operations:
        data.append(['', '', 'Aucune opération sur cette période.', '', '', ''])

    elements.append(Table(data, colWidths=colWidths, repeatRows=1, style=TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#0d3b2e')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTSIZE', (0, 0), (-1, -1), 8),
        ('GRID', (0, 0), (-1, -1), 0.25, colors.HexColor('#D1D5DB')),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F0FAF6')]),
        ('ALIGN', (3, 0), (3, -1), 'RIGHT'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('LEFTPADDING', (0, 0), (-1, -1), 6),
        ('RIGHTPADDING', (0, 0), (-1, -1), 6),
        ('TOPPADDING', (0, 0), (-1, -1), 4),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
    ])))
    doc.build(elements)


def build_operations_financieres_excel(target, operations, periode_label, solde_actuel=None):
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = 'Opérations'

    devise = devise_active()
    total_encaissements = sum(o.montant or 0 for o in operations if o.type == 'encaissement')
    total_decaissements = sum(o.montant or 0 for o in operations if o.type == 'decaissement')

    ws['A1'] = 'Opérations financières'
    ws['A1'].font = Font(bold=True, size=14, color='0D3B2E')
    ws['A2'] = f'Période : {periode_label}'
    ws['A3'] = f'Généré le {datetime.now().strftime("%d/%m/%Y %H:%M")}'
    ws['A4'] = (f"Nombre d'opérations : {len(operations)} | Total encaissements : {total_encaissements:.2f} {devise} | "
                f"Total décaissements : {total_decaissements:.2f} {devise}")
    if solde_actuel is not None:
        ws['A5'] = f'Solde actuel : {solde_actuel:.2f} {devise}'

    header_row = 7
    columns = ['Date', 'Type', f'Montant ({devise})', 'Raison', 'Enregistré par', 'Note']
    header_fill = PatternFill(start_color='0D3B2E', end_color='0D3B2E', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF')
    for ci, col in enumerate(columns, start=1):
        cell = ws.cell(row=header_row, column=ci, value=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')

    row = header_row + 1
    for o in operations:
        ws.cell(row=row, column=1, value=o.created_at.strftime('%d/%m/%Y %H:%M') if o.created_at else '')
        ws.cell(row=row, column=2, value='Encaissement' if o.type == 'encaissement' else 'Décaissement')
        ws.cell(row=row, column=3, value=round(o.montant or 0, 2))
        ws.cell(row=row, column=4, value=o.raison or '')
        ws.cell(row=row, column=5, value=o.created_by_nom or '')
        ws.cell(row=row, column=6, value=o.note or '')
        row += 1

    ws.freeze_panes = f'A{header_row + 1}'
    for col_idx in range(1, len(columns) + 1):
        longueur = max(
            (len(str(ws.cell(row=r, column=col_idx).value or '')) for r in range(1, ws.max_row + 1)),
            default=0
        )
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max(12, longueur + 2), 45)

    wb.save(target)
