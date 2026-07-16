from flask import render_template, redirect, url_for, flash, request, abort, Response, current_app, send_file
from flask_login import login_required, current_user
import queue
import threading
import requests
import socket
import os
from . import admin
from models.user import User
from models.poste import Poste
from models.permission import Permission
from models.fournisseur import Fournisseur
from models.groupe_fournisseur import GroupeFournisseur
from models.rayon import Rayon
from models.famille import Famille
from models.section import Section
from models.produit import Produit
from models.stock import Stock
from models.stock_modification import StockModification
from models.stock_reason import StockReason
from models.stock_exit_log import StockExitLog
from models.groupe_client import GroupeClient
from models.client import Client
from models.client_modification_log import ClientModificationLog
from models.vente import Vente, VenteLigne
from models.setting import Setting
from models.inventaire import Inventaire, InventaireLigne
from models.declaration_impot import DeclarationImpot
from extensions import db
from functools import wraps
from datetime import datetime, timedelta
from collections import defaultdict
import secrets
import json
from urllib.parse import quote
from utils.permissions import FEATURES
from sqlalchemy.exc import SQLAlchemyError
from .ai_tools import AI_TOOLS, call_ai_tool, REPORTS_DIR, REPORT_FILENAME_RE

def superadmin_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if not current_user.is_authenticated or current_user.role != 'superadmin':
            flash("Accès refusé. Vous devez être superadmin.", "danger")
            return redirect(url_for('admin.dashboard'))
        return f(*args, **kwargs)
    return decorated_function

def permission_required(feature):
    def decorator(f):
        @wraps(f)
        def decorated_function(*args, **kwargs):
            if not current_user.is_authenticated or not current_user.has_permission(feature):
                flash(f"Accès refusé. Vous n'avez pas la permission : {feature}", "danger")
                return redirect(url_for('admin.dashboard'))
            return f(*args, **kwargs)
        return decorated_function
    return decorator

def generate_product_code(fournisseur):
    prefix = (fournisseur.prefixe or 'XXXX').upper()
    while True:
        suffix = ''.join(secrets.choice('0123456789') for _ in range(9))
        code = f'{prefix}-{suffix}'[:13]
        if not Produit.query.filter_by(code_produit=code).first():
            return code

def create_stock_modification(stock, produit, action, reason, old_values, new_values, old_qr_tire, new_qr_tire, reason_id=None):
    modification = StockModification(
        stock_id=None,
        produit=produit,
        user_id=current_user.id,
        action=action,
        reason=reason if not reason_id else None,
        reason_id=reason_id,
        numero_bl=stock.numero_bl,
        date_peremption=stock.date_peremption,
        code_suivi=stock.code_suivi,
        old_qr_tire=old_qr_tire,
        new_qr_tire=new_qr_tire,
        old_quantite_unites=old_values[0],
        old_quantite_sous_unites=old_values[1],
        old_quantite_sous_sous_unites=old_values[2],
        new_quantite_unites=new_values[0],
        new_quantite_sous_unites=new_values[1],
        new_quantite_sous_sous_unites=new_values[2]
    )
    db.session.add(modification)

def create_stock_exit_log(stock, reason, old_values, new_values, exit_values):
    fournisseur = stock.produit.fournisseur
    groupe_fournisseur = fournisseur.groupe if fournisseur else None
    stock_creation = (
        StockModification.query
        .filter_by(code_suivi=stock.code_suivi, action='create')
        .order_by(StockModification.created_at.asc())
        .first()
    )
    stocked_by = stock_creation.user if stock_creation and stock_creation.user else None
    prix_unite_ht = float(stock.produit.prix_unite or 0)
    prix_sous_unite_ht = float(stock.produit.prix_sous_unite or 0)
    prix_sous_sous_unite_ht = float(stock.produit.prix_sous_sous_unite or 0)
    prix_unite_ttc = float(stock.produit.prix_unite_ttc or 0)
    prix_sous_unite_ttc = float(stock.produit.prix_sous_unite_ttc or 0)
    prix_sous_sous_unite_ttc = float(stock.produit.prix_sous_sous_unite_ttc or 0)
    total_sortie_ht = (
        exit_values[0] * prix_unite_ht
        + exit_values[1] * prix_sous_unite_ht
        + exit_values[2] * prix_sous_sous_unite_ht
    )
    total_sortie_ttc = (
        exit_values[0] * prix_unite_ttc
        + exit_values[1] * prix_sous_unite_ttc
        + exit_values[2] * prix_sous_sous_unite_ttc
    )

    log = StockExitLog(
        produit_nom=stock.produit.nom,
        produit_code=stock.produit.code_produit,
        fournisseur_nom=fournisseur.nom if fournisseur else None,
        groupe_fournisseur_nom=groupe_fournisseur.nom if groupe_fournisseur else None,
        numero_bl=stock.numero_bl,
        date_peremption=stock.date_peremption,
        code_suivi=stock.code_suivi,
        mise_en_stock_at=stock_creation.created_at if stock_creation else stock.created_at,
        mise_en_stock_user_nom=stocked_by.nom if stocked_by else None,
        mise_en_stock_user_prenom=stocked_by.prenom if stocked_by else None,
        mise_en_stock_user_email=stocked_by.email if stocked_by else None,
        user_nom=current_user.nom,
        user_prenom=current_user.prenom,
        user_email=current_user.email,
        reason_nom=reason.nom,
        quantite_unites_sortie=exit_values[0],
        quantite_sous_unites_sortie=exit_values[1],
        quantite_sous_sous_unites_sortie=exit_values[2],
        prix_unite_ht=prix_unite_ht,
        prix_sous_unite_ht=prix_sous_unite_ht,
        prix_sous_sous_unite_ht=prix_sous_sous_unite_ht,
        prix_unite_ttc=prix_unite_ttc,
        prix_sous_unite_ttc=prix_sous_unite_ttc,
        prix_sous_sous_unite_ttc=prix_sous_sous_unite_ttc,
        tva_pourcentage=float(stock.produit.effectif_tva or 0),
        total_sortie_ht=total_sortie_ht,
        total_sortie_ttc=total_sortie_ttc,
        old_quantite_unites=old_values[0],
        old_quantite_sous_unites=old_values[1],
        old_quantite_sous_sous_unites=old_values[2],
        new_quantite_unites=new_values[0],
        new_quantite_sous_unites=new_values[1],
        new_quantite_sous_sous_unites=new_values[2]
    )
    db.session.add(log)

# --- GESTION DES RAISONS DE STOCK ---
@admin.route('/stock/reasons')
@login_required
@permission_required('gestion_raisons_stock')
def list_stock_reasons():
    reasons = StockReason.query.all()
    return render_template('admin/stock/reasons.html', reasons=reasons)

@admin.route('/stock/reasons/create', methods=['GET', 'POST'])
@login_required
@permission_required('gestion_raisons_stock')
def create_stock_reason():
    if request.method == 'POST':
        new_reason = StockReason(
            nom=request.form.get('nom'),
            type=request.form.get('type'),
            description=request.form.get('description')
        )
        db.session.add(new_reason)
        db.session.commit()
        flash('Raison de stock ajoutée avec succès.', 'success')
        return redirect(url_for('admin.list_stock_reasons'))
    return render_template('admin/stock/reason_form.html', title="Ajouter une raison de stock")

@admin.route('/stock/reasons/edit/<int:id>', methods=['GET', 'POST'])
@login_required
@permission_required('gestion_raisons_stock')
def edit_stock_reason(id):
    reason = StockReason.query.get_or_404(id)
    if request.method == 'POST':
        reason.nom = request.form.get('nom')
        reason.type = request.form.get('type')
        reason.description = request.form.get('description')
        db.session.commit()
        flash('Raison de stock mise à jour avec succès.', 'success')
        return redirect(url_for('admin.list_stock_reasons'))
    return render_template('admin/stock/reason_form.html', reason=reason, title="Modifier la raison de stock")

@admin.route('/stock/reasons/delete/<int:id>', methods=['POST'])
@login_required
@permission_required('gestion_raisons_stock')
def delete_stock_reason(id):
    reason = StockReason.query.get_or_404(id)
    db.session.delete(reason)
    db.session.commit()
    flash('Raison de stock supprimée.', 'success')
    return redirect(url_for('admin.list_stock_reasons'))

def get_lan_ip():
    """Detecte l'IP locale du serveur sur le reseau (pour connexion depuis un mobile)."""
    s = socket.socket(socket.AF_INET, socket.SOCK_DGRAM)
    try:
        s.connect(('8.8.8.8', 80))
        return s.getsockname()[0]
    except Exception:
        return '127.0.0.1'
    finally:
        s.close()

def build_qr_svg_data_uri(value, size=96):
    from reportlab.graphics.barcode import qr
    from reportlab.graphics.shapes import Drawing
    from reportlab.graphics import renderSVG

    widget = qr.QrCodeWidget(value)
    bounds = widget.getBounds()
    width = bounds[2] - bounds[0]
    height = bounds[3] - bounds[1]
    drawing = Drawing(size, size, transform=[size / width, 0, 0, size / height, 0, 0])
    drawing.add(widget)
    svg = renderSVG.drawToString(drawing)
    return f"data:image/svg+xml;utf8,{quote(svg)}"

def build_stock_qr_items(stocks, requested_counts=None):
    requested_counts = requested_counts or {}
    return [
        {
            'stock': stock,
            'qr_count': max(requested_counts.get(stock.id, stock.quantite_unites), 0),
            'qr_image': build_qr_svg_data_uri(stock.code_suivi)
        }
        for stock in stocks
    ]

def get_requested_qr_counts(stock_ids):
    counts = {}
    for stock_id in stock_ids:
        raw_count = request.form.get(f'qr_count_{stock_id}')
        if raw_count is None:
            continue
        try:
            counts[stock_id] = max(int(raw_count), 0)
        except ValueError:
            counts[stock_id] = 0
    return counts

def get_selected_stock_ids():
    return [int(stock_id) for stock_id in request.form.getlist('stock_ids') if stock_id.isdigit()]

def get_stocks_in_requested_order(stock_ids):
    stocks = Stock.query.filter(Stock.id.in_(stock_ids)).all()
    stock_by_id = {stock.id: stock for stock in stocks}
    return [stock_by_id[stock_id] for stock_id in stock_ids if stock_id in stock_by_id]

def get_qr_print_info_map(stocks):
    codes = [stock.code_suivi for stock in stocks]
    if not codes:
        return {}

    modifications = (
        StockModification.query
        .filter(StockModification.action == 'qr_print', StockModification.code_suivi.in_(codes))
        .order_by(StockModification.created_at.desc())
        .all()
    )
    info_by_code = {}
    for modification in modifications:
        if modification.code_suivi not in info_by_code:
            info_by_code[modification.code_suivi] = modification
    return info_by_code

def build_stock_qr_report_rows(stocks):
    qr_print_info = get_qr_print_info_map(stocks)
    rows = []
    for stock in stocks:
        print_info = qr_print_info.get(stock.code_suivi)
        user_label = '-'
        if print_info and print_info.user:
            user_label = f'{print_info.user.nom} {print_info.user.prenom}'

        rows.append({
            'produit': stock.produit.nom,
            'code_suivi': stock.code_suivi,
            'numero_bl': stock.numero_bl,
            'date_peremption': stock.date_peremption.strftime('%d/%m/%Y') if stock.date_peremption else '-',
            'unites': stock.quantite_unites,
            'sous_unites': stock.quantite_sous_unites,
            'sous_sous_unites': stock.quantite_sous_sous_unites,
            'qr_tire': 'Oui' if stock.qr_tire else 'Non',
            'date_tirage': print_info.created_at.strftime('%d/%m/%Y %H:%M') if print_info and print_info.created_at else '-',
            'tire_par': user_label
        })
    return rows

# --- DASHBOARD ---
@admin.route('/dashboard')
@login_required
def dashboard():
    today = datetime.utcnow().date()
    today_start = datetime(today.year, today.month, today.day)

    produits_count = Produit.query.count()
    stock_count = Stock.query.count()
    clients_count = Client.query.count()
    today_ventes_count = Vente.query.filter(Vente.created_at >= today_start).count()
    today_ventes_total = db.session.query(
        db.func.coalesce(db.func.sum(Vente.total_ttc), 0)
    ).filter(Vente.created_at >= today_start).scalar()

    port = request.host.split(':')[1] if ':' in request.host else ('443' if request.scheme == 'https' else '80')
    lan_login_url = f"{request.scheme}://{get_lan_ip()}:{port}{url_for('auth.login')}"
    lan_login_qr = build_qr_svg_data_uri(lan_login_url, size=180)

    _activer_inventaires_planifies()
    active_inventaire = Inventaire.query.filter_by(statut='en_cours').first()
    planned_inventaire = Inventaire.query.filter_by(statut='planifie').order_by(Inventaire.date_planifiee.asc()).first()
    active_inventaire_progress = None
    if active_inventaire:
        total_lignes = InventaireLigne.query.filter_by(inventaire_id=active_inventaire.id).count()
        scanned_lignes = InventaireLigne.query.filter_by(inventaire_id=active_inventaire.id, is_scanned=True).count()
        active_inventaire_progress = {'total': total_lignes, 'scanned': scanned_lignes}

    return render_template('admin/dashboard.html',
        produits_count=produits_count,
        stock_count=stock_count,
        clients_count=clients_count,
        today_ventes_count=today_ventes_count,
        today_ventes_total=today_ventes_total,
        lan_login_url=lan_login_url,
        lan_login_qr=lan_login_qr,
        active_inventaire=active_inventaire,
        active_inventaire_progress=active_inventaire_progress,
        planned_inventaire=planned_inventaire
    )

# --- GESTION DES POSTES (METIERS) ---
@admin.route('/postes')
@login_required
@permission_required('gestion_postes')
def list_postes():
    postes = Poste.query.all()
    from sqlalchemy import func
    poste_counts = dict(db.session.query(User.poste, func.count(User.id)).group_by(User.poste).all())
    return render_template('admin/postes/list.html', postes=postes, poste_counts=poste_counts)

@admin.route('/postes/create', methods=['GET', 'POST'])
@login_required
@permission_required('gestion_postes')
def create_poste():
    if request.method == 'POST':
        nom = request.form.get('nom')
        if Poste.query.filter_by(nom=nom).first():
            flash('Ce poste existe déjà.', 'danger')
            return redirect(url_for('admin.create_poste'))
        
        new_poste = Poste(nom=nom, description=request.form.get('description'))
        db.session.add(new_poste)
        db.session.commit()
        flash('Poste ajouté avec succès.', 'success')
        return redirect(url_for('admin.list_postes'))
    return render_template('admin/postes/form.html', title="Ajouter un Poste")

@admin.route('/postes/edit/<int:id>', methods=['GET', 'POST'])
@login_required
@permission_required('gestion_postes')
def edit_poste(id):
    poste = Poste.query.get_or_404(id)
    if request.method == 'POST':
        poste.nom = request.form.get('nom')
        poste.description = request.form.get('description')
        
        # Gestion des permissions du poste (uniquement par superadmin)
        if current_user.role == 'superadmin':
            for feature in FEATURES:
                is_allowed = True if request.form.get(f'perm_{feature}') else False
                perm = Permission.query.filter_by(feature=feature, poste_id=poste.id).first()
                if perm:
                    perm.is_allowed = is_allowed
                else:
                    new_perm = Permission(feature=feature, poste_id=poste.id, is_allowed=is_allowed)
                    db.session.add(new_perm)
        
        db.session.commit()
        flash('Poste mis à jour avec succès.', 'success')
        return redirect(url_for('admin.list_postes'))
    
    poste_perms = {p.feature: p.is_allowed for p in Permission.query.filter_by(poste_id=id).all()}
    return render_template('admin/postes/form.html', poste=poste, title="Modifier le Poste", features=FEATURES, poste_perms=poste_perms)

@admin.route('/postes/<int:id>/permission/toggle', methods=['POST'])
@login_required
@permission_required('gestion_postes')
def toggle_poste_permission(id):
    if current_user.role != 'superadmin':
        return {'success': False, 'message': "Action réservée au superadmin."}, 403

    poste = Poste.query.get_or_404(id)
    feature = request.form.get('feature')
    if feature not in FEATURES:
        return {'success': False, 'message': 'Permission inconnue.'}, 400

    is_allowed = request.form.get('is_allowed') == 'true'
    perm = Permission.query.filter_by(feature=feature, poste_id=poste.id).first()
    if perm:
        perm.is_allowed = is_allowed
    else:
        perm = Permission(feature=feature, poste_id=poste.id, is_allowed=is_allowed)
        db.session.add(perm)
    db.session.commit()

    return {'success': True, 'feature': feature, 'is_allowed': is_allowed}

@admin.route('/postes/delete/<int:id>', methods=['POST'])
@login_required
@permission_required('gestion_postes')
def delete_poste(id):
    poste = Poste.query.get_or_404(id)
    db.session.delete(poste)
    db.session.commit()
    flash('Poste supprimé.', 'success')
    return redirect(url_for('admin.list_postes'))

@admin.route('/postes/bulk-delete', methods=['POST'])
@login_required
@permission_required('gestion_postes')
def bulk_delete_postes():
    ids = request.form.getlist('ids[]')
    if not ids:
        flash("Aucun poste sélectionné.", "warning")
        return redirect(url_for('admin.list_postes'))
    
    deleted_count = 0
    for p_id in ids:
        p = Poste.query.get(p_id)
        if p:
            db.session.delete(p)
            deleted_count += 1
            
    db.session.commit()
    flash(f'{deleted_count} poste(s) supprimé(s).', 'success')
    return redirect(url_for('admin.list_postes'))

# --- GESTION DES UTILISATEURS ---
@admin.route('/users')
@login_required
@permission_required('gestion_employes')
def list_users():
    users = User.query.all()
    total_users = len(users)
    active_users = sum(1 for u in users if u.is_active)
    inactive_users = total_users - active_users
    return render_template('admin/users/list.html',
        users=users,
        total_users=total_users,
        active_users=active_users,
        inactive_users=inactive_users
    )

@admin.route('/users/create', methods=['GET', 'POST'])
@login_required
@permission_required('gestion_employes')
def create_user():
    postes = Poste.query.all()
    if request.method == 'POST':
        email = request.form.get('email')
        username = (request.form.get('username') or '').strip() or None
        if User.query.filter_by(email=email).first():
            flash('Cet email est déjà utilisé.', 'danger')
            return redirect(url_for('admin.create_user'))
        if username and User.query.filter_by(username=username).first():
            flash("Ce nom d'utilisateur est déjà utilisé.", 'danger')
            return redirect(url_for('admin.create_user'))

        date_prise_poste_str = request.form.get('date_prise_poste')
        date_prise_poste = datetime.strptime(date_prise_poste_str, '%Y-%m-%d').date() if date_prise_poste_str else None
        salaire_mensuel = request.form.get('salaire_mensuel')
        salaire_mensuel = float(salaire_mensuel) if salaire_mensuel else None

        new_user = User(
            nom=request.form.get('nom'),
            prenom=request.form.get('prenom'),
            email=email,
            username=username,
            telephone=request.form.get('telephone'),
            adresse=request.form.get('adresse'),
            role=request.form.get('role'),
            poste=request.form.get('poste'),
            date_prise_poste=date_prise_poste,
            salaire_mensuel=salaire_mensuel,
            is_active=True if request.form.get('is_active') else False
        )
        new_user.set_password(request.form.get('password'))
        
        db.session.add(new_user)
        db.session.commit()
        flash('Utilisateur créé avec succès.', 'success')
        return redirect(url_for('admin.list_users'))
    return render_template('admin/users/form.html', title="Créer un utilisateur", postes=postes)

@admin.route('/users/edit/<int:id>', methods=['GET', 'POST'])
@login_required
@permission_required('gestion_employes')
def edit_user(id):
    user = User.query.get_or_404(id)
    postes = Poste.query.all()
    if request.method == 'POST':
        username = (request.form.get('username') or '').strip() or None
        if username and User.query.filter(User.username == username, User.id != user.id).first():
            flash("Ce nom d'utilisateur est déjà utilisé.", 'danger')
            return redirect(url_for('admin.edit_user', id=id))

        user.nom = request.form.get('nom')
        user.prenom = request.form.get('prenom')
        user.email = request.form.get('email')
        user.username = username
        user.telephone = request.form.get('telephone')
        user.adresse = request.form.get('adresse')
        user.role = request.form.get('role')
        user.poste = request.form.get('poste')
        user.is_active = True if request.form.get('is_active') else False
        
        date_prise_poste_str = request.form.get('date_prise_poste')
        user.date_prise_poste = datetime.strptime(date_prise_poste_str, '%Y-%m-%d').date() if date_prise_poste_str else None
        salaire_mensuel = request.form.get('salaire_mensuel')
        user.salaire_mensuel = float(salaire_mensuel) if salaire_mensuel else None
        
        password = request.form.get('password')
        if password:
            user.set_password(password)
            
        # Gestion des permissions (uniquement par superadmin)
        if current_user.role == 'superadmin':
            for feature in FEATURES:
                is_allowed = True if request.form.get(f'perm_{feature}') else False
                perm = Permission.query.filter_by(feature=feature, user_id=user.id).first()
                if perm:
                    perm.is_allowed = is_allowed
                else:
                    new_perm = Permission(feature=feature, user_id=user.id, is_allowed=is_allowed)
                    db.session.add(new_perm)

        db.session.commit()
        flash('Utilisateur mis à jour avec succès.', 'success')
        return redirect(url_for('admin.list_users'))
    
    # Récupérer les permissions actuelles de l'utilisateur
    user_perms = {p.feature: p.is_allowed for p in Permission.query.filter_by(user_id=id).all()}
    return render_template('admin/users/form.html', user=user, title="Modifier l'utilisateur", postes=postes, features=FEATURES, user_perms=user_perms)

@admin.route('/users/bulk-delete', methods=['POST'])
@login_required
@permission_required('gestion_employes')
def bulk_delete_users():
    ids = request.form.getlist('ids[]')
    if not ids:
        flash("Aucun utilisateur sélectionné.", "warning")
        return redirect(url_for('admin.list_users'))
    
    deleted_count = 0
    for u_id in ids:
        u = User.query.get(u_id)
        if u and u.id != current_user.id:
            db.session.delete(u)
            deleted_count += 1
            
    db.session.commit()
    flash(f'{deleted_count} utilisateur(s) supprimé(s).', 'success')
    return redirect(url_for('admin.list_users'))

@admin.route('/users/toggle-active/<int:id>')
@login_required
@permission_required('gestion_employes')
def toggle_user_active(id):
    user = User.query.get_or_404(id)
    if user.id == current_user.id:
        flash("Vous ne pouvez pas modifier le statut de votre propre compte.", "warning")
        return redirect(url_for('admin.list_users'))

    user.is_active = not user.is_active
    db.session.commit()
    flash(
        'Utilisateur activ? avec succ?s.' if user.is_active else 'Utilisateur d?sactiv? avec succ?s.',
        'success'
    )
    return redirect(url_for('admin.list_users'))

@admin.route('/users/delete/<int:id>', methods=['POST'])
@login_required
@permission_required('gestion_employes')
def delete_user(id):
    user = User.query.get_or_404(id)
    if user.id == current_user.id:
        flash("Vous ne pouvez pas supprimer votre propre compte.", "warning")
        return redirect(url_for('admin.list_users'))

    db.session.delete(user)
    db.session.commit()
    flash('Utilisateur supprim?.', 'success')
    return redirect(url_for('admin.list_users'))

# --- GESTION DES CLIENTS ---
def client_snapshot(client):
    return {
        'matricule': client.matricule,
        'nom': client.nom,
        'prenom': client.prenom,
        'email': client.email,
        'telephone': client.telephone,
        'solde': client.solde,
        'groupe': client.groupe.nom if client.groupe else None
    }

def groupe_client_snapshot(groupe):
    return {
        'nom': groupe.nom,
        'description': groupe.description,
        'solde': groupe.solde,
        'pourcentage_absorption': groupe.pourcentage_absorption,
        'clients_count': len(groupe.clients) if groupe.clients is not None else 0
    }

def add_client_modification_log(entity_type, action, reference, label, old_values=None, new_values=None, reason=None):
    log = ClientModificationLog(
        entity_type=entity_type,
        action=action,
        reference=reference,
        label=label,
        old_values=json.dumps(old_values, ensure_ascii=False, default=str) if old_values is not None else None,
        new_values=json.dumps(new_values, ensure_ascii=False, default=str) if new_values is not None else None,
        reason=reason,
        user_nom=current_user.nom,
        user_prenom=current_user.prenom,
        user_email=current_user.email
    )
    db.session.add(log)

def parse_log_values(raw_values):
    if not raw_values:
        return {}
    try:
        return json.loads(raw_values)
    except ValueError:
        return {}

def generate_client_matricule():
    while True:
        suffix = ''.join(secrets.choice('0123456789') for _ in range(7))
        matricule = f'CL-{suffix}'
        if not Client.query.filter_by(matricule=matricule).first():
            return matricule

def parse_date_filter(value):
    try:
        return datetime.strptime(value, '%Y-%m-%d').date() if value else None
    except ValueError:
        return None

def get_filtered_clients():
    clients = Client.query.all()
    query = (request.args.get('q') or '').strip().lower()
    groupe_id = (request.args.get('groupe_id') or '').strip()
    solde_status = (request.args.get('solde_status') or '').strip()
    date_from = parse_date_filter((request.args.get('date_from') or '').strip())
    date_to = parse_date_filter((request.args.get('date_to') or '').strip())
    sort = (request.args.get('sort') or 'created_at').strip()
    direction = (request.args.get('direction') or 'desc').strip()

    if query:
        clients = [
            client for client in clients
            if query in ' '.join([
                client.matricule or '',
                client.nom or '',
                client.prenom or '',
                client.email or '',
                client.telephone or '',
                client.groupe.nom if client.groupe else ''
            ]).lower()
        ]
    if groupe_id == 'none':
        clients = [client for client in clients if not client.groupe]
    elif groupe_id:
        try:
            selected_groupe_id = int(groupe_id)
            clients = [client for client in clients if client.groupe_id == selected_groupe_id]
        except ValueError:
            clients = []
    if solde_status == 'positive':
        clients = [client for client in clients if (client.solde or 0) > 0]
    elif solde_status == 'zero':
        clients = [client for client in clients if (client.solde or 0) == 0]
    elif solde_status == 'negative':
        clients = [client for client in clients if (client.solde or 0) < 0]
    if date_from:
        clients = [client for client in clients if client.created_at and client.created_at.date() >= date_from]
    if date_to:
        clients = [client for client in clients if client.created_at and client.created_at.date() <= date_to]

    sorters = {
        'matricule': lambda client: (client.matricule or '').lower(),
        'nom': lambda client: (client.nom or '').lower(),
        'prenom': lambda client: (client.prenom or '').lower(),
        'email': lambda client: (client.email or '').lower(),
        'telephone': lambda client: (client.telephone or '').lower(),
        'groupe': lambda client: (client.groupe.nom if client.groupe else '').lower(),
        'solde': lambda client: client.solde or 0,
        'created_at': lambda client: client.created_at or datetime.min,
        'updated_at': lambda client: client.updated_at or datetime.min
    }
    clients.sort(key=sorters.get(sort, sorters['created_at']), reverse=direction != 'asc')
    return clients

def get_filtered_groupes_clients():
    groupes = GroupeClient.query.all()
    query = (request.args.get('q') or '').strip().lower()
    solde_status = (request.args.get('solde_status') or '').strip()
    absorption_min_raw = (request.args.get('absorption_min') or '').strip()
    absorption_max_raw = (request.args.get('absorption_max') or '').strip()
    date_from = parse_date_filter((request.args.get('date_from') or '').strip())
    date_to = parse_date_filter((request.args.get('date_to') or '').strip())
    sort = (request.args.get('sort') or 'nom').strip()
    direction = (request.args.get('direction') or 'asc').strip()

    try:
        absorption_min = float(absorption_min_raw) if absorption_min_raw else None
    except ValueError:
        absorption_min = None
    try:
        absorption_max = float(absorption_max_raw) if absorption_max_raw else None
    except ValueError:
        absorption_max = None

    if query:
        groupes = [
            groupe for groupe in groupes
            if query in ' '.join([groupe.nom or '', groupe.description or '']).lower()
        ]
    if solde_status == 'positive':
        groupes = [groupe for groupe in groupes if (groupe.solde or 0) > 0]
    elif solde_status == 'zero':
        groupes = [groupe for groupe in groupes if (groupe.solde or 0) == 0]
    elif solde_status == 'negative':
        groupes = [groupe for groupe in groupes if (groupe.solde or 0) < 0]
    if absorption_min is not None:
        groupes = [groupe for groupe in groupes if (groupe.pourcentage_absorption or 0) >= absorption_min]
    if absorption_max is not None:
        groupes = [groupe for groupe in groupes if (groupe.pourcentage_absorption or 0) <= absorption_max]
    if date_from:
        groupes = [groupe for groupe in groupes if groupe.created_at and groupe.created_at.date() >= date_from]
    if date_to:
        groupes = [groupe for groupe in groupes if groupe.created_at and groupe.created_at.date() <= date_to]

    sorters = {
        'nom': lambda groupe: (groupe.nom or '').lower(),
        'solde': lambda groupe: groupe.solde or 0,
        'pourcentage_absorption': lambda groupe: groupe.pourcentage_absorption or 0,
        'clients_count': lambda groupe: len(groupe.clients),
        'created_at': lambda groupe: groupe.created_at or datetime.min,
        'updated_at': lambda groupe: groupe.updated_at or datetime.min
    }
    groupes.sort(key=sorters.get(sort, sorters['nom']), reverse=direction != 'asc')
    return groupes

@admin.route('/clients')
@login_required
@permission_required('gestion_clients')
def list_clients():
    clients = get_filtered_clients()
    groupes = GroupeClient.query.order_by(GroupeClient.nom.asc()).all()
    return render_template('admin/clients/list.html', clients=clients, groupes=groupes, export_query=request.args.to_dict())

@admin.route('/clients/export/excel')
@login_required
@permission_required('gestion_clients')
def export_clients_excel():
    import io
    import pandas as pd
    from flask import send_file
    from openpyxl.styles import Alignment, Font, PatternFill, Border, Side

    clients = get_filtered_clients()
    generated_at = datetime.now()
    rows = []
    for client in clients:
        rows.append({
            'Matricule': client.matricule,
            'Nom': client.nom,
            'Prenom': client.prenom,
            'Email': client.email or '',
            'Telephone': client.telephone or '',
            'Solde client': client.solde or 0,
            'Groupe': client.groupe.nom if client.groupe else '',
            'Solde groupe': client.groupe.solde if client.groupe else '',
            'Absorption groupe (%)': client.groupe.pourcentage_absorption if client.groupe else '',
            'Part client (%)': 100 - client.groupe.pourcentage_absorption if client.groupe else '',
            'Creation': client.created_at.strftime('%d/%m/%Y %H:%M') if client.created_at else '',
            'Modification': client.updated_at.strftime('%d/%m/%Y %H:%M') if client.updated_at else ''
        })

    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        pd.DataFrame(rows).to_excel(writer, index=False, sheet_name='Clients', startrow=4)
        worksheet = writer.sheets['Clients']
        worksheet['A1'] = 'Clients - ReflexPharma'
        worksheet['A2'] = f'Date du tirage : {generated_at.strftime("%d/%m/%Y %H:%M")}'
        worksheet['A3'] = f'Tire par : {current_user.nom} {current_user.prenom}'
        worksheet['A4'] = f'Lignes exportees : {len(rows)}'
        header_fill = PatternFill(start_color='1F2937', end_color='1F2937', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF')
        thin_border = Border(bottom=Side(style='thin', color='D1D5DB'))
        for cell in worksheet[5]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        for row in worksheet.iter_rows(min_row=6, max_row=worksheet.max_row):
            for cell in row:
                cell.border = thin_border
                cell.alignment = Alignment(vertical='top', wrap_text=True)
        widths = [16, 18, 18, 28, 18, 16, 22, 16, 20, 16, 18, 18]
        for index, width in enumerate(widths, start=1):
            worksheet.column_dimensions[chr(64 + index)].width = width
        worksheet.freeze_panes = 'A6'

    output.seek(0)
    filename = f'clients_{generated_at.strftime("%Y%m%d_%H%M")}.xlsx'
    return send_file(output, download_name=filename, as_attachment=True)

@admin.route('/clients/export/pdf')
@login_required
@permission_required('gestion_clients')
def export_clients_pdf():
    import io
    from flask import send_file
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
    from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle
    from xml.sax.saxutils import escape

    clients = get_filtered_clients()
    generated_at = datetime.now()
    output = io.BytesIO()
    doc = SimpleDocTemplate(output, pagesize=landscape(A4), topMargin=12, bottomMargin=12, leftMargin=12, rightMargin=12)
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle('ClientsTitle', parent=styles['Title'], fontSize=12, leading=14)
    meta_style = ParagraphStyle('ClientsMeta', parent=styles['Normal'], fontSize=7, leading=9)
    cell_style = ParagraphStyle('ClientsCell', parent=styles['Normal'], fontSize=6, leading=7)
    elements = [
        Paragraph('Clients - ReflexPharma', title_style),
        Paragraph(f'Date du tirage : {generated_at.strftime("%d/%m/%Y %H:%M")} | Tire par : {current_user.nom} {current_user.prenom} | Lignes : {len(clients)}', meta_style),
        Spacer(1, 6)
    ]
    data = [['Matricule', 'Client', 'Contact', 'Groupe', 'Solde', 'Abs.', 'Creation', 'Modif.']]
    for client in clients:
        data.append([
            client.matricule,
            Paragraph(escape(client.nom_complet or '-'), cell_style),
            Paragraph(escape(f'{client.telephone or "-"}\n{client.email or "-"}'), cell_style),
            Paragraph(escape(client.groupe.nom if client.groupe else '-'), cell_style),
            f'{client.solde or 0:.2f}',
            f'{client.groupe.pourcentage_absorption:.2f}%' if client.groupe else '-',
            client.created_at.strftime('%d/%m/%Y') if client.created_at else '-',
            client.updated_at.strftime('%d/%m/%Y') if client.updated_at else '-'
        ])
    table = Table(data, repeatRows=1, colWidths=[70, 110, 145, 105, 65, 45, 60, 60])
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1F2937')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 6),
        ('LEADING', (0, 0), (-1, -1), 7),
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ('GRID', (0, 0), (-1, -1), 0.25, colors.HexColor('#D1D5DB')),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F9FAFB')]),
        ('TOPPADDING', (0, 0), (-1, -1), 2),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 2),
        ('LEFTPADDING', (0, 0), (-1, -1), 2),
        ('RIGHTPADDING', (0, 0), (-1, -1), 2),
    ]))
    elements.append(table)
    doc.build(elements)
    output.seek(0)
    filename = f'clients_{generated_at.strftime("%Y%m%d_%H%M")}.pdf'
    return send_file(output, download_name=filename, as_attachment=True)

@admin.route('/clients/create', methods=['GET', 'POST'])
@login_required
@permission_required('gestion_clients')
def create_client():
    groupes = GroupeClient.query.order_by(GroupeClient.nom.asc()).all()
    if request.method == 'POST':
        matricule = (request.form.get('matricule') or '').strip().upper() or generate_client_matricule()
        email = (request.form.get('email') or '').strip() or None
        if Client.query.filter_by(matricule=matricule).first():
            flash('Ce matricule client existe dÃ©jÃ .', 'danger')
            return redirect(url_for('admin.create_client'))
        if email and Client.query.filter_by(email=email).first():
            flash('Cet email est dÃ©jÃ  utilisÃ© par un client.', 'danger')
            return redirect(url_for('admin.create_client'))

        groupe_id = request.form.get('groupe_id')
        client = Client(
            matricule=matricule,
            nom=request.form.get('nom'),
            prenom=request.form.get('prenom'),
            email=email,
            telephone=request.form.get('telephone'),
            solde=float(request.form.get('solde') or 0),
            groupe_id=int(groupe_id) if groupe_id else None
        )
        db.session.add(client)
        db.session.flush()
        add_client_modification_log(
            entity_type='client',
            action='create',
            reference=client.matricule,
            label=client.nom_complet,
            new_values=client_snapshot(client)
        )
        db.session.commit()
        flash('Client ajoutÃ© avec succÃ¨s.', 'success')
        return redirect(url_for('admin.list_clients'))

    return render_template(
        'admin/clients/form.html',
        title='Ajouter un client',
        groupes=groupes,
        suggested_matricule=generate_client_matricule()
    )

@admin.route('/clients/edit/<int:id>', methods=['GET', 'POST'])
@login_required
@permission_required('gestion_clients')
def edit_client(id):
    client = Client.query.get_or_404(id)
    groupes = GroupeClient.query.order_by(GroupeClient.nom.asc()).all()
    if request.method == 'POST':
        old_values = client_snapshot(client)
        matricule = (request.form.get('matricule') or '').strip().upper()
        email = (request.form.get('email') or '').strip() or None
        existing_matricule = Client.query.filter_by(matricule=matricule).first()
        if existing_matricule and existing_matricule.id != client.id:
            flash('Ce matricule client existe dÃ©jÃ .', 'danger')
            return redirect(url_for('admin.edit_client', id=id))
        if email:
            existing_email = Client.query.filter_by(email=email).first()
            if existing_email and existing_email.id != client.id:
                flash('Cet email est dÃ©jÃ  utilisÃ© par un client.', 'danger')
                return redirect(url_for('admin.edit_client', id=id))

        groupe_id = request.form.get('groupe_id')
        client.matricule = matricule
        client.nom = request.form.get('nom')
        client.prenom = request.form.get('prenom')
        client.email = email
        client.telephone = request.form.get('telephone')
        client.solde = float(request.form.get('solde') or 0)
        client.groupe_id = int(groupe_id) if groupe_id else None
        db.session.flush()
        add_client_modification_log(
            entity_type='client',
            action='edit',
            reference=client.matricule,
            label=client.nom_complet,
            old_values=old_values,
            new_values=client_snapshot(client)
        )
        db.session.commit()
        flash('Client mis Ã  jour avec succÃ¨s.', 'success')
        return redirect(url_for('admin.list_clients'))

    return render_template('admin/clients/form.html', title='Modifier le client', client=client, groupes=groupes)

@admin.route('/clients/delete/<int:id>', methods=['POST'])
@login_required
@permission_required('gestion_clients')
def delete_client(id):
    client = Client.query.get_or_404(id)
    old_values = client_snapshot(client)
    reference = client.matricule
    label = client.nom_complet
    add_client_modification_log(
        entity_type='client',
        action='delete',
        reference=reference,
        label=label,
        old_values=old_values
    )
    db.session.delete(client)
    db.session.commit()
    flash('Client supprimÃ©.', 'success')
    return redirect(url_for('admin.list_clients'))

@admin.route('/clients/bulk-delete', methods=['POST'])
@login_required
@permission_required('gestion_clients')
def bulk_delete_clients():
    ids = request.form.getlist('ids[]')
    deleted_count = 0
    for client_id in ids:
        client = Client.query.get(client_id)
        if client:
            add_client_modification_log(
                entity_type='client',
                action='delete',
                reference=client.matricule,
                label=client.nom_complet,
                old_values=client_snapshot(client),
                reason='Suppression groupée'
            )
            db.session.delete(client)
            deleted_count += 1
    db.session.commit()
    flash(f'{deleted_count} client(s) supprimÃ©(s).', 'success')
    return redirect(url_for('admin.list_clients'))

# --- GESTION DES GROUPES CLIENTS ---
@admin.route('/clients/groupes')
@login_required
@permission_required('gestion_groupes_clients')
def list_groupes_clients():
    groupes = get_filtered_groupes_clients()
    return render_template('admin/clients/groupes_list.html', groupes=groupes, export_query=request.args.to_dict())

@admin.route('/clients/groupes/export/excel')
@login_required
@permission_required('gestion_groupes_clients')
def export_groupes_clients_excel():
    import io
    import pandas as pd
    from flask import send_file
    from openpyxl.styles import Alignment, Font, PatternFill, Border, Side

    groupes = get_filtered_groupes_clients()
    generated_at = datetime.now()
    rows = []
    for groupe in groupes:
        rows.append({
            'Groupe': groupe.nom,
            'Description': groupe.description or '',
            'Solde groupe': groupe.solde or 0,
            'Absorption groupe (%)': groupe.pourcentage_absorption or 0,
            'Part client (%)': 100 - (groupe.pourcentage_absorption or 0),
            'Nombre de clients': len(groupe.clients),
            'Clients': ', '.join([f'{client.prenom} {client.nom} ({client.matricule})' for client in groupe.clients]),
            'Creation': groupe.created_at.strftime('%d/%m/%Y %H:%M') if groupe.created_at else '',
            'Modification': groupe.updated_at.strftime('%d/%m/%Y %H:%M') if groupe.updated_at else ''
        })

    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        pd.DataFrame(rows).to_excel(writer, index=False, sheet_name='Groupes clients', startrow=4)
        worksheet = writer.sheets['Groupes clients']
        worksheet['A1'] = 'Groupes clients - ReflexPharma'
        worksheet['A2'] = f'Date du tirage : {generated_at.strftime("%d/%m/%Y %H:%M")}'
        worksheet['A3'] = f'Tire par : {current_user.nom} {current_user.prenom}'
        worksheet['A4'] = f'Lignes exportees : {len(rows)}'
        header_fill = PatternFill(start_color='1F2937', end_color='1F2937', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF')
        thin_border = Border(bottom=Side(style='thin', color='D1D5DB'))
        for cell in worksheet[5]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        for row in worksheet.iter_rows(min_row=6, max_row=worksheet.max_row):
            for cell in row:
                cell.border = thin_border
                cell.alignment = Alignment(vertical='top', wrap_text=True)
        widths = [24, 36, 16, 22, 18, 18, 54, 18, 18]
        for index, width in enumerate(widths, start=1):
            worksheet.column_dimensions[chr(64 + index)].width = width
        worksheet.freeze_panes = 'A6'

    output.seek(0)
    filename = f'groupes_clients_{generated_at.strftime("%Y%m%d_%H%M")}.xlsx'
    return send_file(output, download_name=filename, as_attachment=True)

@admin.route('/clients/groupes/export/pdf')
@login_required
@permission_required('gestion_groupes_clients')
def export_groupes_clients_pdf():
    import io
    from flask import send_file
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
    from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle
    from xml.sax.saxutils import escape

    groupes = get_filtered_groupes_clients()
    generated_at = datetime.now()
    output = io.BytesIO()
    doc = SimpleDocTemplate(output, pagesize=landscape(A4), topMargin=12, bottomMargin=12, leftMargin=12, rightMargin=12)
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle('GroupesClientsTitle', parent=styles['Title'], fontSize=12, leading=14)
    meta_style = ParagraphStyle('GroupesClientsMeta', parent=styles['Normal'], fontSize=7, leading=9)
    cell_style = ParagraphStyle('GroupesClientsCell', parent=styles['Normal'], fontSize=6, leading=7)
    elements = [
        Paragraph('Groupes clients - ReflexPharma', title_style),
        Paragraph(f'Date du tirage : {generated_at.strftime("%d/%m/%Y %H:%M")} | Tire par : {current_user.nom} {current_user.prenom} | Lignes : {len(groupes)}', meta_style),
        Spacer(1, 6)
    ]
    data = [['Groupe', 'Description', 'Solde', 'Abs.', 'Part client', 'Clients', 'Creation', 'Modif.']]
    for groupe in groupes:
        data.append([
            Paragraph(escape(groupe.nom or '-'), cell_style),
            Paragraph(escape(groupe.description or '-'), cell_style),
            f'{groupe.solde or 0:.2f}',
            f'{groupe.pourcentage_absorption or 0:.2f}%',
            f'{100 - (groupe.pourcentage_absorption or 0):.2f}%',
            str(len(groupe.clients)),
            groupe.created_at.strftime('%d/%m/%Y') if groupe.created_at else '-',
            groupe.updated_at.strftime('%d/%m/%Y') if groupe.updated_at else '-'
        ])
    table = Table(data, repeatRows=1, colWidths=[125, 190, 70, 55, 65, 50, 60, 60])
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1F2937')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 6),
        ('LEADING', (0, 0), (-1, -1), 7),
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ('GRID', (0, 0), (-1, -1), 0.25, colors.HexColor('#D1D5DB')),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F9FAFB')]),
        ('TOPPADDING', (0, 0), (-1, -1), 2),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 2),
        ('LEFTPADDING', (0, 0), (-1, -1), 2),
        ('RIGHTPADDING', (0, 0), (-1, -1), 2),
    ]))
    elements.append(table)
    doc.build(elements)
    output.seek(0)
    filename = f'groupes_clients_{generated_at.strftime("%Y%m%d_%H%M")}.pdf'
    return send_file(output, download_name=filename, as_attachment=True)

@admin.route('/clients/groupes/create', methods=['GET', 'POST'])
@login_required
@permission_required('gestion_groupes_clients')
def create_groupe_client():
    if request.method == 'POST':
        nom = request.form.get('nom')
        if GroupeClient.query.filter_by(nom=nom).first():
            flash('Ce groupe client existe dÃ©jÃ .', 'danger')
            return redirect(url_for('admin.create_groupe_client'))
        groupe = GroupeClient(
            nom=nom,
            description=request.form.get('description'),
            solde=float(request.form.get('solde') or 0),
            pourcentage_absorption=min(max(float(request.form.get('pourcentage_absorption') or 0), 0), 100)
        )
        db.session.add(groupe)
        db.session.flush()
        add_client_modification_log(
            entity_type='groupe_client',
            action='create',
            reference=groupe.nom,
            label=groupe.nom,
            new_values=groupe_client_snapshot(groupe)
        )
        db.session.commit()
        flash('Groupe client ajoutÃ© avec succÃ¨s.', 'success')
        return redirect(url_for('admin.list_groupes_clients'))
    return render_template('admin/clients/groupe_form.html', title='Ajouter un groupe client')

@admin.route('/clients/groupes/edit/<int:id>', methods=['GET', 'POST'])
@login_required
@permission_required('gestion_groupes_clients')
def edit_groupe_client(id):
    groupe = GroupeClient.query.get_or_404(id)
    if request.method == 'POST':
        old_values = groupe_client_snapshot(groupe)
        nom = request.form.get('nom')
        existing = GroupeClient.query.filter_by(nom=nom).first()
        if existing and existing.id != groupe.id:
            flash('Ce groupe client existe dÃ©jÃ .', 'danger')
            return redirect(url_for('admin.edit_groupe_client', id=id))
        groupe.nom = nom
        groupe.description = request.form.get('description')
        groupe.solde = float(request.form.get('solde') or 0)
        groupe.pourcentage_absorption = min(max(float(request.form.get('pourcentage_absorption') or 0), 0), 100)
        db.session.flush()
        add_client_modification_log(
            entity_type='groupe_client',
            action='edit',
            reference=groupe.nom,
            label=groupe.nom,
            old_values=old_values,
            new_values=groupe_client_snapshot(groupe)
        )
        db.session.commit()
        flash('Groupe client mis Ã  jour avec succÃ¨s.', 'success')
        return redirect(url_for('admin.list_groupes_clients'))
    return render_template('admin/clients/groupe_form.html', title='Modifier le groupe client', groupe=groupe)

@admin.route('/clients/groupes/delete/<int:id>', methods=['POST'])
@login_required
@permission_required('gestion_groupes_clients')
def delete_groupe_client(id):
    groupe = GroupeClient.query.get_or_404(id)
    old_values = groupe_client_snapshot(groupe)
    reference = groupe.nom
    label = groupe.nom
    for client in groupe.clients:
        client.groupe_id = None
    add_client_modification_log(
        entity_type='groupe_client',
        action='delete',
        reference=reference,
        label=label,
        old_values=old_values,
        reason='Suppression du groupe et détachement des clients'
    )
    db.session.delete(groupe)
    db.session.commit()
    flash('Groupe client supprimÃ©. Les clients associÃ©s ont Ã©tÃ© dÃ©tachÃ©s.', 'success')
    return redirect(url_for('admin.list_groupes_clients'))

# --- GESTION DES VENTES ---
def generate_numero_vente():
    prefix = datetime.now().strftime('V%Y%m%d')
    count = Vente.query.filter(Vente.numero_vente.like(f'{prefix}-%')).count() + 1
    while True:
        numero = f'{prefix}-{count:04d}'
        if not Vente.query.filter_by(numero_vente=numero).first():
            return numero
        count += 1

def get_product_unit_price(produit, unite):
    if unite == 'sous_unite':
        prix_ht = produit.prix_sous_unite if produit.prix_sous_unite is not None else produit.prix_unite
        prix_ttc = produit.prix_sous_unite_ttc if produit.prix_sous_unite_ttc is not None else produit.prix_unite_ttc
    elif unite == 'sous_sous_unite':
        prix_ht = produit.prix_sous_sous_unite if produit.prix_sous_sous_unite is not None else produit.prix_unite
        prix_ttc = produit.prix_sous_sous_unite_ttc if produit.prix_sous_sous_unite_ttc is not None else produit.prix_unite_ttc
    else:
        prix_ht = produit.prix_unite
        prix_ttc = produit.prix_unite_ttc
    return float(prix_ht or 0), float(prix_ttc or 0)

def normalize_product_unit(produit, unite, stock_summary=None):
    conditionnement = min(max(int(produit.conditionnement or 1), 1), 3)
    if stock_summary:
        if (stock_summary.get('sous_sous_unite') or 0) > 0:
            conditionnement = max(conditionnement, 3)
        elif (stock_summary.get('sous_unite') or 0) > 0:
            conditionnement = max(conditionnement, 2)
    if unite == 'sous_sous_unite' and conditionnement >= 3:
        return unite
    if unite == 'sous_unite' and conditionnement >= 2:
        return unite
    return 'unite'

def get_product_stock_summary(produit):
    return {
        'unite': sum((stock.quantite_unites or 0) for stock in produit.stocks),
        'sous_unite': sum((stock.quantite_sous_unites or 0) for stock in produit.stocks),
        'sous_sous_unite': sum((stock.quantite_sous_sous_unites or 0) for stock in produit.stocks)
    }

def get_products_stock_totals(produits):
    return {produit.id: get_product_stock_summary(produit) for produit in produits}

def get_products_stock_tracking_codes(produits):
    codes_by_product = {}
    for produit in produits:
        codes = []
        for stock in produit.stocks:
            total = (
                (stock.quantite_unites or 0)
                + (stock.quantite_sous_unites or 0)
                + (stock.quantite_sous_sous_unites or 0)
            )
            if total <= 0:
                continue
            codes.append(
                f'{stock.code_suivi} '
                f'(U:{stock.quantite_unites or 0}, SU:{stock.quantite_sous_unites or 0}, SSU:{stock.quantite_sous_sous_unites or 0})'
            )
        codes_by_product[produit.id] = codes
    return codes_by_product

def get_products_stock_expiry_dates(produits):
    expiries_by_product = {}
    for produit in produits:
        expiries = []
        for stock in sorted(produit.stocks, key=lambda item: item.date_peremption):
            total = (
                (stock.quantite_unites or 0)
                + (stock.quantite_sous_unites or 0)
                + (stock.quantite_sous_sous_unites or 0)
            )
            label = stock.date_peremption.strftime('%d/%m/%Y') if stock.date_peremption else '-'
            expiries.append(
                f'{label} '
                f'(U:{stock.quantite_unites or 0}, SU:{stock.quantite_sous_unites or 0}, SSU:{stock.quantite_sous_sous_unites or 0}, Total:{total})'
            )
        expiries_by_product[produit.id] = expiries
    return expiries_by_product

def consume_product_stock_for_sale(produit, unite, quantite, numero_vente):
    field_by_unit = {
        'unite': 'quantite_unites',
        'sous_unite': 'quantite_sous_unites',
        'sous_sous_unite': 'quantite_sous_sous_unites'
    }
    field = field_by_unit.get(unite, 'quantite_unites')
    remaining = quantite
    stocks = sorted(produit.stocks, key=lambda item: item.date_peremption)
    for stock in stocks:
        available = float(getattr(stock, field) or 0)
        if available <= 0 or remaining <= 0:
            continue
        consumed = min(available, remaining)
        old_values = (stock.quantite_unites, stock.quantite_sous_unites, stock.quantite_sous_sous_unites)
        setattr(stock, field, int(round(available - consumed)))
        new_values = (stock.quantite_unites, stock.quantite_sous_unites, stock.quantite_sous_sous_unites)
        create_stock_modification(
            stock=stock,
            produit=produit,
            action='sortie',
            reason=f'Vente {numero_vente}',
            reason_id=None,
            old_values=old_values,
            new_values=new_values,
            old_qr_tire=stock.qr_tire,
            new_qr_tire=stock.qr_tire
        )
        remaining -= consumed
    return remaining <= 0.0001

def get_filtered_ventes(default_today=False):
    ventes = Vente.query.order_by(Vente.created_at.desc()).all()
    query = (request.args.get('q') or '').strip().lower()
    statut = (request.args.get('statut') or '').strip()
    mode_paiement = (request.args.get('mode_paiement') or '').strip()
    
    date_from_raw = (request.args.get('date_from') or '').strip()
    date_to_raw = (request.args.get('date_to') or '').strip()
    
    date_from = parse_date_filter(date_from_raw)
    date_to = parse_date_filter(date_to_raw)
    
    # If default_today is True and no specific filters are applied, default to today
    if default_today and not date_from and not date_to and not query:
        date_from = datetime.now().date()
        date_to = datetime.now().date()
        
    client_id = request.args.get('client_id', type=int)
    auteur_id = request.args.get('auteur_id', type=int)
    min_ttc = request.args.get('min_ttc', type=float)
    max_ttc = request.args.get('max_ttc', type=float)
    sort = (request.args.get('sort') or 'created_at').strip()
    direction = (request.args.get('direction') or 'desc').strip()

    if query:
        ventes = [
            vente for vente in ventes
            if query in ' '.join([
                vente.numero_vente or '',
                vente.client_matricule or '',
                vente.client_nom or '',
                vente.client_prenom or '',
                vente.client_email or '',
                vente.groupe_client_nom or '',
                vente.auteur_nom or '',
                vente.auteur_prenom or '',
                vente.mode_paiement or '',
                vente.statut or ''
            ]).lower()
        ]
    if statut:
        ventes = [vente for vente in ventes if vente.statut == statut]
    if mode_paiement:
        ventes = [vente for vente in ventes if vente.mode_paiement == mode_paiement]
    if date_from:
        ventes = [vente for vente in ventes if vente.created_at and vente.created_at.date() >= date_from]
    if date_to:
        ventes = [vente for vente in ventes if vente.created_at and vente.created_at.date() <= date_to]
    if client_id:
        ventes = [vente for vente in ventes if vente.client_id == client_id]
    if auteur_id:
        ventes = [vente for vente in ventes if vente.auteur_id == auteur_id]
    if min_ttc is not None:
        ventes = [vente for vente in ventes if vente.total_ttc >= min_ttc]
    if max_ttc is not None:
        ventes = [vente for vente in ventes if vente.total_ttc <= max_ttc]

    sorters = {
        'numero_vente': lambda vente: vente.numero_vente or '',
        'client': lambda vente: vente.client_label.lower(),
        'mode_paiement': lambda vente: vente.mode_paiement or '',
        'statut': lambda vente: vente.statut or '',
        'total_ht': lambda vente: vente.total_ht or 0,
        'total_tva': lambda vente: vente.total_tva or 0,
        'total_ttc': lambda vente: vente.total_ttc or 0,
        'auteur': lambda vente: f'{vente.auteur_prenom or ""} {vente.auteur_nom or ""}'.lower(),
        'created_at': lambda vente: vente.created_at or datetime.min
    }
    ventes.sort(key=sorters.get(sort, sorters['created_at']), reverse=direction != 'asc')
    return ventes

def compute_ventes_totals_reels(ventes):
    """Agrege benefice (marge coefficient) et TVA effective sur un ensemble de ventes,
    en une seule requete SQL plutot que ligne par ligne en Python."""
    numero_list = [vente.numero_vente for vente in ventes]
    if not numero_list:
        return {'tva_reelle': 0.0, 'benefice': 0.0}
    tva_expr = VenteLigne.total_ht * (VenteLigne.tva_pourcentage / 100.0)
    benefice_expr = db.func.greatest(VenteLigne.total_ttc - VenteLigne.total_ht - tva_expr, 0.0)
    tva_sum, benefice_sum = db.session.query(
        db.func.coalesce(db.func.sum(tva_expr), 0.0),
        db.func.coalesce(db.func.sum(benefice_expr), 0.0)
    ).filter(VenteLigne.numero_vente.in_(numero_list)).one()
    return {'tva_reelle': float(tva_sum or 0), 'benefice': float(benefice_sum or 0)}

def compute_tva_breakdown(ventes):
    """Regroupe les lignes de vente par taux de TVA - utile pour la declaration fiscale."""
    numero_list = [vente.numero_vente for vente in ventes]
    if not numero_list:
        return []
    tva_expr = VenteLigne.total_ht * (VenteLigne.tva_pourcentage / 100.0)
    rows = db.session.query(
        VenteLigne.tva_pourcentage,
        db.func.coalesce(db.func.sum(VenteLigne.total_ht), 0.0),
        db.func.coalesce(db.func.sum(tva_expr), 0.0)
    ).filter(
        VenteLigne.numero_vente.in_(numero_list)
    ).group_by(VenteLigne.tva_pourcentage).order_by(VenteLigne.tva_pourcentage).all()
    return [
        {'taux': float(taux or 0), 'ht': float(ht or 0), 'tva': float(tva or 0)}
        for taux, ht, tva in rows
    ]

def money_value(value):
    return float(value or 0)

def sale_employee_label(vente):
    return f'{vente.auteur_prenom or ""} {vente.auteur_nom or ""}'.strip() or (vente.auteur_email or 'Inconnu')

def sale_client_label(vente):
    return vente.client_label if vente.client_label != 'Client comptoir' else 'Client comptoir'

def add_stat_bucket(buckets, key, vente):
    bucket = buckets[key]
    bucket['label'] = key
    bucket['count'] += 1
    bucket['ht'] += money_value(vente.total_ht)
    bucket['benefice'] += money_value(vente.total_benefice)
    bucket['tva'] += money_value(vente.total_tva_reelle)
    bucket['ttc'] += money_value(vente.total_ttc)
    bucket['hors_solde'] += money_value(vente.montant_hors_solde)
    bucket['solde_client'] += money_value(vente.montant_solde_client)
    bucket['solde_groupe'] += money_value(vente.montant_solde_groupe)

def sorted_stat_rows(buckets, limit=None):
    rows = sorted(buckets.values(), key=lambda row: row['ttc'], reverse=True)
    return rows[:limit] if limit else rows

def build_vente_stats(ventes):
    base_bucket = lambda: {
        'label': '',
        'count': 0,
        'ht': 0.0,
        'benefice': 0.0,
        'tva': 0.0,
        'ttc': 0.0,
        'hors_solde': 0.0,
        'solde_client': 0.0,
        'solde_groupe': 0.0
    }
    product_bucket = lambda: {'label': '', 'count': 0, 'quantity': 0.0, 'ht': 0.0, 'benefice': 0.0, 'tva': 0.0, 'ttc': 0.0}
    employee_buckets = defaultdict(base_bucket)
    client_buckets = defaultdict(base_bucket)
    mode_buckets = defaultdict(base_bucket)
    status_buckets = defaultdict(base_bucket)
    daily_buckets = defaultdict(base_bucket)
    monthly_buckets = defaultdict(base_bucket)
    product_buckets = defaultdict(product_bucket)
    supplier_buckets = defaultdict(product_bucket)
    family_buckets = defaultdict(product_bucket)
    weekday_buckets = defaultdict(base_bucket)
    month_employee_buckets = defaultdict(base_bucket)
    latest_month = None

    totals = {
        'count': len(ventes),
        'ht': sum(money_value(vente.total_ht) for vente in ventes),
        'benefice': sum(money_value(vente.total_benefice) for vente in ventes),
        'tva': sum(money_value(vente.total_tva_reelle) for vente in ventes),
        'ttc': sum(money_value(vente.total_ttc) for vente in ventes),
        'hors_solde': sum(money_value(vente.montant_hors_solde) for vente in ventes),
        'solde_client': sum(money_value(vente.montant_solde_client) for vente in ventes),
        'solde_groupe': sum(money_value(vente.montant_solde_groupe) for vente in ventes),
        'monnaie': sum(money_value(vente.monnaie_rendue) for vente in ventes)
    }
    totals['panier_moyen'] = totals['ttc'] / totals['count'] if totals['count'] else 0

    for vente in ventes:
        employee = sale_employee_label(vente)
        client = sale_client_label(vente)
        mode = vente.mode_paiement or 'Non precise'
        status = vente.statut or 'Non precise'
        day = vente.created_at.strftime('%Y-%m-%d') if vente.created_at else 'Sans date'
        month = vente.created_at.strftime('%Y-%m') if vente.created_at else 'Sans date'
        weekday = vente.created_at.strftime('%A') if vente.created_at else 'Sans date'
        latest_month = max(latest_month, month) if latest_month else month

        for buckets, key in [
            (employee_buckets, employee),
            (client_buckets, client),
            (mode_buckets, mode),
            (status_buckets, status),
            (daily_buckets, day),
            (monthly_buckets, month),
            (weekday_buckets, weekday)
        ]:
            add_stat_bucket(buckets, key, vente)

        for ligne in vente.lignes:
            for buckets, key in [
                (product_buckets, ligne.produit_nom or 'Produit inconnu'),
                (supplier_buckets, ligne.produit_fournisseur or 'Fournisseur inconnu'),
                (family_buckets, ligne.produit_famille or 'Famille inconnue')
            ]:
                bucket = buckets[key]
                bucket['label'] = key
                bucket['count'] += 1
                bucket['quantity'] += money_value(ligne.quantite)
                bucket['ht'] += money_value(ligne.total_ht)
                bucket['benefice'] += money_value(ligne.benefice)
                bucket['tva'] += money_value(ligne.tva_reelle)
                bucket['ttc'] += money_value(ligne.total_ttc)

    if latest_month:
        for vente in ventes:
            if vente.created_at and vente.created_at.strftime('%Y-%m') == latest_month:
                add_stat_bucket(month_employee_buckets, sale_employee_label(vente), vente)

    # Fill gaps in daily stats with zeros
    if daily_buckets:
        all_dates = sorted(daily_buckets.keys())
        first_date = datetime.strptime(all_dates[0], '%Y-%m-%d').date()
        last_date = datetime.strptime(all_dates[-1], '%Y-%m-%d').date()
        
        curr = first_date
        while curr <= last_date:
            d_str = curr.strftime('%Y-%m-%d')
            if d_str not in daily_buckets:
                daily_buckets[d_str] = base_bucket()
                daily_buckets[d_str]['label'] = d_str
            curr += timedelta(days=1)

    daily_rows = sorted(daily_buckets.values(), key=lambda row: row['label'])
    monthly_rows = sorted(monthly_buckets.values(), key=lambda row: row['label'])
    employee_rows = sorted_stat_rows(employee_buckets)
    employee_month_rows = sorted_stat_rows(month_employee_buckets)

    return {
        'totals': totals,
        'daily': daily_rows,
        'monthly': monthly_rows,
        'employees': employee_rows,
        'employee_of_month': employee_month_rows[0] if employee_month_rows else None,
        'employee_of_month_period': latest_month,
        'clients': sorted_stat_rows(client_buckets),
        'products': sorted(product_buckets.values(), key=lambda row: row['ttc'], reverse=True),
        'suppliers': sorted(supplier_buckets.values(), key=lambda row: row['ttc'], reverse=True),
        'families': sorted(family_buckets.values(), key=lambda row: row['ttc'], reverse=True),
        'modes': sorted_stat_rows(mode_buckets),
        'statuses': sorted_stat_rows(status_buckets),
        'weekdays': sorted_stat_rows(weekday_buckets)
    }

@admin.route('/ventes')
@login_required
@permission_required('gestion_ventes')
def list_ventes():
    # Use request.args to get pagination parameters
    page = request.args.get('page', 1, type=int)
    per_page = request.args.get('per_page', 150, type=int)
    
    # Get all filtered ventes
    all_ventes = get_filtered_ventes(default_today=True)
    total_count = len(all_ventes)
    
    # Slice the list for the current page
    start = (page - 1) * per_page
    end = start + per_page
    ventes_paginated = all_ventes[start:end]
    
    # Calculate total pages
    total_pages = (total_count + per_page - 1) // per_page if total_count > 0 else 1
    
    # Pre-fill date filters for the UI if they were defaulted to today
    date_from = request.args.get('date_from')
    date_to = request.args.get('date_to')
    if not date_from and not date_to and not request.args.get('q'):
        today_str = datetime.now().strftime('%Y-%m-%d')
        date_from = today_str
        date_to = today_str
    
    totals_reels = compute_ventes_totals_reels(all_ventes)

    return render_template(
        'admin/ventes/list.html',
        ventes=ventes_paginated,
        all_ventes=all_ventes, # Keep this for totals in the UI if needed
        total_count=total_count,
        page=page,
        per_page=per_page,
        total_pages=total_pages,
        clients=Client.query.order_by(Client.nom.asc()).all(),
        users=User.query.filter_by(is_active=True).order_by(User.nom.asc()).all(),
        date_from_val=date_from,
        date_to_val=date_to,
        page_total_benefice=totals_reels['benefice'],
        page_total_tva_reelle=totals_reels['tva_reelle']
    )

@admin.route('/ventes/all')
@login_required
@permission_required('gestion_ventes')
def list_all_ventes():
    ventes = get_filtered_ventes()
    totals_reels = compute_ventes_totals_reels(ventes)
    return render_template(
        'admin/ventes/all.html',
        ventes=ventes,
        total_benefice=totals_reels['benefice'],
        total_tva_reelle=totals_reels['tva_reelle']
    )


@admin.route('/ventes/stats')
@login_required
@permission_required('stats_ventes')
def ventes_stats():
    ventes = get_filtered_ventes()
    stats = build_vente_stats(ventes)
    return render_template(
        'admin/ventes/stats.html', 
        ventes=ventes, 
        stats=stats, 
        export_query=request.args.to_dict(),
        clients=Client.query.order_by(Client.nom.asc()).all(),
        users=User.query.filter_by(is_active=True).order_by(User.nom.asc()).all()
    )

@admin.route('/ventes/stats/export/excel')
@login_required
@permission_required('stats_ventes')
def export_ventes_stats_excel():
    import io
    import pandas as pd
    from flask import send_file
    from openpyxl.chart import BarChart, LineChart, PieChart, Reference
    from openpyxl.styles import Alignment, Font, PatternFill, Border, Side

    ventes = get_filtered_ventes()
    stats = build_vente_stats(ventes)
    generated_at = datetime.now()
    output = io.BytesIO()

    def table_rows(rows, quantity=False):
        result = []
        for row in rows:
            data = {
                'Libelle': row['label'],
                'Ventes': row.get('count', 0),
                'HT': row.get('ht', 0),
                'Benefice': row.get('benefice', 0),
                'TVA': row.get('tva', 0),
                'TTC': row.get('ttc', 0)
            }
            if quantity:
                data['Quantite'] = row.get('quantity', 0)
            data['Hors solde'] = row.get('hors_solde', 0)
            data['Solde client'] = row.get('solde_client', 0)
            data['Solde groupe'] = row.get('solde_groupe', 0)
            return_row = data
            result.append(return_row)
        return result

    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        summary_rows = [
            {'Indicateur': 'Nombre de ventes', 'Valeur': stats['totals']['count']},
            {'Indicateur': 'Total HT', 'Valeur': stats['totals']['ht']},
            {'Indicateur': 'Total Benefice', 'Valeur': stats['totals']['benefice']},
            {'Indicateur': 'Total TVA effective', 'Valeur': stats['totals']['tva']},
            {'Indicateur': 'Total TTC', 'Valeur': stats['totals']['ttc']},
            {'Indicateur': 'Panier moyen', 'Valeur': stats['totals']['panier_moyen']},
            {'Indicateur': 'Hors solde', 'Valeur': stats['totals']['hors_solde']},
            {'Indicateur': 'Solde client', 'Valeur': stats['totals']['solde_client']},
            {'Indicateur': 'Solde groupe', 'Valeur': stats['totals']['solde_groupe']},
            {'Indicateur': 'Monnaie rendue', 'Valeur': stats['totals']['monnaie']},
            {'Indicateur': 'Employe du mois', 'Valeur': stats['employee_of_month']['label'] if stats['employee_of_month'] else '-'}
        ]
        pd.DataFrame(summary_rows).to_excel(writer, index=False, sheet_name='Synthese', startrow=4)
        pd.DataFrame(table_rows(stats['daily'])).to_excel(writer, index=False, sheet_name='Par jour')
        pd.DataFrame(table_rows(stats['monthly'])).to_excel(writer, index=False, sheet_name='Par mois')
        pd.DataFrame(table_rows(stats['employees'])).to_excel(writer, index=False, sheet_name='Employes')
        pd.DataFrame(table_rows(stats['clients'][:50])).to_excel(writer, index=False, sheet_name='Clients')
        pd.DataFrame(table_rows(stats['products'][:50], quantity=True)).to_excel(writer, index=False, sheet_name='Produits')
        pd.DataFrame(table_rows(stats['suppliers'][:50], quantity=True)).to_excel(writer, index=False, sheet_name='Fournisseurs')
        pd.DataFrame(table_rows(stats['modes'])).to_excel(writer, index=False, sheet_name='Paiements')

        for sheet_name, worksheet in writer.sheets.items():
            worksheet.freeze_panes = 'A2' if sheet_name != 'Synthese' else 'A6'
            worksheet.column_dimensions['A'].width = 28
            for column in ['B', 'C', 'D', 'E', 'F', 'G', 'H', 'I']:
                worksheet.column_dimensions[column].width = 15
            header_row = 5 if sheet_name == 'Synthese' else 1
            for cell in worksheet[header_row]:
                cell.font = Font(bold=True, color='FFFFFF')
                cell.fill = PatternFill(start_color='1F2937', end_color='1F2937', fill_type='solid')
                cell.alignment = Alignment(horizontal='center')
            for row in worksheet.iter_rows(min_row=header_row + 1):
                for cell in row:
                    cell.border = Border(bottom=Side(style='thin', color='D1D5DB'))

        synth = writer.sheets['Synthese']
        synth['A1'] = 'Statistiques des ventes - ReflexPharma'
        synth['A2'] = f'Date du tirage : {generated_at.strftime("%d/%m/%Y %H:%M")}'
        synth['A3'] = f'Tire par : {current_user.nom} {current_user.prenom}'
        synth['A4'] = f'Ventes filtrees : {len(ventes)}'

        def add_bar_chart(sheet_name, title, category_col=1, value_col=5, anchor='J2', max_row=12):
            ws = writer.sheets[sheet_name]
            last_row = min(ws.max_row, max_row + 1)
            if last_row < 2:
                return
            chart = BarChart()
            chart.title = title
            chart.y_axis.title = 'TTC'
            chart.x_axis.title = 'Libelle'
            data = Reference(ws, min_col=value_col, min_row=1, max_row=last_row)
            cats = Reference(ws, min_col=category_col, min_row=2, max_row=last_row)
            chart.add_data(data, titles_from_data=True)
            chart.set_categories(cats)
            chart.height = 8
            chart.width = 16
            ws.add_chart(chart, anchor)

        def add_line_chart(sheet_name, title, anchor='J2'):
            ws = writer.sheets[sheet_name]
            if ws.max_row < 2:
                return
            chart = LineChart()
            chart.title = title
            chart.y_axis.title = 'TTC'
            data = Reference(ws, min_col=5, min_row=1, max_row=ws.max_row)
            cats = Reference(ws, min_col=1, min_row=2, max_row=ws.max_row)
            chart.add_data(data, titles_from_data=True)
            chart.set_categories(cats)
            chart.height = 8
            chart.width = 16
            ws.add_chart(chart, anchor)

        def add_pie_chart(sheet_name, title, anchor='J2'):
            ws = writer.sheets[sheet_name]
            if ws.max_row < 2:
                return
            chart = PieChart()
            chart.title = title
            data = Reference(ws, min_col=5, min_row=1, max_row=min(ws.max_row, 8))
            cats = Reference(ws, min_col=1, min_row=2, max_row=min(ws.max_row, 8))
            chart.add_data(data, titles_from_data=True)
            chart.set_categories(cats)
            chart.height = 8
            chart.width = 12
            ws.add_chart(chart, anchor)

        add_line_chart('Par jour', 'Evolution quotidienne TTC')
        add_bar_chart('Employes', 'Top employes')
        add_bar_chart('Produits', 'Top produits')
        add_pie_chart('Paiements', 'Repartition des paiements')

    output.seek(0)
    filename = f'statistiques_ventes_{generated_at.strftime("%Y%m%d_%H%M")}.xlsx'
    return send_file(output, download_name=filename, as_attachment=True)

@admin.route('/ventes/stats/export/pdf')
@login_required
@permission_required('stats_ventes')
def export_ventes_stats_pdf():
    import io
    from flask import send_file
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle
    from reportlab.graphics.shapes import Drawing, String
    from reportlab.graphics.charts.barcharts import HorizontalBarChart
    from reportlab.graphics.charts.linecharts import HorizontalLineChart
    from reportlab.graphics.charts.piecharts import Pie

    ventes = get_filtered_ventes()
    stats = build_vente_stats(ventes)
    generated_at = datetime.now()
    output = io.BytesIO()
    doc = SimpleDocTemplate(output, pagesize=A4, topMargin=18, bottomMargin=18, leftMargin=18, rightMargin=18)
    styles = getSampleStyleSheet()
    styles.add(ParagraphStyle('Small', parent=styles['Normal'], fontSize=7, leading=9))
    elements = [
        Paragraph('Statistiques des ventes - ReflexPharma', styles['Title']),
        Paragraph(f'Date du tirage : {generated_at.strftime("%d/%m/%Y %H:%M")} | Tire par : {current_user.nom} {current_user.prenom} | Ventes : {len(ventes)}', styles['Small']),
        Spacer(1, 8)
    ]

    summary = [
        ['Ventes', stats['totals']['count'], 'Total TTC', f"{stats['totals']['ttc']:.2f}"],
        ['Panier moyen', f"{stats['totals']['panier_moyen']:.2f}", 'Benefice', f"{stats['totals']['benefice']:.2f}"],
        ['TVA effective', f"{stats['totals']['tva']:.2f}", 'Solde client', f"{stats['totals']['solde_client']:.2f}"],
        ['Hors solde', f"{stats['totals']['hors_solde']:.2f}", 'Solde groupe', f"{stats['totals']['solde_groupe']:.2f}"],
        ['Employe du mois', stats['employee_of_month']['label'] if stats['employee_of_month'] else '-', '', '']
    ]
    elements.append(Table(summary, colWidths=[90, 100, 90, 200], style=[
        ('GRID', (0, 0), (-1, -1), 0.25, colors.HexColor('#D1D5DB')),
        ('BACKGROUND', (0, 0), (-1, -1), colors.HexColor('#F9FAFB')),
        ('FONTSIZE', (0, 0), (-1, -1), 8),
    ]))
    elements.append(Spacer(1, 10))

    def add_horizontal_chart(title, rows, max_items=8):
        rows = rows[:max_items]
        if not rows:
            return
        drawing = Drawing(500, 185)
        drawing.add(String(0, 170, title, fontSize=10, fillColor=colors.HexColor('#1F2937')))
        chart = HorizontalBarChart()
        chart.x = 130
        chart.y = 25
        chart.width = 330
        chart.height = 125
        chart.data = [[row['ttc'] for row in rows]]
        chart.categoryAxis.categoryNames = [row['label'][:24] for row in rows]
        chart.categoryAxis.labels.fontSize = 6
        chart.valueAxis.valueMin = 0
        chart.bars[0].fillColor = colors.HexColor('#198754')
        drawing.add(chart)
        elements.append(drawing)
        elements.append(Spacer(1, 8))

    def add_line_chart(title, rows, max_items=18):
        rows = rows[-max_items:]
        if not rows:
            return
        drawing = Drawing(500, 175)
        drawing.add(String(0, 160, title, fontSize=10, fillColor=colors.HexColor('#1F2937')))
        chart = HorizontalLineChart()
        chart.x = 35
        chart.y = 35
        chart.height = 105
        chart.width = 420
        chart.data = [[row['ttc'] for row in rows]]
        chart.categoryAxis.categoryNames = [row['label'] for row in rows]
        chart.categoryAxis.labels.angle = 35
        chart.categoryAxis.labels.fontSize = 5
        chart.valueAxis.valueMin = 0
        chart.lines[0].strokeColor = colors.HexColor('#0D6EFD')
        drawing.add(chart)
        elements.append(drawing)
        elements.append(Spacer(1, 8))

    def add_pie_chart(title, rows, max_items=6):
        rows = rows[:max_items]
        if not rows:
            return
        drawing = Drawing(500, 170)
        drawing.add(String(0, 155, title, fontSize=10, fillColor=colors.HexColor('#1F2937')))
        pie = Pie()
        pie.x = 20
        pie.y = 20
        pie.width = 120
        pie.height = 120
        pie.data = [row['ttc'] for row in rows]
        pie.labels = [row['label'][:16] for row in rows]
        drawing.add(pie)
        y = 130
        for row in rows:
            drawing.add(String(175, y, f"{row['label'][:36]}: {row['ttc']:.2f}", fontSize=7, fillColor=colors.HexColor('#374151')))
            y -= 16
        elements.append(drawing)
        elements.append(Spacer(1, 8))

    add_line_chart('Evolution quotidienne TTC', stats['daily'])
    add_horizontal_chart('Top employes par TTC', stats['employees'])
    add_horizontal_chart('Top produits par TTC', stats['products'])
    add_horizontal_chart('Top clients par TTC', stats['clients'])
    add_pie_chart('Repartition par mode de paiement', stats['modes'])

    def add_table(title, rows, quantity=False):
        elements.append(Paragraph(title, styles['Heading3']))
        headers = ['Libelle', 'Ventes', 'TTC'] if not quantity else ['Libelle', 'Lignes', 'Qte', 'TTC']
        data = [headers]
        for row in rows[:10]:
            if quantity:
                data.append([row['label'][:42], row.get('count', 0), f"{row.get('quantity', 0):.2f}", f"{row['ttc']:.2f}"])
            else:
                data.append([row['label'][:48], row.get('count', 0), f"{row['ttc']:.2f}"])
        elements.append(Table(data, repeatRows=1, style=[
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#374151')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('GRID', (0, 0), (-1, -1), 0.25, colors.HexColor('#D1D5DB')),
            ('FONTSIZE', (0, 0), (-1, -1), 7),
        ]))
        elements.append(Spacer(1, 8))

    add_table('Employes', stats['employees'])
    add_table('Clients', stats['clients'])
    add_table('Produits', stats['products'], quantity=True)
    add_table('Fournisseurs', stats['suppliers'], quantity=True)

    doc.build(elements)
    output.seek(0)
    filename = f'statistiques_ventes_{generated_at.strftime("%Y%m%d_%H%M")}.pdf'
    return send_file(output, download_name=filename, as_attachment=True)

@admin.route('/ventes/scan-lookup', methods=['GET'])
@login_required
@permission_required('gestion_ventes')
def vente_scan_lookup():
    """Resout un code scanne (QR d'un lot, ou code produit) vers un produit a ajouter au panier de vente."""
    code = (request.args.get('code') or '').strip()
    if not code:
        return {'success': False, 'message': 'Code vide.'}, 400

    stock = Stock.query.filter_by(code_suivi=code).first()
    produit = stock.produit if stock else None

    if not produit:
        produit = Produit.query.filter_by(code_produit=code).first()

    if not produit:
        stock = Stock.query.filter(Stock.code_suivi.ilike(f'%{code}%')).first()
        produit = stock.produit if stock else None

    if not produit:
        return {'success': False, 'message': f"Aucun produit trouve pour le code '{code}'."}, 404

    return {
        'success': True,
        'produit_id': produit.id,
        'nom': produit.nom,
        'code_produit': produit.code_produit,
        'code_suivi': stock.code_suivi if stock else None
    }

@admin.route('/ventes/create', methods=['GET', 'POST'])
@login_required
@permission_required('gestion_ventes')
def create_vente():
    clients = Client.query.order_by(Client.nom.asc(), Client.prenom.asc()).all()
    produits = Produit.query.order_by(Produit.nom.asc()).all()
    stock_totals = get_products_stock_totals(produits)
    stock_tracking_codes = get_products_stock_tracking_codes(produits)
    stock_expiry_dates = get_products_stock_expiry_dates(produits)
    if request.method == 'POST':
        def form_amount(name):
            try:
                return max(float(request.form.get(name) or 0), 0)
            except ValueError:
                return 0

        validation_password = request.form.get('validation_password') or ''
        if not current_user.check_password(validation_password):
            flash('Mot de passe incorrect. La vente n a pas ete validee.', 'danger')
            return redirect(url_for('admin.create_vente'))

        client = None
        client_id = request.form.get('client_id')
        if client_id:
            client = Client.query.get(client_id)

        numero_vente = generate_numero_vente()
        vente = Vente(
            numero_vente=numero_vente,
            statut=request.form.get('statut') or 'validee',
            mode_paiement=request.form.get('mode_paiement') or 'especes',
            note=request.form.get('note'),
            client_id=None,
            client_matricule=client.matricule if client else None,
            client_nom=client.nom if client else None,
            client_prenom=client.prenom if client else None,
            client_email=client.email if client else None,
            groupe_client_id=None,
            groupe_client_nom=client.groupe.nom if client and client.groupe else None,
            groupe_absorption_pourcentage=client.groupe.pourcentage_absorption if client and client.groupe else 0,
            auteur_id=None,
            auteur_nom=current_user.nom,
            auteur_prenom=current_user.prenom,
            auteur_email=current_user.email
        )
        db.session.add(vente)

        produit_ids = request.form.getlist('produit_id[]')
        unites = request.form.getlist('unite[]')
        quantites = request.form.getlist('quantite[]')
        total_ht = 0
        total_tva = 0
        total_ttc = 0
        lignes_count = 0
        requested_quantities = {}

        for index, produit_id in enumerate(produit_ids):
            if not produit_id:
                continue
            produit = Produit.query.get(produit_id)
            if not produit:
                continue
            stock_summary = stock_totals.get(produit.id, get_product_stock_summary(produit))
            unite = normalize_product_unit(produit, unites[index] if index < len(unites) else 'unite', stock_summary)
            try:
                quantite = float(quantites[index] if index < len(quantites) else 1)
            except ValueError:
                quantite = 1
            if quantite <= 0:
                continue
            if not quantite.is_integer():
                db.session.rollback()
                flash(f'La quantite vendue pour {produit.nom} doit etre un nombre entier.', 'danger')
                return redirect(url_for('admin.create_vente'))

            available_quantity = float(stock_summary.get(unite, 0) or 0)
            request_key = (produit.id, unite)
            requested_quantity = requested_quantities.get(request_key, 0) + quantite
            if available_quantity <= 0:
                db.session.rollback()
                flash(f'{produit.nom} n a pas de stock disponible pour cette unite.', 'danger')
                return redirect(url_for('admin.create_vente'))
            if requested_quantity > available_quantity:
                db.session.rollback()
                flash(f'Stock insuffisant pour {produit.nom}. Disponible : {available_quantity:g}, demande : {requested_quantity:g}.', 'danger')
                return redirect(url_for('admin.create_vente'))
            requested_quantities[request_key] = requested_quantity

            prix_ht, prix_ttc = get_product_unit_price(produit, unite)
            ligne_total_ht = prix_ht * quantite
            ligne_total_ttc = prix_ttc * quantite
            ligne_total_tva = max(ligne_total_ttc - ligne_total_ht, 0)
            db.session.add(VenteLigne(
                numero_vente=numero_vente,
                produit_id=None,
                produit_code=produit.code_produit,
                produit_nom=produit.nom,
                produit_fournisseur=produit.fournisseur.nom if produit.fournisseur else None,
                produit_groupe_fournisseur=produit.fournisseur.groupe.nom if produit.fournisseur and produit.fournisseur.groupe else None,
                produit_rayon=produit.rayon.nom if produit.rayon else None,
                produit_famille=produit.famille.nom if produit.famille else None,
                produit_section=produit.section.nom if produit.section else None,
                produit_conditionnement=produit.conditionnement or 1,
                produit_codes_suivi=' | '.join(stock_tracking_codes.get(produit.id, [])),
                produit_dates_peremption=' | '.join(stock_expiry_dates.get(produit.id, [])),
                stock_unite_avant=float(stock_summary.get('unite', 0) or 0),
                stock_sous_unite_avant=float(stock_summary.get('sous_unite', 0) or 0),
                stock_sous_sous_unite_avant=float(stock_summary.get('sous_sous_unite', 0) or 0),
                unite=unite,
                quantite=quantite,
                prix_unitaire_ht=prix_ht,
                prix_unitaire_ttc=prix_ttc,
                tva_pourcentage=produit.effectif_tva or 0,
                total_ht=ligne_total_ht,
                total_tva=ligne_total_tva,
                total_ttc=ligne_total_ttc
            ))
            if not consume_product_stock_for_sale(produit, unite, quantite, numero_vente):
                db.session.rollback()
                flash(f'Impossible de sortir le stock pour {produit.nom}.', 'danger')
                return redirect(url_for('admin.create_vente'))
            total_ht += ligne_total_ht
            total_tva += ligne_total_tva
            total_ttc += ligne_total_ttc
            lignes_count += 1

        if lignes_count == 0:
            db.session.rollback()
            flash('Ajoutez au moins un produit valide a la vente.', 'danger')
            return redirect(url_for('admin.create_vente'))

        montant_hors_solde = form_amount('montant_hors_solde')
        montant_solde_client = form_amount('montant_solde_client')
        use_group_balance = request.form.get('use_group_balance') == '1'
        montant_solde_groupe = 0
        if use_group_balance and client and client.groupe:
            absorption = min(max(float(client.groupe.pourcentage_absorption or 0), 0), 100)
            montant_solde_groupe = total_ttc * (absorption / 100)
        montant_recu = form_amount('montant_recu')
        montant_couvert = montant_hors_solde + montant_solde_client + montant_solde_groupe
        if montant_recu + 0.0001 < montant_hors_solde:
            db.session.rollback()
            flash(f'Montant recu insuffisant pour le paiement hors solde. Recu : {montant_recu:.2f}, hors solde : {montant_hors_solde:.2f}.', 'danger')
            return redirect(url_for('admin.create_vente'))
        if montant_couvert + 0.0001 < total_ttc:
            db.session.rollback()
            flash(f'Paiement insuffisant. Total TTC : {total_ttc:.2f}, montant couvert : {montant_couvert:.2f}.', 'danger')
            return redirect(url_for('admin.create_vente'))
        if montant_couvert - total_ttc > 0.01:
            db.session.rollback()
            flash(f'Paiement trop eleve. Ajustez les montants appliques a la vente. Total TTC : {total_ttc:.2f}, montant couvert : {montant_couvert:.2f}.', 'danger')
            return redirect(url_for('admin.create_vente'))

        client_balance_before = float(client.solde or 0) if client else 0
        group_balance_before = float(client.groupe.solde or 0) if client and client.groupe else 0
        if montant_solde_client > 0 and not client:
            db.session.rollback()
            flash('Selectionnez un client pour debiter un solde client.', 'danger')
            return redirect(url_for('admin.create_vente'))
        if montant_solde_client > client_balance_before:
            db.session.rollback()
            flash(f'Solde client insuffisant. Disponible : {client_balance_before:.2f}.', 'danger')
            return redirect(url_for('admin.create_vente'))
        if montant_solde_groupe > 0 and not (client and client.groupe):
            db.session.rollback()
            flash('Selectionnez un client avec groupe pour debiter un solde groupe.', 'danger')
            return redirect(url_for('admin.create_vente'))
        if montant_solde_groupe > group_balance_before:
            db.session.rollback()
            flash(f'Solde groupe insuffisant. Disponible : {group_balance_before:.2f}.', 'danger')
            return redirect(url_for('admin.create_vente'))

        if client and montant_solde_client > 0:
            client.solde = client_balance_before - montant_solde_client
        if client and client.groupe and montant_solde_groupe > 0:
            client.groupe.solde = group_balance_before - montant_solde_groupe

        vente.total_ht = total_ht
        vente.total_tva = total_tva
        vente.total_ttc = total_ttc
        vente.montant_recu = montant_recu
        vente.montant_hors_solde = montant_hors_solde
        vente.montant_solde_client = montant_solde_client
        vente.montant_solde_groupe = montant_solde_groupe
        vente.monnaie_rendue = max(montant_recu - montant_hors_solde, 0)
        vente.solde_client_avant = client_balance_before
        vente.solde_client_apres = client.solde if client else 0
        vente.solde_groupe_avant = group_balance_before
        vente.solde_groupe_apres = client.groupe.solde if client and client.groupe else 0
        db.session.commit()
        flash(f'Vente {numero_vente} enregistrée avec succès.', 'success')
        return redirect(url_for('admin.create_vente'))

    return render_template(
        'admin/ventes/form.html',
        clients=clients,
        produits=produits,
        stock_totals=stock_totals,
        stock_tracking_codes=stock_tracking_codes,
        stock_expiry_dates=stock_expiry_dates,
        suggested_numero=generate_numero_vente(),
        pharmacy_name=Setting.get_value('pharmacy_name', 'REFLEXPHARMA'),
        auto_print_enabled=Setting.get_value('auto_print_enabled', 'true') == 'true'
    )

@admin.route('/ventes/validate-password', methods=['POST'])
@login_required
@permission_required('gestion_ventes')
def validate_vente_password():
    password = request.form.get('validation_password') or ''
    return {'valid': current_user.check_password(password)}

@admin.route('/ventes/<int:id>')
@login_required
@permission_required('gestion_ventes')
def detail_vente(id):
    vente = Vente.query.get_or_404(id)
    return render_template(
        'admin/ventes/detail.html',
        vente=vente,
        pharmacy_name=Setting.get_value('pharmacy_name', 'REFLEXPHARMA'),
        auto_print_enabled=Setting.get_value('auto_print_enabled', 'true') == 'true'
    )

@admin.route('/clients/<int:id>/achats')
@login_required
@permission_required('gestion_clients')
def client_purchase_history(id):
    client = Client.query.get_or_404(id)
    ventes = Vente.query.filter_by(client_matricule=client.matricule).order_by(Vente.created_at.desc()).all()
    return render_template('admin/clients/purchase_history.html', client=client, ventes=ventes)

@admin.route('/clients/<int:id>/achats/export/excel')
@login_required
@permission_required('gestion_clients')
def export_client_purchase_history_excel(id):
    import io
    import pandas as pd
    from flask import send_file
    from openpyxl.styles import Alignment, Font, PatternFill, Border, Side

    client = Client.query.get_or_404(id)
    ventes = Vente.query.filter_by(client_matricule=client.matricule).order_by(Vente.created_at.desc()).all()
    generated_at = datetime.now()

    rows = []
    for vente in ventes:
        produits = ', '.join(f"{l.produit_nom} x{l.quantite:g}{l.unite}" for l in vente.lignes)
        rows.append({
            'Date': vente.created_at.strftime('%d/%m/%Y %H:%M') if vente.created_at else '',
            'Vente': vente.numero_vente,
            'Statut': vente.statut,
            'Produits': produits,
            'Total TTC': vente.total_ttc or 0,
            'Payé (hors solde)': vente.montant_hors_solde or 0,
            'Prélevé solde client': vente.montant_solde_client or 0,
            'Prélevé solde groupe': vente.montant_solde_groupe or 0,
            'Mode de paiement': vente.mode_paiement,
        })

    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        pd.DataFrame(rows).to_excel(writer, index=False, sheet_name='Achats', startrow=5)
        worksheet = writer.sheets['Achats']
        worksheet['A1'] = f'Historique des achats - {client.nom_complet}'
        worksheet['A2'] = f'Matricule : {client.matricule} | Groupe : {client.groupe.nom if client.groupe else "-"}'
        worksheet['A3'] = f'Date du tirage : {generated_at.strftime("%d/%m/%Y %H:%M")} | Tire par : {current_user.nom} {current_user.prenom}'
        worksheet['A4'] = f'Achats : {len(rows)} | Total TTC : {sum(v.total_ttc or 0 for v in ventes):.2f} | Solde actuel : {client.solde:.2f}'
        header_fill = PatternFill(start_color='1F2937', end_color='1F2937', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF')
        thin_border = Border(bottom=Side(style='thin', color='D1D5DB'))
        for cell in worksheet[6]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        for row in worksheet.iter_rows(min_row=7, max_row=worksheet.max_row):
            for cell in row:
                cell.border = thin_border
                cell.alignment = Alignment(vertical='top', wrap_text=True)
        widths = [16, 16, 12, 45, 12, 14, 16, 16, 16]
        for index, width in enumerate(widths, start=1):
            worksheet.column_dimensions[chr(64 + index)].width = width
        worksheet.freeze_panes = 'A7'

    output.seek(0)
    filename = f"achats_{client.matricule}_{generated_at.strftime('%Y%m%d_%H%M')}.xlsx"
    return send_file(output, download_name=filename, as_attachment=True)

@admin.route('/clients/<int:id>/achats/export/pdf')
@login_required
@permission_required('gestion_clients')
def export_client_purchase_history_pdf(id):
    import io
    from flask import send_file
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
    from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle
    from xml.sax.saxutils import escape

    client = Client.query.get_or_404(id)
    ventes = Vente.query.filter_by(client_matricule=client.matricule).order_by(Vente.created_at.desc()).all()
    generated_at = datetime.now()
    total_ttc = sum(v.total_ttc or 0 for v in ventes)

    output = io.BytesIO()
    doc = SimpleDocTemplate(output, pagesize=A4, topMargin=30, bottomMargin=30, leftMargin=30, rightMargin=30)
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle('PHTitle', parent=styles['Heading1'], fontSize=18, leading=22, textColor=colors.HexColor('#2c3e50'), alignment=1, spaceAfter=6)
    meta_style = ParagraphStyle('PHMeta', parent=styles['Normal'], fontSize=9, leading=12, alignment=1)
    cell_style = ParagraphStyle('PHCell', parent=styles['Normal'], fontSize=7, leading=8.5)

    elements = [
        Paragraph('Historique des achats', title_style),
        Paragraph(f"{escape(client.nom_complet)} &middot; Matricule {escape(client.matricule)}", meta_style),
        Paragraph(f"Édité le {generated_at.strftime('%d/%m/%Y à %H:%M')}", meta_style),
        Spacer(1, 12),
    ]

    summary_data = [[
        'Nombre d\'achats', 'Total TTC cumulé', 'Solde actuel du client'
    ], [
        str(len(ventes)),
        f'{total_ttc:.2f}',
        f'{client.solde:.2f}'
    ]]
    summary_table = Table(summary_data, colWidths=[178, 178, 178])
    summary_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2c3e50')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTNAME', (0, 1), (-1, 1), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 9),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('TOPPADDING', (0, 0), (-1, -1), 5),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 5),
    ]))
    elements.append(summary_table)
    elements.append(Spacer(1, 14))

    data = [['Date', 'Vente', 'Produits achetés', 'TTC', 'Payé', 'Solde client', 'Solde groupe']]
    for vente in ventes:
        produits_str = '\n'.join(f"{l.produit_nom} x{l.quantite:g}{l.unite}" for l in vente.lignes)
        data.append([
            vente.created_at.strftime('%d/%m/%Y\n%H:%M') if vente.created_at else '-',
            Paragraph(escape(vente.numero_vente), cell_style),
            Paragraph(escape(produits_str), cell_style),
            f'{vente.total_ttc or 0:.2f}',
            f'{vente.montant_hors_solde or 0:.2f}',
            f'{vente.montant_solde_client or 0:.2f}',
            f'{vente.montant_solde_groupe or 0:.2f}',
        ])

    table = Table(data, repeatRows=1, colWidths=[60, 75, 195, 50, 45, 55, 55])
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1abc9c')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 7.5),
        ('FONTSIZE', (0, 1), (-1, -1), 6.5),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('ALIGN', (3, 0), (-1, -1), 'RIGHT'),
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F0FBF8')]),
        ('LEFTPADDING', (0, 0), (-1, -1), 3),
        ('RIGHTPADDING', (0, 0), (-1, -1), 3),
        ('TOPPADDING', (0, 0), (-1, -1), 2.5),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 2.5),
    ]))
    elements.append(table)

    if not ventes:
        elements.append(Spacer(1, 10))
        elements.append(Paragraph('Aucun achat enregistré pour ce client.', styles['Normal']))

    doc.build(elements)
    output.seek(0)
    filename = f"achats_{client.matricule}_{generated_at.strftime('%Y%m%d_%H%M')}.pdf"
    return send_file(output, download_name=filename, as_attachment=True)

@admin.route('/clients/historique')
@login_required
@permission_required('historique_clients')
def list_client_modification_logs():
    logs = ClientModificationLog.query.order_by(ClientModificationLog.created_at.desc()).all()
    parsed_logs = [
        {
            'log': log,
            'old_values': parse_log_values(log.old_values),
            'new_values': parse_log_values(log.new_values)
        }
        for log in logs
    ]
    return render_template('admin/clients/history.html', parsed_logs=parsed_logs)

def get_filtered_client_modification_logs():
    logs = ClientModificationLog.query.order_by(ClientModificationLog.created_at.desc()).all()
    query = (request.args.get('q') or '').strip().lower()
    entity_type = (request.args.get('type') or '').strip()
    action = (request.args.get('action') or '').strip()
    direction = (request.args.get('direction') or 'desc').strip()
    date_from_raw = (request.args.get('date_from') or '').strip()
    date_to_raw = (request.args.get('date_to') or '').strip()
    try:
        date_from = datetime.strptime(date_from_raw, '%Y-%m-%d').date() if date_from_raw else None
    except ValueError:
        date_from = None
    try:
        date_to = datetime.strptime(date_to_raw, '%Y-%m-%d').date() if date_to_raw else None
    except ValueError:
        date_to = None

    if entity_type:
        logs = [log for log in logs if log.entity_type == entity_type]
    if action:
        logs = [log for log in logs if log.action == action]
    if date_from:
        logs = [log for log in logs if log.created_at and log.created_at.date() >= date_from]
    if date_to:
        logs = [log for log in logs if log.created_at and log.created_at.date() <= date_to]
    if query:
        logs = [
            log for log in logs
            if query in ' '.join([
                log.created_at.strftime('%d/%m/%Y %H:%M') if log.created_at else '',
                log.entity_type or '',
                log.action or '',
                log.reference or '',
                log.label or '',
                log.user_prenom or '',
                log.user_nom or '',
                log.user_email or '',
                log.old_values or '',
                log.new_values or ''
            ]).lower()
        ]
    logs.sort(key=lambda log: log.created_at or datetime.min, reverse=direction != 'asc')
    return logs

@admin.route('/clients/historique/export/excel')
@login_required
@permission_required('historique_clients')
def export_client_modification_logs_excel():
    import io
    import pandas as pd
    from flask import send_file
    from openpyxl.styles import Alignment, Font, PatternFill, Border, Side

    logs = get_filtered_client_modification_logs()
    generated_at = datetime.now()
    rows = []
    for log in logs:
        rows.append({
            'Date': log.created_at.strftime('%d/%m/%Y %H:%M') if log.created_at else '-',
            'Type': 'Client' if log.entity_type == 'client' else 'Groupe client',
            'Action': log.action,
            'Reference': log.reference,
            'Libelle': log.label,
            'Avant': log.old_values or '',
            'Apres': log.new_values or '',
            'Fait par': f'{log.user_prenom} {log.user_nom}',
            'Email': log.user_email,
            'Raison': log.reason or ''
        })

    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        pd.DataFrame(rows).to_excel(writer, index=False, sheet_name='Historique clients', startrow=4)
        worksheet = writer.sheets['Historique clients']
        worksheet['A1'] = 'Historique clients et groupes clients - ReflexPharma'
        worksheet['A2'] = f'Date du tirage : {generated_at.strftime("%d/%m/%Y %H:%M")}'
        worksheet['A3'] = f'Tire par : {current_user.nom} {current_user.prenom}'
        worksheet['A4'] = f'Lignes exportees : {len(rows)}'

        header_fill = PatternFill(start_color='1F2937', end_color='1F2937', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF')
        thin_border = Border(bottom=Side(style='thin', color='D1D5DB'))
        for cell in worksheet[5]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        for row in worksheet.iter_rows(min_row=6, max_row=worksheet.max_row):
            for cell in row:
                cell.border = thin_border
                cell.alignment = Alignment(vertical='top', wrap_text=True)
        widths = [18, 16, 12, 18, 28, 42, 42, 22, 28, 28]
        for index, width in enumerate(widths, start=1):
            worksheet.column_dimensions[chr(64 + index)].width = width
        worksheet.freeze_panes = 'A6'

    output.seek(0)
    filename = f'historique_clients_{generated_at.strftime("%Y%m%d_%H%M")}.xlsx'
    return send_file(output, download_name=filename, as_attachment=True)

@admin.route('/clients/historique/export/pdf')
@login_required
@permission_required('historique_clients')
def export_client_modification_logs_pdf():
    import io
    from flask import send_file
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
    from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle
    from xml.sax.saxutils import escape

    logs = get_filtered_client_modification_logs()
    generated_at = datetime.now()
    output = io.BytesIO()
    doc = SimpleDocTemplate(output, pagesize=landscape(A4), topMargin=12, bottomMargin=12, leftMargin=12, rightMargin=12)
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle('ClientHistoryTitle', parent=styles['Title'], fontSize=12, leading=14)
    meta_style = ParagraphStyle('ClientHistoryMeta', parent=styles['Normal'], fontSize=7, leading=9)
    cell_style = ParagraphStyle('ClientHistoryCell', parent=styles['Normal'], fontSize=6, leading=7)

    elements = [
        Paragraph('Historique clients et groupes clients - ReflexPharma', title_style),
        Paragraph(f'Date du tirage : {generated_at.strftime("%d/%m/%Y %H:%M")} | Tire par : {current_user.nom} {current_user.prenom} | Lignes : {len(logs)}', meta_style),
        Spacer(1, 6)
    ]
    data = [['Date', 'Type', 'Action', 'Reference', 'Libelle', 'Avant', 'Apres', 'Fait par']]
    for log in logs:
        data.append([
            log.created_at.strftime('%d/%m/%Y %H:%M') if log.created_at else '-',
            'Client' if log.entity_type == 'client' else 'Groupe',
            log.action,
            Paragraph(escape(str(log.reference or '-')), cell_style),
            Paragraph(escape(str(log.label or '-')), cell_style),
            Paragraph(escape(str(log.old_values or '-'))[:700], cell_style),
            Paragraph(escape(str(log.new_values or '-'))[:700], cell_style),
            Paragraph(escape(f'{log.user_prenom} {log.user_nom}'), cell_style)
        ])

    table = Table(data, repeatRows=1, colWidths=[58, 52, 42, 70, 90, 180, 180, 80])
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1F2937')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 6),
        ('LEADING', (0, 0), (-1, -1), 7),
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ('GRID', (0, 0), (-1, -1), 0.25, colors.HexColor('#D1D5DB')),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F9FAFB')]),
        ('TOPPADDING', (0, 0), (-1, -1), 2),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 2),
        ('LEFTPADDING', (0, 0), (-1, -1), 2),
        ('RIGHTPADDING', (0, 0), (-1, -1), 2),
    ]))
    elements.append(table)
    doc.build(elements)
    output.seek(0)
    filename = f'historique_clients_{generated_at.strftime("%Y%m%d_%H%M")}.pdf'
    return send_file(output, download_name=filename, as_attachment=True)

# --- GESTION DES FOURNISSEURS ---
@admin.route('/fournisseurs')
@login_required
@permission_required('gestion_fournisseurs')
def list_fournisseurs():
    fournisseurs = Fournisseur.query.all()
    groupes = GroupeFournisseur.query.all()
    return render_template('admin/fournisseurs/list.html', fournisseurs=fournisseurs, groupes=groupes)

@admin.route('/fournisseurs/bulk-delete', methods=['POST'])
@login_required
@permission_required('gestion_fournisseurs')
def bulk_delete_fournisseurs():
    ids = request.form.getlist('ids[]')
    if not ids:
        flash("Aucun fournisseur sélectionné.", "warning")
        return redirect(url_for('admin.list_fournisseurs'))
    
    deleted_count = 0
    for f_id in ids:
        f = Fournisseur.query.get(f_id)
        if f:
            db.session.delete(f)
            deleted_count += 1
            
    db.session.commit()
    flash(f'{deleted_count} fournisseur(s) supprimé(s).', 'success')
    return redirect(url_for('admin.list_fournisseurs'))

@admin.route('/fournisseurs/create', methods=['GET', 'POST'])
@login_required
@permission_required('gestion_fournisseurs')
def create_fournisseur():
    groupes = GroupeFournisseur.query.all()
    if request.method == 'POST':
        nom = request.form.get('nom')
        prefixe_propose = request.form.get('prefixe').upper()
        
        import re
        base_prefix = re.sub(r'[0-9]+$', '', prefixe_propose)
        if not base_prefix: base_prefix = prefixe_propose
        
        prefixe = prefixe_propose
        counter = 1
        match = re.search(r'([0-9]+)$', prefixe_propose)
        if match:
            counter = int(match.group(1))
            base_prefix = prefixe_propose[:match.start()]

        while Fournisseur.query.filter_by(prefixe=prefixe).first():
            prefixe = f"{base_prefix}{counter}"
            counter += 1
        
        coeff = request.form.get('coefficient')
        tva = request.form.get('tva')
        groupe_id = request.form.get('groupe_id')

        new_fournisseur = Fournisseur(
            nom=nom,
            site_web=request.form.get('site_web'),
            contact=request.form.get('contact'),
            prefixe=prefixe,
            coefficient=float(coeff) if coeff else None,
            tva=float(tva) if tva else None,
            groupe_id=int(groupe_id) if groupe_id else None
        )
        db.session.add(new_fournisseur)
        db.session.commit()
        flash(f'Fournisseur ajouté avec succès. Préfixe utilisé : {prefixe}', 'success')
        return redirect(url_for('admin.list_fournisseurs'))
    return render_template('admin/fournisseurs/form.html', title="Ajouter un Fournisseur", groupes=groupes)

@admin.route('/fournisseurs/edit/<int:id>', methods=['GET', 'POST'])
@login_required
@permission_required('gestion_fournisseurs')
def edit_fournisseur(id):
    fournisseur = Fournisseur.query.get_or_404(id)
    groupes = GroupeFournisseur.query.all()
    if request.method == 'POST':
        prefixe = request.form.get('prefixe')
        existing = Fournisseur.query.filter_by(prefixe=prefixe).first()
        if existing and existing.id != id:
            flash('Ce préfixe est déjà utilisé.', 'danger')
            return redirect(url_for('admin.edit_fournisseur', id=id))
            
        fournisseur.nom = request.form.get('nom')
        fournisseur.site_web = request.form.get('site_web')
        fournisseur.contact = request.form.get('contact')
        fournisseur.prefixe = prefixe
        
        coeff = request.form.get('coefficient')
        fournisseur.coefficient = float(coeff) if coeff else None
        tva = request.form.get('tva')
        fournisseur.tva = float(tva) if tva else None
        groupe_id = request.form.get('groupe_id')
        fournisseur.groupe_id = int(groupe_id) if groupe_id else None
        
        db.session.commit()
        flash('Fournisseur mis à jour avec succès.', 'success')
        return redirect(url_for('admin.list_fournisseurs'))
    return render_template('admin/fournisseurs/form.html', fournisseur=fournisseur, title="Modifier le Fournisseur", groupes=groupes)

@admin.route('/fournisseurs/delete/<int:id>', methods=['POST'])
@login_required
@permission_required('gestion_fournisseurs')
def delete_fournisseur(id):
    fournisseur = Fournisseur.query.get_or_404(id)
    db.session.delete(fournisseur)
    db.session.commit()
    flash('Fournisseur supprimé.', 'success')
    return redirect(url_for('admin.list_fournisseurs'))

# --- GESTION DES GROUPES DE FOURNISSEURS ---
@admin.route('/fournisseurs/groupes')
@login_required
@permission_required('gestion_groupes_fournisseurs')
def list_groupes_fournisseurs():
    groupes = GroupeFournisseur.query.all()
    return render_template('admin/fournisseurs/groupes_list.html', groupes=groupes)

@admin.route('/fournisseurs/groupes/create', methods=['GET', 'POST'])
@login_required
@permission_required('gestion_groupes_fournisseurs')
def create_groupe_fournisseur():
    if request.method == 'POST':
        nom = request.form.get('nom')
        if GroupeFournisseur.query.filter_by(nom=nom).first():
            flash('Ce groupe existe déjà.', 'danger')
            return redirect(url_for('admin.create_groupe_fournisseur'))
            
        coeff = request.form.get('coefficient_defaut')
        tva = request.form.get('tva_defaut')
        
        new_groupe = GroupeFournisseur(
            nom=nom,
            coefficient_defaut=float(coeff) if coeff else 1.0,
            tva_defaut=float(tva) if tva else 20.0
        )
        db.session.add(new_groupe)
        db.session.commit()
        flash('Groupe de fournisseurs créé avec succès.', 'success')
        return redirect(url_for('admin.list_groupes_fournisseurs'))
    return render_template('admin/fournisseurs/groupe_form.html', title="Créer un Groupe")

@admin.route('/fournisseurs/groupes/edit/<int:id>', methods=['GET', 'POST'])
@login_required
@permission_required('gestion_groupes_fournisseurs')
def edit_groupe_fournisseur(id):
    groupe = GroupeFournisseur.query.get_or_404(id)
    if request.method == 'POST':
        groupe.nom = request.form.get('nom')
        coeff = request.form.get('coefficient_defaut')
        groupe.coefficient_defaut = float(coeff) if coeff else 1.0
        tva = request.form.get('tva_defaut')
        groupe.tva_defaut = float(tva) if tva else 20.0
        
        db.session.commit()
        flash('Groupe mis à jour avec succès.', 'success')
        return redirect(url_for('admin.list_groupes_fournisseurs'))
    return render_template('admin/fournisseurs/groupe_form.html', groupe=groupe, title="Modifier le Groupe")

@admin.route('/fournisseurs/groupes/delete/<int:id>', methods=['POST'])
@login_required
@permission_required('gestion_groupes_fournisseurs')
def delete_groupe_fournisseur(id):
    groupe = GroupeFournisseur.query.get_or_404(id)
    # Optionnel: gérer les fournisseurs orphelins si nécessaire
    db.session.delete(groupe)
    db.session.commit()
    flash('Groupe supprimé.', 'success')
    return redirect(url_for('admin.list_groupes_fournisseurs'))

@admin.route('/fournisseurs/groupes/bulk-delete', methods=['POST'])
@login_required
@permission_required('gestion_groupes_fournisseurs')
def bulk_delete_groupes_fournisseurs():
    ids = request.form.getlist('ids[]')
    if not ids:
        flash("Aucun groupe sélectionné.", "warning")
        return redirect(url_for('admin.list_groupes_fournisseurs'))
    
    deleted_count = 0
    for g_id in ids:
        g = GroupeFournisseur.query.get(g_id)
        if g:
            db.session.delete(g)
            deleted_count += 1
            
    db.session.commit()
    flash(f'{deleted_count} groupe(s) supprimé(s).', 'success')
    return redirect(url_for('admin.list_groupes_fournisseurs'))

# --- GESTION DES RAYONS ---
@admin.route('/rayons')
@login_required
@permission_required('gestion_rayons')
def list_rayons():
    rayons = Rayon.query.all()
    total_produits = sum(len(r.produits) for r in rayons)
    return render_template('admin/rayons/list.html', rayons=rayons, total_produits=total_produits)

@admin.route('/rayons/create', methods=['GET', 'POST'])
@login_required
@permission_required('gestion_rayons')
def create_rayon():
    if request.method == 'POST':
        nom = request.form.get('nom')
        if Rayon.query.filter_by(nom=nom).first():
            flash('Ce rayon existe déjà.', 'danger')
            return redirect(url_for('admin.create_rayon'))
        new_rayon = Rayon(nom=nom, description=request.form.get('description'))
        db.session.add(new_rayon)
        db.session.commit()
        flash('Rayon ajouté avec succès.', 'success')
        return redirect(url_for('admin.list_rayons'))
    return render_template('admin/rayons/form.html', title='Ajouter un Rayon')

@admin.route('/rayons/edit/<int:id>', methods=['GET', 'POST'])
@login_required
@permission_required('gestion_rayons')
def edit_rayon(id):
    rayon = Rayon.query.get_or_404(id)
    if request.method == 'POST':
        rayon.nom = request.form.get('nom')
        rayon.description = request.form.get('description')
        db.session.commit()
        flash('Rayon mis à jour.', 'success')
        return redirect(url_for('admin.list_rayons'))
    return render_template('admin/rayons/form.html', rayon=rayon, title='Modifier le Rayon')

@admin.route('/rayons/bulk-delete', methods=['POST'])
@login_required
@permission_required('gestion_rayons')
def bulk_delete_rayons():
    ids = request.form.getlist('ids[]')
    deleted_count = 0
    for r_id in ids:
        r = Rayon.query.get(r_id)
        if r:
            db.session.delete(r)
            deleted_count += 1
    db.session.commit()
    flash(f'{deleted_count} rayon(s) supprimé(s).', 'success')
    return redirect(url_for('admin.list_rayons'))

@admin.route('/rayons/delete/<int:id>', methods=['POST'])
@login_required
@permission_required('gestion_rayons')
def delete_rayon(id):
    item = Rayon.query.get_or_404(id)
    db.session.delete(item)
    db.session.commit()
    flash('Rayon supprimé.', 'success')
    return redirect(url_for('admin.list_rayons'))

# --- GESTION DES FAMILLES ---
@admin.route('/familles')
@login_required
@permission_required('gestion_familles')
def list_familles():
    familles = Famille.query.all()
    total_produits = sum(len(f.produits) for f in familles)
    return render_template('admin/familles/list.html', familles=familles, total_produits=total_produits)

@admin.route('/familles/create', methods=['GET', 'POST'])
@login_required
@permission_required('gestion_familles')
def create_famille():
    if request.method == 'POST':
        nom = request.form.get('nom')
        if Famille.query.filter_by(nom=nom).first():
            flash('Cette famille existe déjà.', 'danger')
            return redirect(url_for('admin.create_famille'))
        new_famille = Famille(nom=nom, description=request.form.get('description'))
        db.session.add(new_famille)
        db.session.commit()
        flash('Famille ajoutée.', 'success')
        return redirect(url_for('admin.list_familles'))
    return render_template('admin/familles/form.html', title='Ajouter une Famille')

@admin.route('/familles/edit/<int:id>', methods=['GET', 'POST'])
@login_required
@permission_required('gestion_familles')
def edit_famille(id):
    famille = Famille.query.get_or_404(id)
    if request.method == 'POST':
        famille.nom = request.form.get('nom')
        famille.description = request.form.get('description')
        db.session.commit()
        flash('Famille mise à jour.', 'success')
        return redirect(url_for('admin.list_familles'))
    return render_template('admin/familles/form.html', famille=famille, title='Modifier la Famille')

@admin.route('/familles/bulk-delete', methods=['POST'])
@login_required
@permission_required('gestion_familles')
def bulk_delete_familles():
    ids = request.form.getlist('ids[]')
    deleted_count = 0
    for f_id in ids:
        f = Famille.query.get(f_id)
        if f:
            db.session.delete(f)
            deleted_count += 1
    db.session.commit()
    flash(f'{deleted_count} famille(s) supprimée(s).', 'success')
    return redirect(url_for('admin.list_familles'))

@admin.route('/familles/delete/<int:id>', methods=['POST'])
@login_required
@permission_required('gestion_familles')
def delete_famille(id):
    item = Famille.query.get_or_404(id)
    db.session.delete(item)
    db.session.commit()
    flash('Famille supprimée.', 'success')
    return redirect(url_for('admin.list_familles'))

# --- GESTION DES SECTIONS ---
@admin.route('/sections')
@login_required
@permission_required('gestion_sections')
def list_sections():
    sections = Section.query.all()
    total_produits = sum(len(s.produits) for s in sections)
    return render_template('admin/sections/list.html', sections=sections, total_produits=total_produits)

@admin.route('/sections/create', methods=['GET', 'POST'])
@login_required
@permission_required('gestion_sections')
def create_section():
    if request.method == 'POST':
        nom = request.form.get('nom')
        if Section.query.filter_by(nom=nom).first():
            flash('Cette section existe déjà.', 'danger')
            return redirect(url_for('admin.create_section'))
        new_section = Section(nom=nom, description=request.form.get('description'))
        db.session.add(new_section)
        db.session.commit()
        flash('Section ajoutée.', 'success')
        return redirect(url_for('admin.list_sections'))
    return render_template('admin/sections/form.html', title='Ajouter une Section')

@admin.route('/sections/edit/<int:id>', methods=['GET', 'POST'])
@login_required
@permission_required('gestion_sections')
def edit_section(id):
    section = Section.query.get_or_404(id)
    if request.method == 'POST':
        section.nom = request.form.get('nom')
        section.description = request.form.get('description')
        db.session.commit()
        flash('Section mise à jour.', 'success')
        return redirect(url_for('admin.list_sections'))
    return render_template('admin/sections/form.html', section=section, title='Modifier la Section')

@admin.route('/sections/bulk-delete', methods=['POST'])
@login_required
@permission_required('gestion_sections')
def bulk_delete_sections():
    ids = request.form.getlist('ids[]')
    deleted_count = 0
    for s_id in ids:
        s = Section.query.get(s_id)
        if s:
            db.session.delete(s)
            deleted_count += 1
    db.session.commit()
    flash(f'{deleted_count} section(s) supprimée(s).', 'success')
    return redirect(url_for('admin.list_sections'))

@admin.route('/sections/delete/<int:id>', methods=['POST'])
@login_required
@permission_required('gestion_sections')
def delete_section(id):
    item = Section.query.get_or_404(id)
    db.session.delete(item)
    db.session.commit()
    flash('Section supprimée.', 'success')
    return redirect(url_for('admin.list_sections'))

# --- GESTION DES PRODUITS ---
@admin.route('/produits')
@login_required
@permission_required('gestion_produits')
def list_produits():
    produits = Produit.query.all()
    fournisseurs_list = Fournisseur.query.all()
    rayons_list = Rayon.query.all()
    familles_list = Famille.query.all()
    return render_template('admin/produits/list.html', produits=produits, 
                           fournisseurs_list=fournisseurs_list, 
                           rayons_list=rayons_list, 
                           familles_list=familles_list)

@admin.route('/produits/create', methods=['GET', 'POST'])
@login_required
@permission_required('gestion_produits')
def create_produit():
    fournisseurs = Fournisseur.query.all()
    rayons = Rayon.query.all()
    familles = Famille.query.all()
    sections = Section.query.all()
    
    if request.method == 'POST':
        f_id = int(request.form.get('fournisseur_id'))
        fournisseur = Fournisseur.query.get_or_404(f_id)
        
        new_produit = Produit(
            nom=request.form.get('nom'),
            code_produit=generate_product_code(fournisseur),
            fournisseur_id=f_id,
            rayon_id=int(request.form.get('rayon_id')) if request.form.get('rayon_id') else None,
            famille_id=int(request.form.get('famille_id')) if request.form.get('famille_id') else None,
            section_id=int(request.form.get('section_id')) if request.form.get('section_id') else None,
            conditionnement=int(request.form.get('conditionnement')),
            prix_unite=float(request.form.get('prix_unite') or 0),
            prix_sous_unite=float(request.form.get('prix_sous_unite') or 0),
            prix_sous_sous_unite=float(request.form.get('prix_sous_sous_unite') or 0),
            coefficient=float(request.form.get('coefficient')) if request.form.get('coefficient') else None,
            tva=float(request.form.get('tva')) if request.form.get('tva') else None,
            stock_securite=int(request.form.get('stock_securite') or 0)
        )
        db.session.add(new_produit)
        db.session.commit()
        
        flash(f'Produit créé avec le code : {new_produit.code_produit}', 'success')
        return redirect(url_for('admin.list_produits'))
        
    return render_template('admin/produits/form.html', title='Ajouter un Produit', 
                           fournisseurs=fournisseurs, rayons=rayons, familles=familles, sections=sections)

@admin.route('/produits/edit/<int:id>', methods=['GET', 'POST'])
@login_required
@permission_required('gestion_produits')
def edit_produit(id):
    produit = Produit.query.get_or_404(id)
    fournisseurs = Fournisseur.query.all()
    rayons = Rayon.query.all()
    familles = Famille.query.all()
    sections = Section.query.all()
    
    if request.method == 'POST':
        produit.nom = request.form.get('nom')
        produit.rayon_id = int(request.form.get('rayon_id')) if request.form.get('rayon_id') else None
        produit.famille_id = int(request.form.get('famille_id')) if request.form.get('famille_id') else None
        produit.section_id = int(request.form.get('section_id')) if request.form.get('section_id') else None
        produit.conditionnement = int(request.form.get('conditionnement'))
        produit.prix_unite = float(request.form.get('prix_unite') or 0)
        produit.prix_sous_unite = float(request.form.get('prix_sous_unite') or 0)
        produit.prix_sous_sous_unite = float(request.form.get('prix_sous_sous_unite') or 0)
        produit.coefficient = float(request.form.get('coefficient')) if request.form.get('coefficient') else None
        produit.tva = float(request.form.get('tva')) if request.form.get('tva') else None
        produit.stock_securite = int(request.form.get('stock_securite') or 0)

        db.session.commit()
        flash('Produit mis à jour.', 'success')
        return redirect(url_for('admin.list_produits'))
        
    return render_template('admin/produits/form.html', produit=produit, title='Modifier le Produit',
                           fournisseurs=fournisseurs, rayons=rayons, familles=familles, sections=sections)

@admin.route('/produits/bulk-delete', methods=['POST'])
@login_required
@permission_required('gestion_produits')
def bulk_delete_produits():
    ids = request.form.getlist('ids[]')
    deleted_count = 0
    for p_id in ids:
        p = Produit.query.get(p_id)
        if p:
            db.session.delete(p)
            deleted_count += 1
    db.session.commit()
    flash(f'{deleted_count} produit(s) supprimé(s).', 'success')
    return redirect(url_for('admin.list_produits'))

@admin.route('/produits/delete/<int:id>', methods=['POST'])
@login_required
@permission_required('gestion_produits')
def delete_produit(id):
    produit = Produit.query.get_or_404(id)
    db.session.delete(produit)
    db.session.commit()
    flash('Produit supprimé du catalogue.', 'success')
    return redirect(url_for('admin.list_produits'))

# --- GESTION DU STOCK ---
@admin.route('/stock', methods=['GET', 'POST'])
@login_required
@permission_required('gestion_stock')
def manage_stock():
    produits = Produit.query.order_by(Produit.nom.asc()).all()
    reasons = StockReason.query.filter_by(type='ajout').all()

    if request.method == 'POST':
        produit_id = int(request.form.get('produit_id'))
        produit = Produit.query.get_or_404(produit_id)
        reason_id = request.form.get('reason_id')
        reason_text = (request.form.get('reason') or '').strip()
        numero_bl_raw = (request.form.get('numero_bl') or '').strip()
        date_peremption_str = (request.form.get('date_peremption') or '').strip()
        
        if not reason_id and not reason_text:
            flash('Veuillez préciser la raison de cette entrée en stock.', 'warning')
            return redirect(url_for('admin.manage_stock'))
        if not numero_bl_raw:
            flash('Veuillez préciser le numéro du BL.', 'warning')
            return redirect(url_for('admin.manage_stock'))
        if not date_peremption_str:
            flash('Veuillez préciser la date de péremption.', 'warning')
            return redirect(url_for('admin.manage_stock'))

        try:
            date_peremption = datetime.strptime(date_peremption_str, '%Y-%m-%d').date()
        except ValueError:
            flash('La date de péremption est invalide.', 'danger')
            return redirect(url_for('admin.manage_stock'))

        numero_bl = Stock.normalize_bl(numero_bl_raw)

        quantite_unites = int(request.form.get('quantite_unites') or 0)
        quantite_sous_unites = int(request.form.get('quantite_sous_unites') or 0)
        quantite_sous_sous_unites = int(request.form.get('quantite_sous_sous_unites') or 0)
        code_suivi = Stock.build_tracking_code(produit.code_produit, numero_bl, date_peremption)

        stock = Stock.query.filter_by(
            produit_id=produit_id,
            numero_bl=numero_bl,
            date_peremption=date_peremption
        ).first()
        if stock is None:
            stock = Stock(
                produit_id=produit_id,
                numero_bl=numero_bl,
                date_peremption=date_peremption,
                code_suivi=code_suivi,
                quantite_unites=quantite_unites,
                quantite_sous_unites=quantite_sous_unites,
                quantite_sous_sous_unites=quantite_sous_sous_unites
            )
            db.session.add(stock)
            db.session.flush()
            create_stock_modification(
                stock=stock,
                produit=produit,
                action='create',
                reason=reason_text,
                reason_id=reason_id,
                old_values=(0, 0, 0),
                new_values=(quantite_unites, quantite_sous_unites, quantite_sous_sous_unites),
                old_qr_tire=False,
                new_qr_tire=stock.qr_tire
            )
            message = f'Stock initial ajouté pour {produit.nom}.'
        else:
            old_values = (stock.quantite_unites, stock.quantite_sous_unites, stock.quantite_sous_sous_unites)
            stock.quantite_unites += quantite_unites
            stock.quantite_sous_unites += quantite_sous_unites
            stock.quantite_sous_sous_unites += quantite_sous_sous_unites
            create_stock_modification(
                stock=stock,
                produit=produit,
                action='adjust',
                reason=reason_text,
                reason_id=reason_id,
                old_values=old_values,
                new_values=(stock.quantite_unites, stock.quantite_sous_unites, stock.quantite_sous_sous_unites),
                old_qr_tire=stock.qr_tire,
                new_qr_tire=stock.qr_tire
            )
            message = f'Stock mis à jour pour {produit.nom} ({stock.code_suivi}).'

        db.session.commit()
        flash(message, 'success')
        return redirect(url_for('admin.manage_stock'))

    stocks = Stock.query.join(Produit).order_by(Produit.nom.asc(), Stock.date_peremption.asc()).all()
    total_stock_entries = len(stocks)
    total_produits_en_stock = len(set(s.produit_id for s in stocks))
    total_quantite = sum(s.quantite_totale for s in stocks)
    qr_non_tires = sum(1 for s in stocks if not s.qr_tire)
    return render_template('admin/stock/list.html',
        produits=produits, stocks=stocks, reasons=reasons,
        total_stock_entries=total_stock_entries,
        total_produits_en_stock=total_produits_en_stock,
        total_quantite=total_quantite,
        qr_non_tires=qr_non_tires
    )

@admin.route('/stock/edit/<int:id>', methods=['POST'])
@login_required
@permission_required('gestion_stock')
def edit_stock(id):
    stock = Stock.query.get_or_404(id)
    reason = (request.form.get('reason') or '').strip()
    if not reason:
        flash('Veuillez préciser la raison de la modification du stock.', 'warning')
        return redirect(url_for('admin.manage_stock'))

    old_values = (stock.quantite_unites, stock.quantite_sous_unites, stock.quantite_sous_sous_unites)
    old_qr_tire = stock.qr_tire
    stock.quantite_unites = int(request.form.get('quantite_unites') or 0)
    stock.quantite_sous_unites = int(request.form.get('quantite_sous_unites') or 0)
    stock.quantite_sous_sous_unites = int(request.form.get('quantite_sous_sous_unites') or 0)
    stock.qr_tire = True if request.form.get('qr_tire') else False
    create_stock_modification(
        stock=stock,
        produit=stock.produit,
        action='edit',
        reason=reason,
        old_values=old_values,
        new_values=(stock.quantite_unites, stock.quantite_sous_unites, stock.quantite_sous_sous_unites),
        old_qr_tire=old_qr_tire,
        new_qr_tire=stock.qr_tire
    )
    db.session.commit()
    flash(f'Stock mis à jour pour {stock.produit.nom} ({stock.code_suivi}).', 'success')
    return redirect(url_for('admin.manage_stock'))

@admin.route('/stock/delete/<int:id>', methods=['POST'])
@login_required
@permission_required('gestion_stock')
def delete_stock(id):
    stock = Stock.query.get_or_404(id)
    reason = (request.form.get('reason') or '').strip()
    if not reason:
        flash('Veuillez préciser la raison de la suppression du stock.', 'warning')
        return redirect(url_for('admin.manage_stock'))

    produit_nom = stock.produit.nom
    code_suivi = stock.code_suivi
    old_values = (stock.quantite_unites, stock.quantite_sous_unites, stock.quantite_sous_sous_unites)
    create_stock_modification(
        stock=stock,
        produit=stock.produit,
        action='delete',
        reason=reason,
        old_values=old_values,
        new_values=(0, 0, 0),
        old_qr_tire=stock.qr_tire,
        new_qr_tire=stock.qr_tire
    )
    db.session.delete(stock)
    db.session.commit()
    flash(f'Stock supprimé pour {produit_nom} ({code_suivi}).', 'success')
    return redirect(url_for('admin.manage_stock'))

@admin.route('/stock/exit', methods=['GET', 'POST'])
@login_required
@permission_required('effectuer_sortie_stock')
def stock_exit():
    reasons = StockReason.query.filter_by(type='sortie').all()
    if request.method == 'POST':
        stock_id = request.form.get('stock_id')
        stock = Stock.query.get_or_404(stock_id)
        reason_id = request.form.get('reason_id')
        reason = StockReason.query.filter_by(id=reason_id, type='sortie').first() if reason_id else None

        if reason is None:
            flash("Veuillez preciser une raison de sortie valide.", "warning")
            return redirect(url_for('admin.stock_exit', stock_id=stock.id))
        
        q_u = int(request.form.get('quantite_unites') or 0)
        q_su = int(request.form.get('quantite_sous_unites') or 0)
        q_ssu = int(request.form.get('quantite_sous_sous_unites') or 0)

        if q_u < 0 or q_su < 0 or q_ssu < 0:
            flash("Les quantites de sortie ne peuvent pas etre negatives.", "danger")
            return redirect(url_for('admin.stock_exit', stock_id=stock.id))

        if q_u == 0 and q_su == 0 and q_ssu == 0:
            flash("Veuillez preciser au moins une quantite a sortir.", "warning")
            return redirect(url_for('admin.stock_exit', stock_id=stock.id))

        if q_u > stock.quantite_unites or q_su > stock.quantite_sous_unites or q_ssu > stock.quantite_sous_sous_unites:
            flash("La quantité de sortie dépasse le stock disponible pour ce lot.", "danger")
            return redirect(url_for('admin.manage_stock'))

        old_values = (stock.quantite_unites, stock.quantite_sous_unites, stock.quantite_sous_sous_unites)
        stock.quantite_unites -= q_u
        stock.quantite_sous_unites -= q_su
        stock.quantite_sous_sous_unites -= q_ssu

        create_stock_modification(
            stock=stock,
            produit=stock.produit,
            action='sortie',
            reason=None,
            reason_id=reason_id,
            old_values=old_values,
            new_values=(stock.quantite_unites, stock.quantite_sous_unites, stock.quantite_sous_sous_unites),
            old_qr_tire=stock.qr_tire,
            new_qr_tire=stock.qr_tire
        )
        create_stock_exit_log(
            stock=stock,
            reason=reason,
            old_values=old_values,
            new_values=(stock.quantite_unites, stock.quantite_sous_unites, stock.quantite_sous_sous_unites),
            exit_values=(q_u, q_su, q_ssu)
        )
        db.session.commit()
        flash(f'Sortie de stock effectuée pour {stock.produit.nom}.', 'success')
        return redirect(url_for('admin.manage_stock'))
    
    # Si GET, on peut passer un stock_id pour pré-remplir
    stock_id = request.args.get('stock_id')
    stock = Stock.query.get(stock_id) if stock_id else None
    stocks = Stock.query.join(Produit).order_by(Produit.nom.asc()).all()
    return render_template('admin/stock/exit_form.html', stock=stock, stocks=stocks, reasons=reasons)

@admin.route('/stock/<int:id>/qr-preview')
@login_required
@permission_required('gestion_stock')
def preview_stock_qr_codes(id):
    stock = Stock.query.get_or_404(id)
    if stock.qr_tire:
        flash('Les QR codes de ce lot sont déjà marqués comme tirés.', 'info')
        return redirect(url_for('admin.manage_stock'))

    qr_items = build_stock_qr_items([stock])
    return render_template(
        'admin/stock/qr_preview.html',
        qr_items=qr_items,
        total_qr_count=sum(item['qr_count'] for item in qr_items),
        selected_ids=[stock.id]
    )

@admin.route('/stock/qr-preview', methods=['POST'])
@login_required
@permission_required('gestion_stock')
def preview_selected_stock_qr_codes():
    selected_ids = get_selected_stock_ids()
    if not selected_ids:
        flash('Veuillez selectionner au moins une ligne de stock pour tirer les QR codes.', 'warning')
        return redirect(url_for('admin.manage_stock'))

    stocks = get_stocks_in_requested_order(selected_ids)
    if not stocks:
        flash('Aucune ligne de stock valide selectionnee.', 'warning')
        return redirect(url_for('admin.manage_stock'))

    qr_items = build_stock_qr_items(stocks, get_requested_qr_counts(selected_ids))
    return render_template(
        'admin/stock/qr_preview.html',
        qr_items=qr_items,
        total_qr_count=sum(item['qr_count'] for item in qr_items),
        selected_ids=[stock.id for stock in stocks]
    )

@admin.route('/stock/<int:id>/mark-qr-printed', methods=['POST'])
@login_required
@permission_required('gestion_stock')
def mark_stock_qr_printed(id):
    stock = Stock.query.get_or_404(id)
    reason = (request.form.get('reason') or '').strip()
    if not reason:
        flash('Veuillez préciser la raison du tirage des QR codes.', 'warning')
        return redirect(url_for('admin.preview_stock_qr_codes', id=id))

    old_values = (stock.quantite_unites, stock.quantite_sous_unites, stock.quantite_sous_sous_unites)
    old_qr_tire = stock.qr_tire
    stock.qr_tire = True
    create_stock_modification(
        stock=stock,
        produit=stock.produit,
        action='qr_print',
        reason=reason,
        old_values=old_values,
        new_values=old_values,
        old_qr_tire=old_qr_tire,
        new_qr_tire=stock.qr_tire
    )
    db.session.commit()
    flash(f'QR codes marqués comme tirés pour {stock.produit.nom} ({stock.code_suivi}).', 'success')
    return redirect(url_for('admin.manage_stock'))

@admin.route('/stock/mark-qr-printed', methods=['POST'])
@login_required
@permission_required('gestion_stock')
def mark_selected_stock_qr_printed():
    selected_ids = get_selected_stock_ids()
    reason = (request.form.get('reason') or '').strip()

    if not selected_ids:
        flash('Veuillez selectionner au moins une ligne de stock.', 'warning')
        return redirect(url_for('admin.manage_stock'))
    if not reason:
        flash('Veuillez preciser la raison du tirage des QR codes.', 'warning')
        return redirect(url_for('admin.manage_stock'))

    stocks = get_stocks_in_requested_order(selected_ids)
    updated_count = 0
    for stock in stocks:
        old_values = (stock.quantite_unites, stock.quantite_sous_unites, stock.quantite_sous_sous_unites)
        old_qr_tire = stock.qr_tire
        stock.qr_tire = True
        create_stock_modification(
            stock=stock,
            produit=stock.produit,
            action='qr_print',
            reason=reason,
            old_values=old_values,
            new_values=old_values,
            old_qr_tire=old_qr_tire,
            new_qr_tire=stock.qr_tire
        )
        updated_count += 1

    db.session.commit()
    flash(f'QR codes marques comme tires pour {updated_count} lot(s).', 'success')
    return redirect(url_for('admin.manage_stock'))

@admin.route('/stock/export/qr/excel', methods=['POST'])
@login_required
@permission_required('gestion_stock')
def export_selected_stock_qr_excel():
    selected_ids = get_selected_stock_ids()
    stocks = get_stocks_in_requested_order(selected_ids)
    if not stocks:
        flash('Veuillez selectionner au moins une ligne de stock a exporter.', 'warning')
        return redirect(url_for('admin.manage_stock'))

    import pandas as pd
    import io
    from flask import send_file
    from openpyxl.styles import Alignment, Font, PatternFill, Border, Side

    generated_at = datetime.now()
    rows = build_stock_qr_report_rows(stocks)
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df = pd.DataFrame(rows)
        df = df.rename(columns={
            'produit': 'Produit',
            'code_suivi': 'Code suivi',
            'numero_bl': 'BL',
            'date_peremption': 'Peremption',
            'unites': 'Unites',
            'sous_unites': 'Sous-unites',
            'sous_sous_unites': 'Sous-sous-unites',
            'qr_tire': 'QR tire',
            'date_tirage': 'Date tirage QR',
            'tire_par': 'Tire par'
        })
        df.to_excel(writer, index=False, sheet_name='Stock QR', startrow=4)
        worksheet = writer.sheets['Stock QR']
        worksheet['A1'] = 'Rapport Stock QR - ReflexPharma'
        worksheet['A2'] = f'Genere le : {generated_at.strftime("%d/%m/%Y %H:%M")}'
        worksheet['A3'] = f'Genere par : {current_user.nom} {current_user.prenom}'
        worksheet['A4'] = f'Lignes exportees : {len(rows)}'

        title_font = Font(bold=True, size=14, color='1F2937')
        meta_font = Font(size=10, color='374151')
        header_font = Font(bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='1F2937', end_color='1F2937', fill_type='solid')
        thin_border = Border(bottom=Side(style='thin', color='D1D5DB'))

        worksheet['A1'].font = title_font
        for row_number in range(2, 5):
            worksheet[f'A{row_number}'].font = meta_font

        for cell in worksheet[5]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        for row in worksheet.iter_rows(min_row=6, max_row=worksheet.max_row):
            for cell in row:
                cell.border = thin_border
                cell.alignment = Alignment(vertical='top', wrap_text=True)

        widths = [24, 34, 16, 14, 10, 12, 16, 10, 18, 22]
        for index, width in enumerate(widths, start=1):
            worksheet.column_dimensions[chr(64 + index)].width = width
        worksheet.freeze_panes = 'A6'

    output.seek(0)
    filename = f'stock_qr_{generated_at.strftime("%Y%m%d_%H%M")}.xlsx'
    return send_file(output, download_name=filename, as_attachment=True)

@admin.route('/stock/export/qr/pdf', methods=['POST'])
@login_required
@permission_required('gestion_stock')
def export_selected_stock_qr_pdf():
    selected_ids = get_selected_stock_ids()
    stocks = get_stocks_in_requested_order(selected_ids)
    if not stocks:
        flash('Veuillez selectionner au moins une ligne de stock a exporter.', 'warning')
        return redirect(url_for('admin.manage_stock'))

    import io
    from flask import send_file
    from reportlab.lib import colors
    from reportlab.lib.enums import TA_CENTER
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
    from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

    generated_at = datetime.now()
    rows = build_stock_qr_report_rows(stocks)
    output = io.BytesIO()
    doc = SimpleDocTemplate(
        output,
        pagesize=landscape(A4),
        topMargin=12,
        bottomMargin=12,
        leftMargin=12,
        rightMargin=12
    )
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle('StockQrTitle', parent=styles['Title'], fontSize=12, leading=14, spaceAfter=4)
    meta_style = ParagraphStyle('StockQrMeta', parent=styles['Normal'], fontSize=7, leading=9, textColor=colors.HexColor('#374151'))
    cell_style = ParagraphStyle('StockQrCell', parent=styles['Normal'], fontSize=6, leading=7)
    header_style = ParagraphStyle('StockQrHeader', parent=cell_style, alignment=TA_CENTER, textColor=colors.white)

    elements = [
        Paragraph('Rapport Stock QR - ReflexPharma', title_style),
        Paragraph(
            f'Genere le : {generated_at.strftime("%d/%m/%Y %H:%M")} | Genere par : {current_user.nom} {current_user.prenom} | Lignes : {len(rows)}',
            meta_style
        ),
        Spacer(1, 6)
    ]

    data = [[
        Paragraph('Produit', header_style),
        Paragraph('Code suivi', header_style),
        Paragraph('BL', header_style),
        Paragraph('Peremp.', header_style),
        Paragraph('U', header_style),
        Paragraph('S/U', header_style),
        Paragraph('SS/U', header_style),
        Paragraph('QR', header_style),
        Paragraph('Date tirage', header_style),
        Paragraph('Tire par', header_style)
    ]]
    for row in rows:
        data.append([
            Paragraph(str(row['produit']), cell_style),
            Paragraph(str(row['code_suivi']), cell_style),
            Paragraph(str(row['numero_bl']), cell_style),
            row['date_peremption'],
            row['unites'],
            row['sous_unites'],
            row['sous_sous_unites'],
            row['qr_tire'],
            row['date_tirage'],
            Paragraph(str(row['tire_par']), cell_style)
        ])

    table = Table(
        data,
        repeatRows=1,
        colWidths=[95, 155, 70, 52, 28, 32, 36, 30, 75, 85]
    )
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1F2937')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 6),
        ('LEADING', (0, 0), (-1, -1), 7),
        ('ALIGN', (3, 1), (8, -1), 'CENTER'),
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ('GRID', (0, 0), (-1, -1), 0.25, colors.HexColor('#D1D5DB')),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F9FAFB')]),
        ('TOPPADDING', (0, 0), (-1, -1), 2),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 2),
        ('LEFTPADDING', (0, 0), (-1, -1), 2),
        ('RIGHTPADDING', (0, 0), (-1, -1), 2),
    ]))
    elements.append(table)
    doc.build(elements)
    output.seek(0)
    filename = f'stock_qr_{generated_at.strftime("%Y%m%d_%H%M")}.pdf'
    return send_file(output, download_name=filename, as_attachment=True)

@admin.route('/stock/modifications')
@login_required
@permission_required('gestion_modifications_stock')
def list_stock_modifications():
    modifications = StockModification.query.order_by(StockModification.created_at.desc()).all()
    return render_template('admin/stock/modifications.html', modifications=modifications)

@admin.route('/stock/exits')
@login_required
@permission_required('gestion_modifications_stock')
def list_stock_exit_logs():
    exits = StockExitLog.query.order_by(StockExitLog.created_at.desc()).all()
    exit_prices = {item.id: get_exit_log_prices(item) for item in exits}
    exit_suppliers = {item.id: get_exit_log_supplier_info(item) for item in exits}
    return render_template(
        'admin/stock/exit_logs.html',
        exits=exits,
        exit_prices=exit_prices,
        exit_suppliers=exit_suppliers
    )

def parse_date_arg(name):
    value = (request.args.get(name) or '').strip()
    if not value:
        return None
    try:
        return datetime.strptime(value, '%Y-%m-%d').date()
    except ValueError:
        return None

def parse_float_arg(name):
    value = (request.args.get(name) or '').strip()
    if not value:
        return None
    try:
        return float(value)
    except ValueError:
        return None

def get_stock_exit_stats_filters():
    return {
        'date_from': request.args.get('date_from', '').strip(),
        'date_to': request.args.get('date_to', '').strip(),
        'produit': request.args.get('produit', '').strip(),
        'fournisseur': request.args.get('fournisseur', '').strip(),
        'groupe': request.args.get('groupe', '').strip(),
        'sorti_par': request.args.get('sorti_par', '').strip(),
        'mis_en_stock_par': request.args.get('mis_en_stock_par', '').strip(),
        'raison': request.args.get('raison', '').strip(),
        'tva': request.args.get('tva', '').strip(),
        'min_ttc': request.args.get('min_ttc', '').strip(),
        'max_ttc': request.args.get('max_ttc', '').strip()
    }

def get_stock_exit_stats_options(exits):
    def sorted_values(values):
        return sorted(value for value in values if value)

    return {
        'produits': sorted_values({f'{item.produit_nom} ({item.produit_code})' for item in exits}),
        'fournisseurs': sorted_values({item.fournisseur_nom or '-' for item in exits}),
        'groupes': sorted_values({item.groupe_fournisseur_nom or '-' for item in exits}),
        'sorti_par': sorted_values({f'{item.user_prenom} {item.user_nom}'.strip() for item in exits}),
        'mis_en_stock_par': sorted_values({f'{item.mise_en_stock_user_prenom or ""} {item.mise_en_stock_user_nom or ""}'.strip() or '-' for item in exits}),
        'raisons': sorted_values({item.reason_nom or '-' for item in exits}),
        'tvas': sorted_values({f'{get_exit_log_prices(item)["tva_pourcentage"]:.2f}' for item in exits})
    }

def get_filtered_stock_exit_logs(filters=None):
    filters = filters or get_stock_exit_stats_filters()
    exits = StockExitLog.query.order_by(StockExitLog.created_at.asc()).all()
    date_from = parse_date_arg('date_from')
    date_to = parse_date_arg('date_to')
    min_ttc = parse_float_arg('min_ttc')
    max_ttc = parse_float_arg('max_ttc')

    def item_value(item, key):
        if key == 'produit':
            return f'{item.produit_nom} ({item.produit_code})'
        if key == 'fournisseur':
            return item.fournisseur_nom or '-'
        if key == 'groupe':
            return item.groupe_fournisseur_nom or '-'
        if key == 'sorti_par':
            return f'{item.user_prenom} {item.user_nom}'.strip()
        if key == 'mis_en_stock_par':
            return f'{item.mise_en_stock_user_prenom or ""} {item.mise_en_stock_user_nom or ""}'.strip() or '-'
        if key == 'raison':
            return item.reason_nom or '-'
        if key == 'tva':
            return f'{get_exit_log_prices(item)["tva_pourcentage"]:.2f}'
        return ''

    filtered = []
    for item in exits:
        item_date = item.created_at.date() if item.created_at else None
        if date_from and item_date and item_date < date_from:
            continue
        if date_to and item_date and item_date > date_to:
            continue
        if min_ttc is not None and get_exit_log_prices(item)['total_ttc'] < min_ttc:
            continue
        if max_ttc is not None and get_exit_log_prices(item)['total_ttc'] > max_ttc:
            continue
        if any(filters[key] and item_value(item, key) != filters[key] for key in ['produit', 'fournisseur', 'groupe', 'sorti_par', 'mis_en_stock_par', 'raison', 'tva']):
            continue
        filtered.append(item)
    return filtered

def build_stock_exit_stats(exits=None):
    from collections import defaultdict

    exits = exits if exits is not None else StockExitLog.query.order_by(StockExitLog.created_at.asc()).all()
    now = datetime.now()
    today = now.date()

    totals = {
        'count': len(exits),
        'total_ht': 0.0,
        'total_ttc': 0.0,
        'total_taxes': 0.0,
        'today_count': 0,
        'today_ttc': 0.0,
        'today_taxes': 0.0,
        'last_7_count': 0,
        'last_7_ttc': 0.0,
        'last_7_taxes': 0.0,
        'last_30_count': 0,
        'last_30_ttc': 0.0,
        'last_30_taxes': 0.0,
        'expired_count': 0,
        'expired_ttc': 0.0,
        'expired_taxes': 0.0,
        'quantity_total': 0,
        'avg_ttc': 0.0,
        'avg_taxes': 0.0,
        'avg_stock_age_days': 0.0
    }

    daily = defaultdict(lambda: {'count': 0, 'ht': 0.0, 'taxes': 0.0, 'ttc': 0.0})
    monthly = defaultdict(lambda: {'count': 0, 'ht': 0.0, 'taxes': 0.0, 'ttc': 0.0})
    by_product = defaultdict(lambda: {'count': 0, 'ttc': 0.0, 'quantity': 0})
    by_supplier = defaultdict(lambda: {'count': 0, 'ttc': 0.0, 'quantity': 0})
    by_supplier_group = defaultdict(lambda: {'count': 0, 'ttc': 0.0, 'quantity': 0})
    by_exit_user = defaultdict(lambda: {'count': 0, 'ttc': 0.0, 'quantity': 0})
    by_stock_user = defaultdict(lambda: {'count': 0, 'ttc': 0.0, 'quantity': 0})
    by_reason = defaultdict(lambda: {'count': 0, 'ttc': 0.0, 'quantity': 0})
    by_tva = defaultdict(lambda: {'count': 0, 'ttc': 0.0, 'taxes': 0.0})
    by_weekday = defaultdict(lambda: {'count': 0, 'ht': 0.0, 'taxes': 0.0, 'ttc': 0.0})
    by_hour = defaultdict(lambda: {'count': 0, 'ht': 0.0, 'taxes': 0.0, 'ttc': 0.0})
    stock_age_days = []

    for item in exits:
        prices = get_exit_log_prices(item)
        total_ht = prices['total_ht']
        total_ttc = prices['total_ttc']
        total_taxes = max(total_ttc - total_ht, 0)
        quantity = (
            item.quantite_unites_sortie
            + item.quantite_sous_unites_sortie
            + item.quantite_sous_sous_unites_sortie
        )
        created_at = item.created_at or now
        created_date = created_at.date()

        totals['total_ht'] += total_ht
        totals['total_ttc'] += total_ttc
        totals['total_taxes'] += total_taxes
        totals['quantity_total'] += quantity

        if created_date == today:
            totals['today_count'] += 1
            totals['today_ttc'] += total_ttc
            totals['today_taxes'] += total_taxes
        if (today - created_date).days <= 6:
            totals['last_7_count'] += 1
            totals['last_7_ttc'] += total_ttc
            totals['last_7_taxes'] += total_taxes
        if (today - created_date).days <= 29:
            totals['last_30_count'] += 1
            totals['last_30_ttc'] += total_ttc
            totals['last_30_taxes'] += total_taxes
        if item.date_peremption and item.date_peremption < created_date:
            totals['expired_count'] += 1
            totals['expired_ttc'] += total_ttc
            totals['expired_taxes'] += total_taxes
        if item.mise_en_stock_at:
            stock_age_days.append(max((created_at - item.mise_en_stock_at).days, 0))

        day_key = created_at.strftime('%Y-%m-%d')
        month_key = created_at.strftime('%Y-%m')
        daily[day_key]['count'] += 1
        daily[day_key]['ht'] += total_ht
        daily[day_key]['taxes'] += total_taxes
        daily[day_key]['ttc'] += total_ttc
        monthly[month_key]['count'] += 1
        monthly[month_key]['ht'] += total_ht
        monthly[month_key]['taxes'] += total_taxes
        monthly[month_key]['ttc'] += total_ttc

        product_key = f'{item.produit_nom} ({item.produit_code})'
        supplier_key = item.fournisseur_nom or '-'
        supplier_group_key = item.groupe_fournisseur_nom or '-'
        exit_user_key = f'{item.user_prenom} {item.user_nom}'.strip() or '-'
        stock_user_key = f'{item.mise_en_stock_user_prenom or ""} {item.mise_en_stock_user_nom or ""}'.strip() or '-'
        reason_key = item.reason_nom or '-'
        tva_key = f'{prices["tva_pourcentage"]:.2f}%'
        weekday_key = created_at.strftime('%A')
        hour_key = created_at.strftime('%H:00')

        for bucket, key in [
            (by_product, product_key),
            (by_supplier, supplier_key),
            (by_supplier_group, supplier_group_key),
            (by_exit_user, exit_user_key),
            (by_stock_user, stock_user_key),
            (by_reason, reason_key),
        ]:
            bucket[key]['count'] += 1
            bucket[key]['ttc'] += total_ttc
            bucket[key]['quantity'] += quantity

        by_tva[tva_key]['count'] += 1
        by_tva[tva_key]['ttc'] += total_ttc
        by_tva[tva_key]['taxes'] += total_taxes
        by_weekday[weekday_key]['count'] += 1
        by_weekday[weekday_key]['ht'] += total_ht
        by_weekday[weekday_key]['taxes'] += total_taxes
        by_weekday[weekday_key]['ttc'] += total_ttc
        by_hour[hour_key]['count'] += 1
        by_hour[hour_key]['ht'] += total_ht
        by_hour[hour_key]['taxes'] += total_taxes
        by_hour[hour_key]['ttc'] += total_ttc

    totals['avg_ttc'] = totals['total_ttc'] / totals['count'] if totals['count'] else 0
    totals['avg_taxes'] = totals['total_taxes'] / totals['count'] if totals['count'] else 0
    totals['avg_stock_age_days'] = sum(stock_age_days) / len(stock_age_days) if stock_age_days else 0

    def top_rows(bucket, limit=10):
        return [
            {
                'label': label,
                'count': values['count'],
                'quantity': values.get('quantity', 0),
                'ttc': round(values['ttc'], 2),
                'taxes': round(values.get('taxes', 0), 2)
            }
            for label, values in sorted(bucket.items(), key=lambda item: item[1]['ttc'], reverse=True)[:limit]
        ]

    def series_rows(bucket):
        labels = sorted(bucket.keys())
        return {
            'labels': labels,
            'count': [bucket[label]['count'] for label in labels],
            'ht': [round(bucket[label].get('ht', 0), 2) for label in labels],
            'taxes': [round(bucket[label].get('taxes', 0), 2) for label in labels],
            'ttc': [round(bucket[label]['ttc'], 2) for label in labels]
        }

    return {
        'totals': {key: round(value, 2) if isinstance(value, float) else value for key, value in totals.items()},
        'daily': series_rows(daily),
        'monthly': series_rows(monthly),
        'hourly': series_rows(by_hour),
        'weekday': series_rows(by_weekday),
        'top_products': top_rows(by_product),
        'top_suppliers': top_rows(by_supplier),
        'top_supplier_groups': top_rows(by_supplier_group),
        'top_exit_users': top_rows(by_exit_user),
        'top_stock_users': top_rows(by_stock_user),
        'top_reasons': top_rows(by_reason),
        'tva': top_rows(by_tva, limit=20)
    }

@admin.route('/stock/exits/stats')
@login_required
@permission_required('stats_sorties_stock')
def stock_exit_stats():
    filters = get_stock_exit_stats_filters()
    all_exits = StockExitLog.query.order_by(StockExitLog.created_at.asc()).all()
    exits = get_filtered_stock_exit_logs(filters)
    stats = build_stock_exit_stats(exits)
    options = get_stock_exit_stats_options(all_exits)
    return render_template('admin/stock/exit_stats.html', stats=stats, filters=filters, options=options)

@admin.route('/stock/exits/stats/export/excel')
@login_required
@permission_required('stats_sorties_stock')
def export_stock_exit_stats_excel():
    import io
    import pandas as pd
    from flask import send_file
    from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
    from openpyxl.chart import BarChart, LineChart, PieChart, Reference

    filters = get_stock_exit_stats_filters()
    exits = get_filtered_stock_exit_logs(filters)
    stats = build_stock_exit_stats(exits)
    rows = build_stock_exit_log_rows(exits)
    generated_at = datetime.now()
    output = io.BytesIO()

    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        summary_rows = [
            {'Indicateur': 'Sorties totales', 'Valeur': stats['totals']['count']},
            {'Indicateur': 'Quantite totale', 'Valeur': stats['totals']['quantity_total']},
            {'Indicateur': 'Total HT', 'Valeur': stats['totals']['total_ht']},
            {'Indicateur': 'Taxes perdues', 'Valeur': stats['totals']['total_taxes']},
            {'Indicateur': 'Perte / valeur TTC', 'Valeur': stats['totals']['total_ttc']},
            {'Indicateur': 'Sortie moyenne TTC', 'Valeur': stats['totals']['avg_ttc']},
            {'Indicateur': 'Taxes moyennes par sortie', 'Valeur': stats['totals']['avg_taxes']},
            {'Indicateur': 'Age moyen stock avant sortie', 'Valeur': stats['totals']['avg_stock_age_days']},
            {'Indicateur': 'Produits perimes sortis', 'Valeur': stats['totals']['expired_count']},
            {'Indicateur': 'Taxes perdues produits perimes', 'Valeur': stats['totals']['expired_taxes']},
            {'Indicateur': 'Perte TTC produits perimes', 'Valeur': stats['totals']['expired_ttc']}
        ]
        pd.DataFrame(summary_rows).to_excel(writer, index=False, sheet_name='Synthese', startrow=5)
        pd.DataFrame(rows).to_excel(writer, index=False, sheet_name='Details')
        for sheet_name in ['Top produits', 'Top fournisseurs', 'Top raisons', 'Par utilisateur']:
            data = {
                'Top produits': stats['top_products'],
                'Top fournisseurs': stats['top_suppliers'],
                'Top raisons': stats['top_reasons'],
                'Par utilisateur': stats['top_exit_users']
            }[sheet_name]
            pd.DataFrame(data).to_excel(writer, index=False, sheet_name=sheet_name)

        chart_sheet = writer.book.create_sheet('Courbes')
        chart_sheet['A1'] = 'Evolution quotidienne'
        chart_sheet.append(['Date', 'Sorties', 'HT', 'Taxes', 'TTC'])
        for label, count, ht, taxes, ttc in zip(stats['daily']['labels'], stats['daily']['count'], stats['daily']['ht'], stats['daily']['taxes'], stats['daily']['ttc']):
            chart_sheet.append([label, count, ht, taxes, ttc])

        monthly_start = chart_sheet.max_row + 3
        chart_sheet.cell(monthly_start, 1, 'Evolution mensuelle')
        chart_sheet.cell(monthly_start + 1, 1, 'Mois')
        chart_sheet.cell(monthly_start + 1, 2, 'Sorties')
        chart_sheet.cell(monthly_start + 1, 3, 'HT')
        chart_sheet.cell(monthly_start + 1, 4, 'Taxes')
        chart_sheet.cell(monthly_start + 1, 5, 'TTC')
        for offset, (label, count, ht, taxes, ttc) in enumerate(zip(stats['monthly']['labels'], stats['monthly']['count'], stats['monthly']['ht'], stats['monthly']['taxes'], stats['monthly']['ttc']), start=monthly_start + 2):
            chart_sheet.cell(offset, 1, label)
            chart_sheet.cell(offset, 2, count)
            chart_sheet.cell(offset, 3, ht)
            chart_sheet.cell(offset, 4, taxes)
            chart_sheet.cell(offset, 5, ttc)

        products_start = chart_sheet.max_row + 3
        chart_sheet.cell(products_start, 1, 'Top produits')
        chart_sheet.cell(products_start + 1, 1, 'Produit')
        chart_sheet.cell(products_start + 1, 2, 'TTC')
        for offset, row in enumerate(stats['top_products'][:10], start=products_start + 2):
            chart_sheet.cell(offset, 1, row['label'])
            chart_sheet.cell(offset, 2, row['ttc'])

        reasons_start = chart_sheet.max_row + 3
        chart_sheet.cell(reasons_start, 1, 'Top raisons')
        chart_sheet.cell(reasons_start + 1, 1, 'Raison')
        chart_sheet.cell(reasons_start + 1, 2, 'TTC')
        for offset, row in enumerate(stats['top_reasons'][:10], start=reasons_start + 2):
            chart_sheet.cell(offset, 1, row['label'])
            chart_sheet.cell(offset, 2, row['ttc'])

        worksheet = writer.sheets['Synthese']
        worksheet['A1'] = 'Rapport statistiques sorties de stock - ReflexPharma'
        worksheet['A2'] = f'Date du tirage : {generated_at.strftime("%d/%m/%Y %H:%M")}'
        worksheet['A3'] = f'Tire par : {current_user.nom} {current_user.prenom}'
        active_filters = ', '.join(f'{key}={value}' for key, value in filters.items() if value) or 'Aucun filtre'
        worksheet['A4'] = f'Filtres : {active_filters}'
        worksheet['A5'] = f'Lignes analysees : {len(exits)}'

        header_fill = PatternFill(start_color='1F2937', end_color='1F2937', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF')
        thin_border = Border(bottom=Side(style='thin', color='D1D5DB'))
        for ws in writer.sheets.values():
            for row in ws.iter_rows():
                for cell in row:
                    cell.alignment = Alignment(vertical='top', wrap_text=True)
                    cell.border = thin_border
            for cell in ws[1]:
                cell.font = header_font
                cell.fill = header_fill
            for col in range(1, min(ws.max_column, 12) + 1):
                ws.column_dimensions[chr(64 + col)].width = 22

        if stats['daily']['labels']:
            daily_chart = LineChart()
            daily_chart.title = 'Sorties quotidiennes'
            daily_chart.y_axis.title = 'HT / taxes / TTC / sorties'
            daily_chart.x_axis.title = 'Date'
            daily_chart.add_data(Reference(chart_sheet, min_col=2, max_col=5, min_row=2, max_row=2 + len(stats['daily']['labels'])), titles_from_data=True)
            daily_chart.set_categories(Reference(chart_sheet, min_col=1, min_row=3, max_row=2 + len(stats['daily']['labels'])))
            daily_chart.height = 9
            daily_chart.width = 22
            chart_sheet.add_chart(daily_chart, 'E2')

        if stats['monthly']['labels']:
            monthly_chart = BarChart()
            monthly_chart.title = 'Valeur mensuelle HT / taxes / TTC'
            monthly_chart.y_axis.title = 'Valeur'
            monthly_chart.x_axis.title = 'Mois'
            monthly_end = monthly_start + 1 + len(stats['monthly']['labels'])
            monthly_chart.add_data(Reference(chart_sheet, min_col=3, max_col=5, min_row=monthly_start + 1, max_row=monthly_end), titles_from_data=True)
            monthly_chart.set_categories(Reference(chart_sheet, min_col=1, min_row=monthly_start + 2, max_row=monthly_end))
            monthly_chart.height = 8
            monthly_chart.width = 18
            chart_sheet.add_chart(monthly_chart, 'E20')

        if stats['top_products']:
            product_chart = BarChart()
            product_chart.type = 'bar'
            product_chart.title = 'Top produits par perte TTC'
            product_end = products_start + 1 + min(len(stats['top_products']), 10)
            product_chart.add_data(Reference(chart_sheet, min_col=2, min_row=products_start + 1, max_row=product_end), titles_from_data=True)
            product_chart.set_categories(Reference(chart_sheet, min_col=1, min_row=products_start + 2, max_row=product_end))
            product_chart.height = 10
            product_chart.width = 22
            chart_sheet.add_chart(product_chart, 'E36')

        if stats['top_reasons']:
            reason_chart = PieChart()
            reason_chart.title = 'Repartition par raison'
            reason_end = reasons_start + 1 + min(len(stats['top_reasons']), 10)
            reason_chart.add_data(Reference(chart_sheet, min_col=2, min_row=reasons_start + 1, max_row=reason_end), titles_from_data=True)
            reason_chart.set_categories(Reference(chart_sheet, min_col=1, min_row=reasons_start + 2, max_row=reason_end))
            reason_chart.height = 9
            reason_chart.width = 14
            chart_sheet.add_chart(reason_chart, 'E56')

    output.seek(0)
    filename = f'stats_sorties_stock_{generated_at.strftime("%Y%m%d_%H%M")}.xlsx'
    return send_file(output, download_name=filename, as_attachment=True)

@admin.route('/stock/exits/stats/export/pdf')
@login_required
@permission_required('stats_sorties_stock')
def export_stock_exit_stats_pdf():
    import io
    from flask import send_file
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle
    from reportlab.graphics.shapes import Drawing, String
    from reportlab.graphics.charts.barcharts import VerticalBarChart, HorizontalBarChart
    from reportlab.graphics.charts.linecharts import HorizontalLineChart
    from reportlab.graphics.charts.piecharts import Pie

    filters = get_stock_exit_stats_filters()
    exits = get_filtered_stock_exit_logs(filters)
    stats = build_stock_exit_stats(exits)
    generated_at = datetime.now()
    output = io.BytesIO()
    doc = SimpleDocTemplate(output, pagesize=A4, topMargin=18, bottomMargin=18, leftMargin=18, rightMargin=18)
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle('StatsTitle', parent=styles['Title'], fontSize=13, leading=15)
    meta_style = ParagraphStyle('StatsMeta', parent=styles['Normal'], fontSize=8, leading=10)
    elements = [
        Paragraph('Rapport statistiques sorties de stock - ReflexPharma', title_style),
        Paragraph(f'Date du tirage : {generated_at.strftime("%d/%m/%Y %H:%M")} | Tire par : {current_user.nom} {current_user.prenom}', meta_style),
        Paragraph('Filtres : ' + (', '.join(f'{key}={value}' for key, value in filters.items() if value) or 'Aucun filtre'), meta_style),
        Spacer(1, 8)
    ]

    summary = [
        ['Indicateur', 'Valeur'],
        ['Sorties totales', stats['totals']['count']],
        ['Quantite totale', stats['totals']['quantity_total']],
        ['Total HT', f"{stats['totals']['total_ht']:.2f}"],
        ['Taxes perdues', f"{stats['totals']['total_taxes']:.2f}"],
        ['Perte / valeur TTC', f"{stats['totals']['total_ttc']:.2f}"],
        ['Sortie moyenne TTC', f"{stats['totals']['avg_ttc']:.2f}"],
        ['Taxes moyennes par sortie', f"{stats['totals']['avg_taxes']:.2f}"],
        ['Age moyen avant sortie', f"{stats['totals']['avg_stock_age_days']:.1f} j"],
        ['Produits perimes sortis', stats['totals']['expired_count']],
        ['Taxes perdues produits perimes', f"{stats['totals']['expired_taxes']:.2f}"]
    ]
    elements.append(Table(summary, repeatRows=1, colWidths=[260, 240], style=[
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1F2937')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('GRID', (0, 0), (-1, -1), 0.25, colors.HexColor('#D1D5DB')),
        ('FONTSIZE', (0, 0), (-1, -1), 8),
        ('TOPPADDING', (0, 0), (-1, -1), 3),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 3),
    ]))
    elements.append(Spacer(1, 10))

    def add_vertical_chart(title, labels, values, value_label='TTC', max_items=12):
        labels = labels[:max_items]
        values = values[:max_items]
        if not labels or not values:
            return
        drawing = Drawing(500, 185)
        drawing.add(String(0, 170, title, fontSize=10, fillColor=colors.HexColor('#1F2937')))
        chart = VerticalBarChart()
        chart.x = 35
        chart.y = 35
        chart.height = 115
        chart.width = 430
        chart.data = [values]
        chart.categoryAxis.categoryNames = labels
        chart.categoryAxis.labels.angle = 35
        chart.categoryAxis.labels.fontSize = 5
        chart.valueAxis.valueMin = 0
        chart.bars[0].fillColor = colors.HexColor('#DC3545')
        drawing.add(chart)
        drawing.add(String(35, 12, value_label, fontSize=7, fillColor=colors.HexColor('#6B7280')))
        elements.append(drawing)
        elements.append(Spacer(1, 8))

    def add_grouped_vertical_chart(title, labels, series, max_items=12):
        labels = labels[:max_items]
        if not labels or not series:
            return
        drawing = Drawing(500, 205)
        drawing.add(String(0, 190, title, fontSize=10, fillColor=colors.HexColor('#1F2937')))
        chart = VerticalBarChart()
        chart.x = 35
        chart.y = 42
        chart.height = 125
        chart.width = 430
        chart.data = [values[:max_items] for _, values, _ in series]
        chart.categoryAxis.categoryNames = labels
        chart.categoryAxis.labels.angle = 35
        chart.categoryAxis.labels.fontSize = 5
        chart.valueAxis.valueMin = 0
        for index, (_, _, color) in enumerate(series):
            chart.bars[index].fillColor = color
        drawing.add(chart)
        y = 18
        x = 35
        for label, _, color in series:
            drawing.add(String(x, y, label, fontSize=7, fillColor=color))
            x += 80
        elements.append(drawing)
        elements.append(Spacer(1, 8))

    def add_horizontal_chart(title, rows, max_items=8):
        rows = rows[:max_items]
        if not rows:
            return
        drawing = Drawing(500, 210)
        drawing.add(String(0, 195, title, fontSize=10, fillColor=colors.HexColor('#1F2937')))
        chart = HorizontalBarChart()
        chart.x = 125
        chart.y = 25
        chart.height = 155
        chart.width = 330
        chart.data = [[row['ttc'] for row in rows]]
        chart.categoryAxis.categoryNames = [row['label'][:28] for row in rows]
        chart.categoryAxis.labels.fontSize = 6
        chart.valueAxis.valueMin = 0
        chart.bars[0].fillColor = colors.HexColor('#198754')
        drawing.add(chart)
        elements.append(drawing)
        elements.append(Spacer(1, 8))

    def add_line_chart(title, labels, values, max_items=18):
        labels = labels[-max_items:]
        values = values[-max_items:]
        if not labels or not values:
            return
        drawing = Drawing(500, 185)
        drawing.add(String(0, 170, title, fontSize=10, fillColor=colors.HexColor('#1F2937')))
        chart = HorizontalLineChart()
        chart.x = 35
        chart.y = 35
        chart.height = 115
        chart.width = 430
        chart.data = [values]
        chart.categoryAxis.categoryNames = labels
        chart.categoryAxis.labels.angle = 35
        chart.categoryAxis.labels.fontSize = 5
        chart.valueAxis.valueMin = 0
        chart.lines[0].strokeColor = colors.HexColor('#DC3545')
        drawing.add(chart)
        elements.append(drawing)
        elements.append(Spacer(1, 8))

    def add_multi_line_chart(title, labels, series, max_items=18):
        labels = labels[-max_items:]
        if not labels or not series:
            return
        drawing = Drawing(500, 200)
        drawing.add(String(0, 185, title, fontSize=10, fillColor=colors.HexColor('#1F2937')))
        chart = HorizontalLineChart()
        chart.x = 35
        chart.y = 45
        chart.height = 115
        chart.width = 430
        chart.data = [values[-max_items:] for _, values, _ in series]
        chart.categoryAxis.categoryNames = labels
        chart.categoryAxis.labels.angle = 35
        chart.categoryAxis.labels.fontSize = 5
        chart.valueAxis.valueMin = 0
        for index, (_, _, color) in enumerate(series):
            chart.lines[index].strokeColor = color
        drawing.add(chart)
        x = 35
        for label, _, color in series:
            drawing.add(String(x, 18, label, fontSize=7, fillColor=color))
            x += 85
        elements.append(drawing)
        elements.append(Spacer(1, 8))

    def add_pie_chart(title, rows, max_items=6):
        rows = rows[:max_items]
        if not rows:
            return
        drawing = Drawing(500, 180)
        drawing.add(String(0, 165, title, fontSize=10, fillColor=colors.HexColor('#1F2937')))
        pie = Pie()
        pie.x = 20
        pie.y = 25
        pie.width = 130
        pie.height = 130
        pie.data = [row['ttc'] for row in rows]
        pie.labels = [row['label'][:18] for row in rows]
        drawing.add(pie)
        y = 135
        for row in rows:
            drawing.add(String(180, y, f"{row['label'][:40]}: {row['ttc']:.2f}", fontSize=7, fillColor=colors.HexColor('#374151')))
            y -= 16
        elements.append(drawing)
        elements.append(Spacer(1, 8))

    add_multi_line_chart(
        'Courbe quotidienne HT / taxes / TTC',
        stats['daily']['labels'],
        [
            ('HT', stats['daily']['ht'], colors.HexColor('#198754')),
            ('Taxes', stats['daily']['taxes'], colors.HexColor('#FFC107')),
            ('TTC', stats['daily']['ttc'], colors.HexColor('#DC3545'))
        ]
    )
    add_grouped_vertical_chart(
        'Valeur mensuelle HT / taxes / TTC',
        stats['monthly']['labels'],
        [
            ('HT', stats['monthly']['ht'], colors.HexColor('#198754')),
            ('Taxes', stats['monthly']['taxes'], colors.HexColor('#FFC107')),
            ('TTC', stats['monthly']['ttc'], colors.HexColor('#DC3545'))
        ]
    )
    add_horizontal_chart('Top produits par perte TTC', stats['top_products'])
    add_horizontal_chart('Top fournisseurs par perte TTC', stats['top_suppliers'])
    add_pie_chart('Repartition par raison', stats['top_reasons'])

    def add_top_table(title, rows):
        elements.append(Paragraph(title, styles['Heading3']))
        data = [['Libelle', 'Sorties', 'Qte', 'TTC']]
        data.extend([[row['label'], row['count'], row.get('quantity', 0), f"{row['ttc']:.2f}"] for row in rows[:8]])
        elements.append(Table(data, repeatRows=1, colWidths=[290, 60, 60, 90], style=[
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#374151')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('GRID', (0, 0), (-1, -1), 0.25, colors.HexColor('#D1D5DB')),
            ('FONTSIZE', (0, 0), (-1, -1), 7),
            ('TOPPADDING', (0, 0), (-1, -1), 2),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 2),
        ]))
        elements.append(Spacer(1, 8))

    add_top_table('Top produits', stats['top_products'])
    add_top_table('Top fournisseurs', stats['top_suppliers'])
    add_top_table('Top raisons', stats['top_reasons'])
    add_top_table('Top utilisateurs sortie', stats['top_exit_users'])

    doc.build(elements)
    output.seek(0)
    filename = f'stats_sorties_stock_{generated_at.strftime("%Y%m%d_%H%M")}.pdf'
    return send_file(output, download_name=filename, as_attachment=True)

def get_exit_log_prices(item):
    prix_unite_ht = float(item.prix_unite_ht or 0)
    prix_sous_unite_ht = float(item.prix_sous_unite_ht or 0)
    prix_sous_sous_unite_ht = float(item.prix_sous_sous_unite_ht or 0)
    prix_unite_ttc = float(item.prix_unite_ttc or 0)
    prix_sous_unite_ttc = float(item.prix_sous_unite_ttc or 0)
    prix_sous_sous_unite_ttc = float(item.prix_sous_sous_unite_ttc or 0)

    total_ht = float(item.total_sortie_ht or 0)
    total_ttc = float(item.total_sortie_ttc or 0)
    if not total_ht:
        total_ht = (
            item.quantite_unites_sortie * prix_unite_ht
            + item.quantite_sous_unites_sortie * prix_sous_unite_ht
            + item.quantite_sous_sous_unites_sortie * prix_sous_sous_unite_ht
        )
    if not total_ttc:
        total_ttc = (
            item.quantite_unites_sortie * prix_unite_ttc
            + item.quantite_sous_unites_sortie * prix_sous_unite_ttc
            + item.quantite_sous_sous_unites_sortie * prix_sous_sous_unite_ttc
        )

    return {
        'prix_unite_ht': prix_unite_ht,
        'prix_sous_unite_ht': prix_sous_unite_ht,
        'prix_sous_sous_unite_ht': prix_sous_sous_unite_ht,
        'prix_unite_ttc': prix_unite_ttc,
        'prix_sous_unite_ttc': prix_sous_unite_ttc,
        'prix_sous_sous_unite_ttc': prix_sous_sous_unite_ttc,
        'tva_pourcentage': float(item.tva_pourcentage or 0),
        'total_ht': total_ht,
        'total_ttc': total_ttc
    }

def get_exit_log_supplier_info(item):
    return {
        'fournisseur': item.fournisseur_nom or '-',
        'groupe_fournisseur': item.groupe_fournisseur_nom or '-'
    }

def build_stock_exit_log_rows(exits):
    rows = []
    for item in exits:
        prices = get_exit_log_prices(item)
        supplier = get_exit_log_supplier_info(item)
        rows.append({
            'date_sortie': item.created_at.strftime('%d/%m/%Y %H:%M') if item.created_at else '-',
            'produit': item.produit_nom,
            'code_produit': item.produit_code,
            'fournisseur': supplier['fournisseur'],
            'groupe_fournisseur': supplier['groupe_fournisseur'],
            'code_suivi': item.code_suivi,
            'numero_bl': item.numero_bl,
            'date_peremption': item.date_peremption.strftime('%d/%m/%Y') if item.date_peremption else '-',
            'date_mise_en_stock': item.mise_en_stock_at.strftime('%d/%m/%Y %H:%M') if item.mise_en_stock_at else '-',
            'mis_en_stock_par': (
                f'{item.mise_en_stock_user_prenom or ""} {item.mise_en_stock_user_nom or ""}'.strip()
                or '-'
            ),
            'mis_en_stock_email': item.mise_en_stock_user_email or '-',
            'sorti_par': f'{item.user_prenom} {item.user_nom}',
            'email': item.user_email,
            'raison': item.reason_nom,
            'sortie': f'U:{item.quantite_unites_sortie} S/U:{item.quantite_sous_unites_sortie} SS/U:{item.quantite_sous_sous_unites_sortie}',
            'prix_ht': f'U:{prices["prix_unite_ht"]:.2f} S/U:{prices["prix_sous_unite_ht"]:.2f} SS/U:{prices["prix_sous_sous_unite_ht"]:.2f}',
            'prix_ttc': f'U:{prices["prix_unite_ttc"]:.2f} S/U:{prices["prix_sous_unite_ttc"]:.2f} SS/U:{prices["prix_sous_sous_unite_ttc"]:.2f}',
            'tva': f'{prices["tva_pourcentage"]:.2f}',
            'total_ht': f'{prices["total_ht"]:.2f}',
            'total_ttc': f'{prices["total_ttc"]:.2f}',
            'avant': f'U:{item.old_quantite_unites} S/U:{item.old_quantite_sous_unites} SS/U:{item.old_quantite_sous_sous_unites}',
            'apres': f'U:{item.new_quantite_unites} S/U:{item.new_quantite_sous_unites} SS/U:{item.new_quantite_sous_sous_unites}'
        })
    return rows

def get_stock_exit_logs_for_export():
    exits = StockExitLog.query.order_by(StockExitLog.created_at.desc()).all()
    query = (request.args.get('q') or '').strip().lower()
    sort_field = request.args.get('sort') or 'date'
    sort_direction = request.args.get('direction') or 'desc'

    if query:
        exits = [
            item for item in exits
            if query in ' '.join([
                item.created_at.strftime('%d/%m/%Y %H:%M') if item.created_at else '',
                item.produit_nom or '',
                item.produit_code or '',
                get_exit_log_supplier_info(item)['fournisseur'],
                get_exit_log_supplier_info(item)['groupe_fournisseur'],
                item.code_suivi or '',
                item.numero_bl or '',
                item.date_peremption.strftime('%d/%m/%Y') if item.date_peremption else '',
                item.mise_en_stock_at.strftime('%d/%m/%Y %H:%M') if item.mise_en_stock_at else '',
                item.mise_en_stock_user_prenom or '',
                item.mise_en_stock_user_nom or '',
                item.mise_en_stock_user_email or '',
                item.user_prenom or '',
                item.user_nom or '',
                item.user_email or '',
                item.reason_nom or '',
                f'{get_exit_log_prices(item)["tva_pourcentage"]:.2f}',
                f'{get_exit_log_prices(item)["total_ht"]:.2f}',
                f'{get_exit_log_prices(item)["total_ttc"]:.2f}'
            ]).lower()
        ]

    sort_getters = {
        'date': lambda item: item.created_at or datetime.min,
        'produit': lambda item: (item.produit_nom or '').lower(),
        'fournisseur': lambda item: get_exit_log_supplier_info(item)['fournisseur'].lower(),
        'groupe': lambda item: get_exit_log_supplier_info(item)['groupe_fournisseur'].lower(),
        'code': lambda item: (item.code_suivi or '').lower(),
        'bl': lambda item: (item.numero_bl or '').lower(),
        'peremption': lambda item: item.date_peremption or datetime.min.date(),
        'mise_stock': lambda item: item.mise_en_stock_at or datetime.min,
        'mis_par': lambda item: f'{item.mise_en_stock_user_prenom or ""} {item.mise_en_stock_user_nom or ""} {item.mise_en_stock_user_email or ""}'.lower(),
        'auteur': lambda item: f'{item.user_prenom or ""} {item.user_nom or ""} {item.user_email or ""}'.lower(),
        'raison': lambda item: (item.reason_nom or '').lower(),
        'perte': lambda item: get_exit_log_prices(item)['total_ttc']
    }
    exits.sort(key=sort_getters.get(sort_field, sort_getters['date']), reverse=sort_direction != 'asc')
    return exits

@admin.route('/stock/exits/export/excel')
@login_required
@permission_required('gestion_modifications_stock')
def export_stock_exit_logs_excel():
    import io
    import pandas as pd
    from flask import send_file
    from openpyxl.styles import Alignment, Font, PatternFill, Border, Side

    exits = get_stock_exit_logs_for_export()
    rows = build_stock_exit_log_rows(exits)
    generated_at = datetime.now()
    output = io.BytesIO()

    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df = pd.DataFrame(rows)
        df = df.rename(columns={
            'date_sortie': 'Date sortie',
            'produit': 'Produit',
            'code_produit': 'Code produit',
            'fournisseur': 'Fournisseur',
            'groupe_fournisseur': 'Groupe fournisseur',
            'code_suivi': 'Code suivi',
            'numero_bl': 'BL',
            'date_peremption': 'Peremption',
            'date_mise_en_stock': 'Date mise en stock',
            'mis_en_stock_par': 'Mis en stock par',
            'mis_en_stock_email': 'Email mise en stock',
            'sorti_par': 'Sorti par',
            'email': 'Email auteur',
            'raison': 'Raison',
            'sortie': 'Quantite sortie',
            'prix_ht': 'Prix unitaires HT',
            'prix_ttc': 'Prix unitaires TTC',
            'tva': 'TVA %',
            'total_ht': 'Total sortie HT',
            'total_ttc': 'Perte / valeur TTC',
            'avant': 'Avant',
            'apres': 'Apres'
        })
        df.to_excel(writer, index=False, sheet_name='Sorties stock', startrow=4)
        worksheet = writer.sheets['Sorties stock']
        worksheet['A1'] = 'Historique des sorties de stock - ReflexPharma'
        worksheet['A2'] = f'Date du tirage : {generated_at.strftime("%d/%m/%Y %H:%M")}'
        worksheet['A3'] = f'Tire par : {current_user.nom} {current_user.prenom}'
        worksheet['A4'] = f'Lignes exportees : {len(rows)}'

        header_font = Font(bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='1F2937', end_color='1F2937', fill_type='solid')
        meta_font = Font(size=10, color='374151')
        thin_border = Border(bottom=Side(style='thin', color='D1D5DB'))

        worksheet['A1'].font = Font(bold=True, size=14, color='1F2937')
        for row_number in range(2, 5):
            worksheet[f'A{row_number}'].font = meta_font
        for cell in worksheet[5]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        for row in worksheet.iter_rows(min_row=6, max_row=worksheet.max_row):
            for cell in row:
                cell.border = thin_border
                cell.alignment = Alignment(vertical='top', wrap_text=True)

        widths = [17, 24, 15, 22, 22, 34, 14, 13, 17, 20, 26, 20, 26, 20, 21, 24, 24, 10, 16, 18, 21, 21]
        for index, width in enumerate(widths, start=1):
            worksheet.column_dimensions[chr(64 + index)].width = width
        worksheet.freeze_panes = 'A6'

    output.seek(0)
    filename = f'sorties_stock_{generated_at.strftime("%Y%m%d_%H%M")}.xlsx'
    return send_file(output, download_name=filename, as_attachment=True)

@admin.route('/stock/exits/export/pdf')
@login_required
@permission_required('gestion_modifications_stock')
def export_stock_exit_logs_pdf():
    import io
    from flask import send_file
    from reportlab.lib import colors
    from reportlab.lib.enums import TA_CENTER
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
    from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle
    from xml.sax.saxutils import escape

    exits = get_stock_exit_logs_for_export()
    rows = build_stock_exit_log_rows(exits)
    generated_at = datetime.now()
    output = io.BytesIO()
    doc = SimpleDocTemplate(output, pagesize=landscape(A4), topMargin=10, bottomMargin=10, leftMargin=10, rightMargin=10)
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle('ExitLogTitle', parent=styles['Title'], fontSize=11, leading=13, spaceAfter=3)
    meta_style = ParagraphStyle('ExitLogMeta', parent=styles['Normal'], fontSize=7, leading=8, textColor=colors.HexColor('#374151'))
    cell_style = ParagraphStyle('ExitLogCell', parent=styles['Normal'], fontSize=5.5, leading=6.4)
    header_style = ParagraphStyle('ExitLogHeader', parent=cell_style, alignment=TA_CENTER, textColor=colors.white)

    elements = [
        Paragraph('Historique des sorties de stock - ReflexPharma', title_style),
        Paragraph(
            f'Date du tirage : {generated_at.strftime("%d/%m/%Y %H:%M")} | Tire par : {current_user.nom} {current_user.prenom} | Lignes : {len(rows)}',
            meta_style
        ),
        Spacer(1, 4)
    ]

    data = [[
        Paragraph('Date', header_style),
        Paragraph('Produit', header_style),
        Paragraph('Fourn.', header_style),
        Paragraph('Groupe', header_style),
        Paragraph('Code suivi', header_style),
        Paragraph('BL', header_style),
        Paragraph('Peremp.', header_style),
        Paragraph('Mise stock', header_style),
        Paragraph('Mis par', header_style),
        Paragraph('Auteur', header_style),
        Paragraph('Raison', header_style),
        Paragraph('Sortie', header_style),
        Paragraph('TVA', header_style),
        Paragraph('Prix TTC', header_style),
        Paragraph('Perte TTC', header_style),
        Paragraph('Avant', header_style),
        Paragraph('Apres', header_style)
    ]]
    for row in rows:
        produit = escape(str(row['produit']))
        code_produit = escape(str(row['code_produit']))
        data.append([
            row['date_sortie'],
            Paragraph(f"{produit}<br/>{code_produit}", cell_style),
            Paragraph(escape(str(row['fournisseur'])), cell_style),
            Paragraph(escape(str(row['groupe_fournisseur'])), cell_style),
            Paragraph(escape(str(row['code_suivi'])), cell_style),
            Paragraph(escape(str(row['numero_bl'])), cell_style),
            row['date_peremption'],
            row['date_mise_en_stock'],
            Paragraph(escape(str(row['mis_en_stock_par'])), cell_style),
            Paragraph(escape(str(row['sorti_par'])), cell_style),
            Paragraph(escape(str(row['raison'])), cell_style),
            row['sortie'],
            f"{row['tva']}%",
            row['prix_ttc'],
            row['total_ttc'],
            row['avant'],
            row['apres']
        ])

    table = Table(data, repeatRows=1, colWidths=[42, 68, 54, 50, 76, 36, 32, 47, 47, 48, 48, 45, 43, 58, 39, 42, 42])
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1F2937')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 5.5),
        ('LEADING', (0, 0), (-1, -1), 6.4),
        ('ALIGN', (0, 1), (0, -1), 'CENTER'),
        ('ALIGN', (3, 1), (4, -1), 'CENTER'),
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ('GRID', (0, 0), (-1, -1), 0.25, colors.HexColor('#D1D5DB')),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F9FAFB')]),
        ('TOPPADDING', (0, 0), (-1, -1), 1.4),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 1.4),
        ('LEFTPADDING', (0, 0), (-1, -1), 1.5),
        ('RIGHTPADDING', (0, 0), (-1, -1), 1.5),
    ]))
    elements.append(table)
    doc.build(elements)
    output.seek(0)
    filename = f'sorties_stock_{generated_at.strftime("%Y%m%d_%H%M")}.pdf'
    return send_file(output, download_name=filename, as_attachment=True)

# --- VUES FILTRÉES PRODUITS ---
@admin.route('/produits/rayon/<int:id>')
@login_required
@permission_required('gestion_produits')
def list_produits_by_rayon(id):
    rayon = Rayon.query.get_or_404(id)
    produits = Produit.query.filter_by(rayon_id=id).all()
    return render_template('admin/produits/list.html', produits=produits, title=f'Produits - Rayon : {rayon.nom}')

@admin.route('/produits/famille/<int:id>')
@login_required
@permission_required('gestion_produits')
def list_produits_by_famille(id):
    famille = Famille.query.get_or_404(id)
    produits = Produit.query.filter_by(famille_id=id).all()
    return render_template('admin/produits/list.html', produits=produits, title=f'Produits - Famille : {famille.nom}')

@admin.route('/produits/section/<int:id>')
@login_required
@permission_required('gestion_produits')
def list_produits_by_section(id):
    section = Section.query.get_or_404(id)
    produits = Produit.query.filter_by(section_id=id).all()
    return render_template('admin/produits/list.html', produits=produits, title=f'Produits - Section : {section.nom}')

import pandas as pd
from flask import send_file
import io
from reportlab.lib.pagesizes import letter
from reportlab.pdfgen import canvas

@admin.route('/produits/export/excel')
@login_required
@permission_required('gestion_produits')
def export_produits_excel():
    produits = Produit.query.all()
    timestamp = datetime.now().strftime("%Y-%m-%d %H:%M")
    data = [{
        'Code': p.code_produit,
        'Nom': p.nom,
        'Fournisseur': p.fournisseur.nom if p.fournisseur else '',
        'Rayon': p.rayon.nom if p.rayon else '',
        'Famille': p.famille.nom if p.famille else '',
        'Conditionnement': p.conditionnement,
        'Prix Unité': p.prix_unite
    } for p in produits]
    
    df = pd.DataFrame(data)
    output = io.BytesIO()
    
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Produits', startrow=2)
        workbook = writer.book
        worksheet = writer.sheets['Produits']
        
        # Add metadata
        worksheet['A1'] = f"Rapport généré le : {timestamp}"
        worksheet['A2'] = f"Tiré par : {current_user.nom} {current_user.prenom}"
        
        # Style
        from openpyxl.styles import Font, PatternFill, Alignment
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        
        for cell in worksheet[3]: # Header row is now 3
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
            
    output.seek(0)
    return send_file(output, download_name=f'produits_{datetime.now().strftime("%Y%m%d_%H%M")}.xlsx', as_attachment=True)

@admin.route('/produits/export/pdf')
@login_required
@permission_required('gestion_produits')
def export_produits_pdf():
    from reportlab.lib import colors
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.lib.pagesizes import A4
    
    produits = Produit.query.all()
    output = io.BytesIO()
    doc = SimpleDocTemplate(output, pagesize=A4, topMargin=20, bottomMargin=20, leftMargin=20, rightMargin=20)
    elements = []
    styles = getSampleStyleSheet()
    
    elements.append(Paragraph(f'Catalogue Produits - ReflexPharma', styles['Title']))
    elements.append(Paragraph(f'Tiré par : {current_user.nom} {current_user.prenom} | Date : {datetime.now().strftime("%d/%m/%Y %H:%M")}', styles['Normal']))
    elements.append(Spacer(1, 12))
    
    data = [['Code', 'Nom', 'Fournisseur', 'Prix']]
    for p in produits:
        data.append([p.code_produit, p.nom, p.fournisseur.nom if p.fournisseur else '-', f'{p.prix_unite} €'])
        
    table = Table(data, repeatRows=1, colWidths=[100, 200, 150, 80])
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
    ]))
    
    elements.append(table)
    doc.build(elements)
    output.seek(0)
    return send_file(output, download_name=f'produits_{datetime.now().strftime("%Y%m%d_%H%M")}.pdf', as_attachment=True)

# --- GESTION DES PARAMÈTRES ---
@admin.route('/settings', methods=['GET', 'POST'])
@login_required
@permission_required('gestion_parametres')
def app_settings():
    if request.method == 'POST':
        pharmacy_name = request.form.get('pharmacy_name')
        if pharmacy_name:
            Setting.set_value('pharmacy_name', pharmacy_name)

        if request.form.get('form_name') == 'printer':
            auto_print_enabled = 'true' if request.form.get('auto_print_enabled') else 'false'
            Setting.set_value('auto_print_enabled', auto_print_enabled)

        flash('Paramètres mis à jour avec succès.', 'success')
        return redirect(url_for('admin.app_settings'))

    settings = {
        'pharmacy_name': Setting.get_value('pharmacy_name', 'REFLEXPHARMA'),
        'auto_print_enabled': Setting.get_value('auto_print_enabled', 'true') == 'true'
    }
    return render_template('admin/settings.html', settings=settings)


# ==============================================================================
# ASSISTANT IA (MASCOTTE)
# ==============================================================================

_JOURS_FR = ['lundi', 'mardi', 'mercredi', 'jeudi', 'vendredi', 'samedi', 'dimanche']
_MOIS_FR = ['janvier', 'février', 'mars', 'avril', 'mai', 'juin', 'juillet',
            'août', 'septembre', 'octobre', 'novembre', 'décembre']

def build_assistant_system_prompt():
    today = datetime.now()
    date_str = f"{_JOURS_FR[today.weekday()]} {today.day} {_MOIS_FR[today.month - 1]} {today.year}"
    return (
        "Tu es l'assistante virtuelle de ReflexPharma, un logiciel de gestion de pharmacie "
        "(stock, ventes, clients, groupes clients, fournisseurs, inventaires, statistiques, "
        "déclarations d'impôts/taxes). "
        f"Nous sommes le {date_str}. "
        "Tu aides le personnel de la pharmacie de deux façons : "
        "1) l'utilisation du logiciel (où trouver une fonctionnalité, comment faire une action, comprendre un chiffre affiché) ; "
        "2) l'analyse des données réelles de la pharmacie (ventes, chiffre d'affaires, stock, clients, employés) "
        "en utilisant les outils mis à ta disposition. "
        "Dès qu'une question porte sur des chiffres ou des données concrètes (ventes, chiffre d'affaires, stock, "
        "employé du mois, clients, prévisions...), appelle systématiquement l'outil le plus pertinent plutôt que "
        "de deviner ou d'inventer une réponse. Tu peux enchaîner plusieurs appels d'outils si besoin (par exemple "
        "comparer deux périodes). Ne donne JAMAIS un chiffre, un nom (produit, client, employé...), une liste "
        "ou toute autre donnée factuelle que tu n'as pas obtenue via un outil — y compris dans une question de "
        "relance sur un message précédent (ex: après avoir donné un nombre de produits/clients, si on te demande "
        "'lesquels ?', appelle l'outil de listing correspondant plutôt que d'inventer des noms plausibles). "
        "En cas de doute, appelle un outil ou dis que tu ne sais pas plutôt que d'inventer. "
        "Important : les paramètres 'recherche' des outils de listing (produits, clients, fournisseurs, "
        "groupes...) ne filtrent QUE sur du texte (nom, code, matricule...), jamais sur des valeurs "
        "numériques (coefficient, TVA, prix, solde, quantité...). Si on te demande de retrouver un élément "
        "par une valeur numérique (ex: 'quel fournisseur a le coefficient 1.34', 'quel client a un solde de "
        "38560'), appelle l'outil de listing SANS filtre pour récupérer tous les éléments, puis identifie "
        "toi-même celui qui correspond dans les résultats — ne renvoie jamais 'aucun résultat' simplement "
        "parce que le filtre texte ne matchait pas un nombre. Ce principe s'applique à toute variante de "
        "formulation d'une recherche, pas seulement aux exemples ci-dessus. "
        "Les montants sont en euros (€). Réponds de façon concise et claire, avec les chiffres clés mis en avant "
        "(des listes à puces pour les classements, pas de longs paragraphes). "
        "Si l'utilisateur demande explicitement un PDF, un rapport, un document ou une impression des résultats, "
        "récupère d'abord les données via le(s) outil(s) pertinent(s) puis appelle generer_rapport_pdf avec ces "
        "données pour produire un fichier téléchargeable ; ne propose jamais un PDF si ce n'est pas demandé. "
        "Après avoir appelé generer_rapport_pdf, ne redonne PAS l'URL ni un lien de téléchargement dans ta "
        "réponse texte : un bouton de téléchargement s'affiche déjà automatiquement sous ton message. Contente-toi "
        "de confirmer brièvement que le PDF est prêt (et résume les chiffres clés si utile). "
        "Réponds toujours en français. "
        "Les outils de données auxquels tu as accès dépendent des permissions de la personne qui te parle : "
        "certains modules peuvent ne pas t'être proposés du tout dans la liste d'outils disponibles, ou un "
        "outil peut renvoyer une erreur d'accès refusé. N'appelle JAMAIS un outil qui n'est pas explicitement "
        "dans la liste qui t'est fournie pour cette conversation, même s'il existait dans un échange "
        "précédent ou te semble logique — un nom d'outil absent de ta liste actuelle n'existe pas pour cet "
        "utilisateur. Si aucun outil disponible ne permet de répondre à la question, ne tente rien d'autre : "
        "réponds directement en texte que cette information nécessite une permission que cette personne n'a "
        "pas, sans détailler de donnée, et suggère de contacter un administrateur. Idem si un outil renvoie "
        "une erreur d'accès refusé : n'insiste pas et n'essaie pas de contourner via un autre outil. "
        "Si la question sort totalement du cadre du logiciel ou de la pharmacie, réponds brièvement puis "
        "recentre poliment sur ce que tu peux faire ici."
    )

@admin.route('/assistant/chat', methods=['POST'])
@login_required
def assistant_chat():
    data = request.get_json(silent=True) or {}
    message = (data.get('message') or '').strip()
    history = data.get('history') or []

    if not message:
        return {'success': False, 'message': 'Message vide.'}, 400
    if len(message) > 2000:
        return {'success': False, 'message': 'Message trop long.'}, 400

    api_key = current_app.config.get('MISTRAL_API_KEY')
    if not api_key:
        return {'success': False, 'message': "L'assistant IA n'est pas configuré (clé API Mistral manquante)."}, 503

    messages = [{'role': 'system', 'content': build_assistant_system_prompt()}]
    if isinstance(history, list):
        for item in history[-12:]:
            if not isinstance(item, dict):
                continue
            role = item.get('role')
            content = (item.get('content') or '').strip()
            if role in ('user', 'assistant') and content:
                messages.append({'role': role, 'content': content[:2000]})
    messages.append({'role': 'user', 'content': message})

    request_payload = {
        'model': 'mistral-small-latest',
        'messages': messages,
        'tools': AI_TOOLS,
        'tool_choice': 'auto',
        'temperature': 0.3,
        'max_tokens': 800,
    }

    pdf_attachment = None
    try:
        for _ in range(4):  # limite le nombre d'aller-retours d'appels d'outils
            response = requests.post(
                'https://api.mistral.ai/v1/chat/completions',
                headers={
                    'Authorization': f'Bearer {api_key}',
                    'Content-Type': 'application/json',
                },
                json=request_payload,
                timeout=30,
            )
            response.raise_for_status()
            payload = response.json()
            assistant_message = payload['choices'][0]['message']
            tool_calls = assistant_message.get('tool_calls')

            if not tool_calls:
                reply = (assistant_message.get('content') or '').strip()
                if not reply:
                    return {'success': False, 'message': "Réponse vide de l'assistant IA."}, 502
                return {'success': True, 'reply': reply, 'attachment': pdf_attachment}

            messages.append(assistant_message)
            for tool_call in tool_calls:
                function_name = (tool_call.get('function') or {}).get('name', '')
                raw_arguments = (tool_call.get('function') or {}).get('arguments') or '{}'
                try:
                    arguments = json.loads(raw_arguments)
                except (ValueError, TypeError):
                    arguments = {}
                result = call_ai_tool(function_name, arguments, current_user)
                if function_name == 'generer_rapport_pdf' and isinstance(result, dict) and result.get('pdf_genere'):
                    pdf_attachment = {
                        'url': result.get('url_telechargement'),
                        'titre': result.get('titre'),
                        'nom_fichier': result.get('nom_fichier'),
                    }
                messages.append({
                    'role': 'tool',
                    'tool_call_id': tool_call.get('id'),
                    'name': function_name,
                    'content': json.dumps(result, ensure_ascii=False, default=str),
                })

        return {'success': False, 'message': "L'assistant n'a pas réussi à conclure sa recherche, reformulez votre question."}, 502
    except requests.exceptions.RequestException:
        return {'success': False, 'message': "Impossible de contacter l'assistant IA pour le moment."}, 502
    except (KeyError, IndexError, ValueError):
        return {'success': False, 'message': "Réponse inattendue de l'assistant IA."}, 502


@admin.route('/assistant/rapport/<path:filename>')
@login_required
def assistant_download_report(filename):
    """Sert un PDF genere par l'assistant IA. Le nom de fichier contient un token
    aleatoire (uuid4) genere a la creation : non devinable, donc seul un lien recu
    dans le chat permet d'y acceder (en plus de la connexion requise)."""
    if not REPORT_FILENAME_RE.match(filename):
        abort(404)
    filepath = os.path.join(REPORTS_DIR, filename)
    if not os.path.isfile(filepath):
        abort(404)
    display_name = filename.split('__', 1)[1]
    return send_file(filepath, mimetype='application/pdf', as_attachment=True, download_name=display_name)


# ==============================================================================
# GESTION DE L'INVENTAIRE
# ==============================================================================

# Diffusion en direct des saisies d'inventaire à tous les appareils connectés (SSE).
# Registre en mémoire : inventaire_id -> liste de queues (une par appareil connecté).
_inventaire_subscribers = {}
_inventaire_subscribers_lock = threading.Lock()

def _inventaire_subscribe(inventaire_id):
    q = queue.Queue()
    with _inventaire_subscribers_lock:
        _inventaire_subscribers.setdefault(inventaire_id, []).append(q)
    return q

def _inventaire_unsubscribe(inventaire_id, q):
    with _inventaire_subscribers_lock:
        subs = _inventaire_subscribers.get(inventaire_id)
        if subs and q in subs:
            subs.remove(q)
            if not subs:
                _inventaire_subscribers.pop(inventaire_id, None)

def _inventaire_publish(inventaire_id, payload):
    with _inventaire_subscribers_lock:
        subs = list(_inventaire_subscribers.get(inventaire_id, []))
    for q in subs:
        q.put(payload)

def _inventaire_line_payload(line):
    return {
        'id': line.id,
        'is_scanned': line.is_scanned,
        'a_decalage': line.a_decalage,
        'quantite_unites_apres': line.quantite_unites_apres,
        'quantite_sous_unites_apres': line.quantite_sous_unites_apres,
        'quantite_sous_sous_unites_apres': line.quantite_sous_sous_unites_apres,
        'quantite_unites_avant': line.quantite_unites_avant,
        'quantite_sous_unites_avant': line.quantite_sous_unites_avant,
        'quantite_sous_sous_unites_avant': line.quantite_sous_sous_unites_avant,
        'constate_by': f"{line.constate_by.prenom} {line.constate_by.nom}" if line.constate_by else None,
        'constate_at': line.constate_at.strftime('%d/%m %H:%M') if line.constate_at else None,
    }

def _snapshot_stock_into_inventaire(inventaire):
    """Prend un instantane des quantites actuelles de tout le stock pour servir de
    base theorique de comparaison a cet inventaire (appele au demarrage reel, pas a
    la programmation, pour que les quantites 'avant' soient a jour)."""
    stocks = Stock.query.all()
    for s in stocks:
        ligne = InventaireLigne(
            inventaire_id=inventaire.id,
            stock_id=s.id,
            produit_id=s.produit_id,
            code_suivi=s.code_suivi,
            numero_bl=s.numero_bl,
            date_peremption=s.date_peremption,
            quantite_unites_avant=s.quantite_unites,
            quantite_sous_unites_avant=s.quantite_sous_unites,
            quantite_sous_sous_unites_avant=s.quantite_sous_sous_unites
        )
        db.session.add(ligne)


def _activer_inventaires_planifies():
    """Demarre automatiquement le prochain inventaire programme dont l'heure est
    arrivee, s'il n'y a pas deja un inventaire en cours. Il n'y a pas de tache de
    fond dans cette appli : on verifie donc a chaque chargement d'une page
    inventaire/dashboard, comme le reste du "self-healing" de l'appli."""
    if Inventaire.query.filter_by(statut='en_cours').first():
        return
    due = Inventaire.query.filter(
        Inventaire.statut == 'planifie',
        Inventaire.date_planifiee <= datetime.now()
    ).order_by(Inventaire.date_planifiee.asc()).first()
    if not due:
        return
    due.statut = 'en_cours'
    _snapshot_stock_into_inventaire(due)
    db.session.commit()

@admin.route('/inventaire')
@login_required
@permission_required('gestion_inventaire')
def list_inventaires():
    _activer_inventaires_planifies()
    active_inventaire = Inventaire.query.filter_by(statut='en_cours').first()
    planned_inventaire = Inventaire.query.filter_by(statut='planifie').order_by(Inventaire.date_planifiee.asc()).first()
    inventaires = Inventaire.query.order_by(Inventaire.created_at.desc()).all()
    return render_template(
        'admin/inventaire/list.html',
        active_inventaire=active_inventaire,
        planned_inventaire=planned_inventaire,
        inventaires=inventaires
    )

@admin.route('/inventaire/create', methods=['POST'])
@login_required
@permission_required('gestion_inventaire')
def create_inventaire():
    active = Inventaire.query.filter_by(statut='en_cours').first()
    if active:
        flash("Un inventaire est déjà en cours.", "warning")
        return redirect(url_for('admin.show_inventaire', id=active.id))

    titre = (request.form.get('titre') or '').strip()
    date_planifiee_str = (request.form.get('date_planifiee') or '').strip()

    if date_planifiee_str:
        planned = Inventaire.query.filter_by(statut='planifie').first()
        if planned:
            flash("Un inventaire est déjà programmé. Annulez-le avant d'en programmer un autre.", "warning")
            return redirect(url_for('admin.list_inventaires'))

        try:
            date_planifiee = datetime.strptime(date_planifiee_str, '%Y-%m-%dT%H:%M')
        except ValueError:
            flash("Date/heure de programmation invalide.", "danger")
            return redirect(url_for('admin.list_inventaires'))
        if date_planifiee <= datetime.now():
            flash("La date programmée doit être dans le futur. Démarrez l'inventaire immédiatement à la place.", "warning")
            return redirect(url_for('admin.list_inventaires'))

        if not titre:
            titre = f"Inventaire programmé du {date_planifiee.strftime('%d/%m/%Y %H:%M')}"

        new_inv = Inventaire(titre=titre, statut='planifie', created_by_id=current_user.id, date_planifiee=date_planifiee)
        db.session.add(new_inv)
        db.session.commit()
        flash("Inventaire programmé avec succès.", "success")
        return redirect(url_for('admin.list_inventaires'))

    if not titre:
        titre = f"Inventaire du {datetime.now().strftime('%d/%m/%Y %H:%M')}"

    new_inv = Inventaire(titre=titre, statut='en_cours', created_by_id=current_user.id)
    db.session.add(new_inv)
    db.session.flush()

    _snapshot_stock_into_inventaire(new_inv)

    db.session.commit()
    flash("Nouvel inventaire démarré avec succès.", "success")
    return redirect(url_for('admin.show_inventaire', id=new_inv.id))

@admin.route('/inventaire/<int:id>')
@login_required
@permission_required('gestion_inventaire')
def show_inventaire(id):
    inventaire = Inventaire.query.get_or_404(id)
    
    # Récupérer toutes les lignes pour permettre la recherche et le filtrage côté client en JS
    lines = InventaireLigne.query.filter_by(inventaire_id=id).order_by(InventaireLigne.id.asc()).all()

    total_count = len(lines)
    scanned_count = sum(1 for line in lines if line.is_scanned)

    # Personnes ayant saisi au moins une ligne, pour le filtre du rapport PDF
    comptages_par = User.query.join(
        InventaireLigne, InventaireLigne.constate_by_id == User.id
    ).filter(InventaireLigne.inventaire_id == id).distinct().order_by(User.nom.asc()).all()

    return render_template(
        'admin/inventaire/show.html',
        inventaire=inventaire,
        lines=lines,
        total_count=total_count,
        scanned_count=scanned_count,
        comptages_par=comptages_par
    )

@admin.route('/inventaire/<int:id>/line/<int:line_id>/save', methods=['POST'])
@login_required
@permission_required('gestion_inventaire')
def save_inventaire_line(id, line_id):
    inventaire = Inventaire.query.get_or_404(id)
    if inventaire.statut != 'en_cours':
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return {'success': False, 'message': 'Inventaire déjà validé ou annulé.'}, 400
        flash("Impossible de modifier un inventaire fermé.", "danger")
        return redirect(url_for('admin.show_inventaire', id=id))
        
    line = InventaireLigne.query.filter_by(inventaire_id=id, id=line_id).first_or_404()
    
    try:
        u = request.form.get('quantite_unites')
        su = request.form.get('quantite_sous_unites')
        ssu = request.form.get('quantite_sous_sous_unites')
        
        line.quantite_unites_apres = int(u) if u is not None and u != '' else 0
        line.quantite_sous_unites_apres = int(su) if su is not None and su != '' else 0
        line.quantite_sous_sous_unites_apres = int(ssu) if ssu is not None and ssu != '' else 0
        
        line.is_scanned = True
        line.constate_at = datetime.utcnow()
        line.constate_by_id = current_user.id
        db.session.commit()

        # Diffuse la saisie en direct à tous les autres appareils connectés sur cet inventaire
        total_count = InventaireLigne.query.filter_by(inventaire_id=id).count()
        scanned_count = InventaireLigne.query.filter_by(inventaire_id=id, is_scanned=True).count()
        _inventaire_publish(id, {
            'line': _inventaire_line_payload(line),
            'total_count': total_count,
            'scanned_count': scanned_count,
        })

        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return {
                'success': True,
                'message': 'Ligne enregistrée avec succès.',
                'total_apres': line.total_apres,
                'a_decalage': line.a_decalage
            }
        flash("Ligne mise à jour avec succès.", "success")
    except Exception as e:
        db.session.rollback()
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return {'success': False, 'message': str(e)}, 400
        flash(f"Erreur lors de la mise à jour : {str(e)}", "danger")

    return redirect(url_for('admin.show_inventaire', id=id))

@admin.route('/inventaire/<int:id>/updates', methods=['GET'])
@login_required
@permission_required('gestion_inventaire')
def inventaire_updates(id):
    """Retourne les lignes saisies depuis 'since' pour synchroniser les autres appareils sans refresh."""
    Inventaire.query.get_or_404(id)

    query = InventaireLigne.query.filter_by(inventaire_id=id).filter(InventaireLigne.constate_at.isnot(None))

    since_str = request.args.get('since')
    if since_str:
        try:
            since_dt = datetime.fromisoformat(since_str.replace('Z', '+00:00')).replace(tzinfo=None)
            query = query.filter(InventaireLigne.constate_at > since_dt)
        except ValueError:
            pass

    lines = query.order_by(InventaireLigne.constate_at.asc()).all()

    total_count = InventaireLigne.query.filter_by(inventaire_id=id).count()
    scanned_count = InventaireLigne.query.filter_by(inventaire_id=id, is_scanned=True).count()

    return {
        'success': True,
        'server_time': datetime.utcnow().isoformat(),
        'total_count': total_count,
        'scanned_count': scanned_count,
        'lines': [_inventaire_line_payload(line) for line in lines]
    }

@admin.route('/inventaire/<int:id>/stream')
@login_required
@permission_required('gestion_inventaire')
def inventaire_stream(id):
    """Flux SSE : pousse chaque saisie aux autres appareils connectés, quasi instantanément."""
    Inventaire.query.get_or_404(id)
    q = _inventaire_subscribe(id)

    def gen():
        try:
            yield 'retry: 2000\n\n'
            while True:
                try:
                    payload = q.get(timeout=20)
                    yield f'data: {json.dumps(payload)}\n\n'
                except queue.Empty:
                    yield ': keep-alive\n\n'
        except GeneratorExit:
            pass
        finally:
            _inventaire_unsubscribe(id, q)

    return Response(gen(), mimetype='text/event-stream', headers={
        'Cache-Control': 'no-cache',
        'X-Accel-Buffering': 'no',
        'Connection': 'keep-alive',
    })

@admin.route('/inventaire/<int:id>/scan')
@login_required
@permission_required('gestion_inventaire')
def scan_inventaire_qrcode(id):
    inventaire = Inventaire.query.get_or_404(id)
    if inventaire.statut != 'en_cours':
        flash("Cet inventaire n'est plus en cours.", "warning")
        return redirect(url_for('admin.show_inventaire', id=id))
    return render_template('admin/inventaire/scan.html', inventaire=inventaire)

@admin.route('/inventaire/<int:id>/scan/lookup', methods=['GET'])
@login_required
@permission_required('gestion_inventaire')
def scan_lookup(id):
    inventaire = Inventaire.query.get_or_404(id)
    code = request.args.get('code', '').strip()
    if not code:
        return {'success': False, 'message': 'Code requis.'}, 400
        
    line = InventaireLigne.query.filter_by(inventaire_id=id).filter(
        (InventaireLigne.code_suivi == code)
    ).first()
    
    if not line:
        line = InventaireLigne.query.join(Produit).filter(
            (InventaireLigne.inventaire_id == id) &
            ((Produit.code_produit == code) | (InventaireLigne.code_suivi.ilike(f'%{code}%')))
        ).first()
        
    if not line:
        return {'success': False, 'message': 'Aucun lot ou produit correspondant trouvé dans cet inventaire.'}, 404
        
    return {
        'success': True,
        'line_id': line.id,
        'product_name': line.produit.nom,
        'code_produit': line.produit.code_produit,
        'code_suivi': line.code_suivi,
        'numero_bl': line.numero_bl,
        'date_peremption': line.date_peremption.strftime('%d/%m/%Y'),
        'conditionnement': line.produit.conditionnement,
        'quantite_unites_avant': line.quantite_unites_avant,
        'quantite_sous_unites_avant': line.quantite_sous_unites_avant,
        'quantite_sous_sous_unites_avant': line.quantite_sous_sous_unites_avant,
        'quantite_unites_apres': line.quantite_unites_apres if line.quantite_unites_apres is not None else '',
        'quantite_sous_unites_apres': line.quantite_sous_unites_apres if line.quantite_sous_unites_apres is not None else '',
        'quantite_sous_sous_unites_apres': line.quantite_sous_sous_unites_apres if line.quantite_sous_sous_unites_apres is not None else '',
        'is_scanned': line.is_scanned,
        'constate_by_id': line.constate_by_id,
        'constate_by': f"{line.constate_by.prenom} {line.constate_by.nom}" if line.constate_by else None,
        'constate_at': line.constate_at.strftime('%d/%m/%Y à %H:%M') if line.constate_at else None
    }

@admin.route('/inventaire/<int:id>/validate', methods=['POST'])
@login_required
@permission_required('gestion_inventaire')
def validate_inventaire(id):
    inventaire = Inventaire.query.get_or_404(id)
    if inventaire.statut != 'en_cours':
        flash("Inventaire déjà traité.", "danger")
        return redirect(url_for('admin.show_inventaire', id=id))
        
    non_saisis_option = request.form.get('non_saisis', 'theorique')
    lignes = InventaireLigne.query.filter_by(inventaire_id=id).all()
    
    for line in lignes:
        if (line.quantite_unites_apres is None and 
            line.quantite_sous_unites_apres is None and 
            line.quantite_sous_sous_unites_apres is None):
            
            if non_saisis_option == 'theorique':
                line.quantite_unites_apres = line.quantite_unites_avant
                line.quantite_sous_unites_apres = line.quantite_sous_unites_avant
                line.quantite_sous_sous_unites_apres = line.quantite_sous_sous_unites_avant
            else:
                line.quantite_unites_apres = 0
                line.quantite_sous_unites_apres = 0
                line.quantite_sous_sous_unites_apres = 0
                
            line.is_scanned = True
            line.constate_at = datetime.utcnow()
            line.constate_by_id = current_user.id
            
        if line.stock:
            stock = line.stock
            old_vals = (stock.quantite_unites, stock.quantite_sous_unites, stock.quantite_sous_sous_unites)
            new_vals = (line.quantite_unites_apres, line.quantite_sous_unites_apres, line.quantite_sous_sous_unites_apres)
            
            if old_vals != new_vals:
                stock.quantite_unites = line.quantite_unites_apres
                stock.quantite_sous_unites = line.quantite_sous_unites_apres
                stock.quantite_sous_sous_unites = line.quantite_sous_sous_unites_apres
                
                create_stock_modification(
                    stock=stock,
                    produit=stock.produit,
                    action='update',
                    reason=f'Ajustement par inventaire "{inventaire.titre}"',
                    old_values=old_vals,
                    new_values=new_vals,
                    old_qr_tire=stock.qr_tire,
                    new_qr_tire=stock.qr_tire
                )
                
    inventaire.statut = 'valide'
    inventaire.validated_at = datetime.utcnow()
    inventaire.validated_by_id = current_user.id
    
    db.session.commit()
    flash("L'inventaire a été validé et le stock mis à jour.", "success")
    return redirect(url_for('admin.show_inventaire', id=id))

@admin.route('/inventaire/<int:id>/cancel', methods=['POST'])
@login_required
@permission_required('gestion_inventaire')
def cancel_inventaire(id):
    inventaire = Inventaire.query.get_or_404(id)
    if inventaire.statut not in ('en_cours', 'planifie'):
        flash("Seul un inventaire en cours ou programmé peut être annulé.", "danger")
        return redirect(url_for('admin.show_inventaire', id=id))

    inventaire.statut = 'annule'
    db.session.commit()
    flash("L'inventaire a été annulé.", "warning")
    return redirect(url_for('admin.list_inventaires'))

@admin.route('/inventaire/<int:id>/export/comptage')
@login_required
@permission_required('gestion_inventaire')
def export_fiche_comptage_pdf(id):
    inventaire = Inventaire.query.get_or_404(id)
    from reportlab.lib import colors
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.pagesizes import A4
    import io
    
    lignes = InventaireLigne.query.filter_by(inventaire_id=id).all()
    output = io.BytesIO()
    doc = SimpleDocTemplate(output, pagesize=A4, topMargin=30, bottomMargin=30, leftMargin=30, rightMargin=30)
    elements = []
    
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'TitleStyle',
        parent=styles['Heading1'],
        fontSize=20,
        leading=24,
        textColor=colors.HexColor('#2c3e50'),
        alignment=1,
        spaceAfter=15
    )
    normal_style = styles['Normal']
    code_style = ParagraphStyle('CodeCell', parent=styles['Normal'], fontSize=6.5, leading=7.5, wordWrap='CJK')
    cell_style = ParagraphStyle('BodyCell', parent=styles['Normal'], fontSize=6.5, leading=7.5)

    elements.append(Paragraph(f"Fiche de Comptage - {inventaire.titre}", title_style))
    elements.append(Paragraph(f"Créé le : {inventaire.created_at.strftime('%d/%m/%Y %H:%M')} | Auteur : {inventaire.created_by.nom} {inventaire.created_by.prenom}", normal_style))
    elements.append(Spacer(1, 10))

    data = [['Code Lot', 'Produit', 'Emplacement / Rayon', 'Stock Théor.', 'Stock Réel']]
    for l in lignes:
        rayon_nom = l.produit.rayon.nom if l.produit.rayon else '-'
        cond = l.produit.conditionnement
        if cond == 3:
            th_str = f"U:{l.quantite_unites_avant}\nSU:{l.quantite_sous_unites_avant}\nSSU:{l.quantite_sous_sous_unites_avant}"
        elif cond == 2:
            th_str = f"U:{l.quantite_unites_avant}\nSU:{l.quantite_sous_unites_avant}"
        else:
            th_str = f"U:{l.quantite_unites_avant}"

        data.append([
            Paragraph(l.code_suivi, code_style),
            Paragraph(l.produit.nom, cell_style),
            Paragraph(rayon_nom, cell_style),
            th_str,
            '[  ] U\n' + ('[  ] SU\n' if cond >= 2 else '') + ('[  ] SSU' if cond == 3 else '')
        ])

    table = Table(data, repeatRows=1, colWidths=[120, 165, 105, 65, 80])
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2c3e50')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 7.5),
        ('FONTSIZE', (0, 1), (-1, -1), 6.5),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('LEFTPADDING', (0, 0), (-1, -1), 3),
        ('RIGHTPADDING', (0, 0), (-1, -1), 3),
        ('TOPPADDING', (0, 0), (-1, -1), 2.5),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 2.5),
    ]))

    elements.append(table)
    doc.build(elements)
    output.seek(0)
    return send_file(output, download_name=f"fiche_comptage_inv_{id}.pdf", as_attachment=True)

@admin.route('/inventaire/<int:id>/export/rapport')
@login_required
@permission_required('gestion_inventaire')
def export_rapport_inventaire_pdf(id):
    inventaire = Inventaire.query.get_or_404(id)
    from reportlab.lib import colors
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.pagesizes import A4
    import io

    sort_by = request.args.get('sort', 'code')
    personne_id = request.args.get('personne_id', '').strip()

    lignes = InventaireLigne.query.filter_by(inventaire_id=id).all()

    if personne_id:
        try:
            personne_id_int = int(personne_id)
            lignes = [l for l in lignes if l.constate_by_id == personne_id_int]
        except ValueError:
            pass

    def sort_key(l):
        if sort_by == 'personne':
            has_person = 0 if l.constate_by else 1
            name = f"{l.constate_by.nom} {l.constate_by.prenom}".lower() if l.constate_by else ''
            return (has_person, name, l.code_suivi.lower())
        elif sort_by == 'produit':
            return (l.produit.nom.lower(), l.code_suivi.lower())
        elif sort_by == 'ecart':
            return (0 if l.a_decalage else 1, l.code_suivi.lower())
        return (l.code_suivi.lower(),)

    lignes.sort(key=sort_key)

    personne_filtree = None
    if personne_id:
        try:
            personne_filtree = User.query.get(int(personne_id))
        except ValueError:
            pass

    output = io.BytesIO()
    doc = SimpleDocTemplate(output, pagesize=A4, topMargin=30, bottomMargin=30, leftMargin=30, rightMargin=30)
    elements = []
    
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'TitleStyle',
        parent=styles['Heading1'],
        fontSize=20,
        leading=24,
        textColor=colors.HexColor('#1abc9c'),
        alignment=1,
        spaceAfter=15
    )
    normal_style = styles['Normal']
    code_style = ParagraphStyle('CodeCell', parent=styles['Normal'], fontSize=6.5, leading=7.5, wordWrap='CJK')
    cell_style = ParagraphStyle('BodyCell', parent=styles['Normal'], fontSize=6.5, leading=7.5)

    elements.append(Paragraph(f"Rapport d'Inventaire - {inventaire.titre}", title_style))
    statut_label = "Validé" if inventaire.statut == 'valide' else "Annulé" if inventaire.statut == 'annule' else "En cours"
    validated_by_str = f" | Validé par : {inventaire.validated_by.nom} {inventaire.validated_by.prenom} le {inventaire.validated_at.strftime('%d/%m/%Y %H:%M')}" if inventaire.validated_at else ""

    elements.append(Paragraph(f"Statut : {statut_label} | Créé le : {inventaire.created_at.strftime('%d/%m/%Y %H:%M')}{validated_by_str}", normal_style))

    sort_labels = {'code': 'Code lot', 'personne': 'Personne (saisi par)', 'produit': 'Produit', 'ecart': 'Écarts en premier'}
    tri_str = f"Trié par : {sort_labels.get(sort_by, 'Code lot')}"
    if personne_filtree:
        tri_str += f" | Filtré sur : {personne_filtree.prenom} {personne_filtree.nom}"
    elements.append(Paragraph(tri_str, normal_style))
    elements.append(Spacer(1, 10))

    data = [['Code Lot', 'Produit', 'Stock Théor.', 'Stock Constaté', 'Écart', 'Saisi par']]
    for l in lignes:
        cond = l.produit.conditionnement
        
        if cond == 3:
            th_str = f"U:{l.quantite_unites_avant}\nSU:{l.quantite_sous_unites_avant}\nSSU:{l.quantite_sous_sous_unites_avant}"
        elif cond == 2:
            th_str = f"U:{l.quantite_unites_avant}\nSU:{l.quantite_sous_unites_avant}"
        else:
            th_str = f"U:{l.quantite_unites_avant}"
            
        u_ap = l.quantite_unites_apres if l.quantite_unites_apres is not None else l.quantite_unites_avant
        su_ap = l.quantite_sous_unites_apres if l.quantite_sous_unites_apres is not None else l.quantite_sous_unites_avant
        ssu_ap = l.quantite_sous_sous_unites_apres if l.quantite_sous_sous_unites_apres is not None else l.quantite_sous_sous_unites_avant
        
        if cond == 3:
            ap_str = f"U:{u_ap}\nSU:{su_ap}\nSSU:{ssu_ap}"
        elif cond == 2:
            ap_str = f"U:{u_ap}\nSU:{su_ap}"
        else:
            ap_str = f"U:{u_ap}"
            
        diff_u = u_ap - l.quantite_unites_avant
        diff_su = su_ap - l.quantite_sous_unites_avant
        diff_ssu = ssu_ap - l.quantite_sous_sous_unites_avant
        
        diff_parts = []
        if diff_u != 0:
            diff_parts.append(f"U:{'+' if diff_u > 0 else ''}{diff_u}")
        if cond >= 2 and diff_su != 0:
            diff_parts.append(f"SU:{'+' if diff_su > 0 else ''}{diff_su}")
        if cond == 3 and diff_ssu != 0:
            diff_parts.append(f"SSU:{'+' if diff_ssu > 0 else ''}{diff_ssu}")
            
        diff_str = "\n".join(diff_parts) if diff_parts else "Aucun"

        if l.constate_by:
            saisi_par_str = f"{l.constate_by.prenom} {l.constate_by.nom}\n{l.constate_at.strftime('%d/%m %H:%M')}"
        else:
            saisi_par_str = "Non saisi"

        data.append([
            Paragraph(l.code_suivi, code_style),
            Paragraph(l.produit.nom, cell_style),
            th_str,
            ap_str,
            diff_str,
            saisi_par_str
        ])

    table = Table(data, repeatRows=1, colWidths=[105, 135, 65, 65, 55, 110])
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1abc9c')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 7.5),
        ('FONTSIZE', (0, 1), (-1, -1), 6.5),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('LEFTPADDING', (0, 0), (-1, -1), 3),
        ('RIGHTPADDING', (0, 0), (-1, -1), 3),
        ('TOPPADDING', (0, 0), (-1, -1), 2.5),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 2.5),
    ]))
    
    elements.append(table)
    doc.build(elements)
    output.seek(0)
    return send_file(output, download_name=f"rapport_inv_{id}.pdf", as_attachment=True)




# --- MODULE IMPOTS (DECLARATION DES TAXES) ---

def generate_reference_declaration():
    while True:
        suffix = ''.join(secrets.choice('0123456789') for _ in range(6))
        reference = f'IMP-{suffix}'
        if not DeclarationImpot.query.filter_by(reference=reference).first():
            return reference

def get_ventes_declaration(declaration):
    """Ventes validees comprises dans la periode de la declaration (bornes incluses)."""
    debut_dt = datetime(declaration.periode_debut.year, declaration.periode_debut.month, declaration.periode_debut.day)
    fin_dt = datetime(declaration.periode_fin.year, declaration.periode_fin.month, declaration.periode_fin.day, 23, 59, 59)
    return Vente.query.filter(
        Vente.statut == 'validee',
        Vente.created_at >= debut_dt,
        Vente.created_at <= fin_dt
    ).order_by(Vente.created_at.asc()).all()

def compute_impots_summary(ventes):
    totals_reels = compute_ventes_totals_reels(ventes)
    return {
        'count': len(ventes),
        'ht': sum(money_value(v.total_ht) for v in ventes),
        'tva': totals_reels['tva_reelle'],
        'benefice': totals_reels['benefice'],
        'ttc': sum(money_value(v.total_ttc) for v in ventes)
    }

def declaration_totaux_affiches(declaration):
    """Totaux a afficher : geles si la periode est declaree, recalcules en direct sinon."""
    if declaration.est_declaree:
        return {
            'count': declaration.nb_ventes,
            'ht': declaration.total_ht,
            'tva': declaration.total_tva,
            'benefice': declaration.total_benefice,
            'ttc': declaration.total_ttc
        }
    return compute_impots_summary(get_ventes_declaration(declaration))

@admin.route('/impots')
@login_required
@permission_required('module_impots')
def list_declarations_impots():
    declarations = DeclarationImpot.query.order_by(DeclarationImpot.periode_debut.desc()).all()

    query = (request.args.get('q') or '').strip().lower()
    statut = (request.args.get('statut') or '').strip()
    if query:
        declarations = [
            d for d in declarations
            if query in ' '.join([
                d.reference or '',
                d.note or '',
                d.periode_label,
                (d.created_by.prenom + ' ' + d.created_by.nom) if d.created_by else '',
                (d.declared_by.prenom + ' ' + d.declared_by.nom) if d.declared_by else ''
            ]).lower()
        ]
    if statut:
        declarations = [d for d in declarations if d.statut == statut]

    rows = [{'declaration': d, 'totaux': declaration_totaux_affiches(d)} for d in declarations]

    # Conflit de chevauchement renvoye par create_declaration_impot (affiche en modale)
    conflit_ref = (request.args.get('chevauchement') or '').strip()
    conflit = DeclarationImpot.query.filter_by(reference=conflit_ref).first() if conflit_ref else None
    tentative_debut = parse_date_filter((request.args.get('periode_debut') or '').strip())
    tentative_fin = parse_date_filter((request.args.get('periode_fin') or '').strip())

    toutes = DeclarationImpot.query.all()
    declarees = [d for d in toutes if d.est_declaree]
    stats_globales = {
        'total': len(toutes),
        'declarees': len(declarees),
        'en_preparation': len(toutes) - len(declarees),
        'tva_declaree': sum(money_value(d.total_tva) for d in declarees),
        'ttc_declare': sum(money_value(d.total_ttc) for d in declarees)
    }

    return render_template(
        'admin/impots/list.html',
        rows=rows,
        stats_globales=stats_globales,
        conflit=conflit,
        tentative_debut=tentative_debut,
        tentative_fin=tentative_fin
    )

@admin.route('/impots/create', methods=['POST'])
@login_required
@permission_required('module_impots')
def create_declaration_impot():
    date_debut = parse_date_filter((request.form.get('periode_debut') or '').strip())
    date_fin = parse_date_filter((request.form.get('periode_fin') or '').strip())
    note = (request.form.get('note') or '').strip() or None

    if not date_debut or not date_fin:
        flash("Veuillez renseigner les deux dates de la période à déclarer.", "danger")
        return redirect(url_for('admin.list_declarations_impots'))
    if date_debut > date_fin:
        flash("La date de début doit être antérieure ou égale à la date de fin.", "danger")
        return redirect(url_for('admin.list_declarations_impots'))

    chevauchement = DeclarationImpot.query.filter(
        DeclarationImpot.periode_debut <= date_fin,
        DeclarationImpot.periode_fin >= date_debut
    ).first()
    if chevauchement:
        # Affiche sur la page liste une modale de conflit bien visible, avec la saisie
        # conservee dans le formulaire pour correction.
        params = {
            'chevauchement': chevauchement.reference,
            'periode_debut': date_debut.isoformat(),
            'periode_fin': date_fin.isoformat(),
        }
        if note:
            params['note'] = note
        return redirect(url_for('admin.list_declarations_impots', **params))

    declaration = DeclarationImpot(
        reference=generate_reference_declaration(),
        periode_debut=date_debut,
        periode_fin=date_fin,
        note=note,
        created_by_id=current_user.id
    )
    db.session.add(declaration)
    db.session.commit()
    flash(f"Période de déclaration {declaration.reference} créée. Vérifiez le récapitulatif puis exportez le PDF.", "success")
    return redirect(url_for('admin.show_declaration_impot', id=declaration.id))

@admin.route('/impots/<int:id>')
@login_required
@permission_required('module_impots')
def show_declaration_impot(id):
    declaration = DeclarationImpot.query.get_or_404(id)
    ventes = get_ventes_declaration(declaration)
    summary = compute_impots_summary(ventes)
    tva_breakdown = compute_tva_breakdown(ventes)
    return render_template(
        'admin/impots/show.html',
        declaration=declaration,
        ventes=ventes,
        summary=summary,
        tva_breakdown=tva_breakdown
    )

@admin.route('/impots/<int:id>/declarer', methods=['POST'])
@login_required
@permission_required('module_impots')
def declarer_impot(id):
    declaration = DeclarationImpot.query.get_or_404(id)
    if declaration.est_declaree:
        flash("Cette période est déjà marquée comme déclarée.", "warning")
        return redirect(url_for('admin.show_declaration_impot', id=declaration.id))

    summary = compute_impots_summary(get_ventes_declaration(declaration))
    declaration.statut = 'declaree'
    declaration.nb_ventes = summary['count']
    declaration.total_ht = summary['ht']
    declaration.total_tva = summary['tva']
    declaration.total_benefice = summary['benefice']
    declaration.total_ttc = summary['ttc']
    declaration.declared_at = datetime.utcnow()
    declaration.declared_by_id = current_user.id
    db.session.commit()
    flash(f"La période {declaration.periode_label} est marquée comme déclarée. Les totaux sont gelés.", "success")
    return redirect(url_for('admin.show_declaration_impot', id=declaration.id))

@admin.route('/impots/<int:id>/rouvrir', methods=['POST'])
@login_required
@permission_required('module_impots')
def rouvrir_declaration_impot(id):
    declaration = DeclarationImpot.query.get_or_404(id)
    if not declaration.est_declaree:
        flash("Cette période n'est pas encore déclarée.", "warning")
        return redirect(url_for('admin.show_declaration_impot', id=declaration.id))
    declaration.statut = 'en_preparation'
    declaration.declared_at = None
    declaration.declared_by_id = None
    db.session.commit()
    flash(f"La déclaration {declaration.reference} a été rouverte : elle n'est plus marquée comme déclarée.", "info")
    return redirect(url_for('admin.show_declaration_impot', id=declaration.id))

@admin.route('/impots/<int:id>/delete', methods=['POST'])
@login_required
@permission_required('module_impots')
def delete_declaration_impot(id):
    declaration = DeclarationImpot.query.get_or_404(id)
    if declaration.est_declaree:
        flash("Impossible de supprimer une période déjà déclarée. Rouvrez-la d'abord.", "danger")
        return redirect(url_for('admin.show_declaration_impot', id=declaration.id))
    reference = declaration.reference
    db.session.delete(declaration)
    db.session.commit()
    flash(f"La déclaration {reference} a été supprimée.", "success")
    return redirect(url_for('admin.list_declarations_impots'))

@admin.route('/impots/<int:id>/export/pdf')
@login_required
@permission_required('module_impots')
def export_declaration_impot_pdf(id):
    import io
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
    from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

    declaration = DeclarationImpot.query.get_or_404(id)
    ventes = get_ventes_declaration(declaration)
    summary = compute_impots_summary(ventes)
    tva_breakdown = compute_tva_breakdown(ventes)
    pharmacy_name = Setting.get_value('pharmacy_name', 'REFLEXPHARMA')
    generated_at = datetime.now()

    output = io.BytesIO()
    doc = SimpleDocTemplate(output, pagesize=landscape(A4), topMargin=20, bottomMargin=20, leftMargin=20, rightMargin=20)
    styles = getSampleStyleSheet()
    styles.add(ParagraphStyle('Small', parent=styles['Normal'], fontSize=7.5, leading=10))
    styles.add(ParagraphStyle('Cell', parent=styles['Normal'], fontSize=6.5, leading=8))

    statut_label = 'DÉCLARÉE' if declaration.est_declaree else 'EN PRÉPARATION'
    elements = [
        Paragraph(f'Déclaration des taxes - {pharmacy_name}', styles['Title']),
        Paragraph(
            f'Référence : {declaration.reference} | Période : {declaration.periode_label} | Statut : {statut_label}',
            styles['Small']
        ),
        Paragraph(
            f'Date du tirage : {generated_at.strftime("%d/%m/%Y %H:%M")} | Tiré par : {current_user.nom} {current_user.prenom} | Ventes incluses : {summary["count"]}',
            styles['Small']
        )
    ]
    if declaration.est_declaree and declaration.declared_at:
        declarant = f'{declaration.declared_by.prenom} {declaration.declared_by.nom}' if declaration.declared_by else '-'
        elements.append(Paragraph(
            f'Période marquée comme déclarée le {declaration.declared_at.strftime("%d/%m/%Y %H:%M")} par {declarant}.',
            styles['Small']
        ))
    if declaration.note:
        elements.append(Paragraph(f'Note : {declaration.note}', styles['Small']))
    elements.append(Spacer(1, 10))

    elements.append(Paragraph('Récapitulatif de la période', styles['Heading3']))
    recap = [
        ['Nombre de ventes', 'Total HT (€)', 'TVA effective (€)', 'Bénéfice (€)', 'Total TTC (€)'],
        [
            summary['count'],
            f"{summary['ht']:.2f}",
            f"{summary['tva']:.2f}",
            f"{summary['benefice']:.2f}",
            f"{summary['ttc']:.2f}"
        ]
    ]
    elements.append(Table(recap, colWidths=[150, 150, 150, 150, 150], style=[
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#a5670a')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('GRID', (0, 0), (-1, -1), 0.25, colors.HexColor('#D1D5DB')),
        ('FONTSIZE', (0, 0), (-1, -1), 8),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
    ]))
    elements.append(Spacer(1, 10))

    elements.append(Paragraph('Répartition par taux de TVA', styles['Heading3']))
    tva_data = [['Taux de TVA (%)', 'Base HT (€)', 'Montant TVA (€)']]
    for row in tva_breakdown:
        tva_data.append([f"{row['taux']:.2f}", f"{row['ht']:.2f}", f"{row['tva']:.2f}"])
    if not tva_breakdown:
        tva_data.append(['-', '0.00', '0.00'])
    elements.append(Table(tva_data, colWidths=[150, 200, 200], repeatRows=1, style=[
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#374151')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('GRID', (0, 0), (-1, -1), 0.25, colors.HexColor('#D1D5DB')),
        ('FONTSIZE', (0, 0), (-1, -1), 8),
        ('ALIGN', (1, 0), (-1, -1), 'RIGHT'),
    ]))
    elements.append(Spacer(1, 12))

    elements.append(Paragraph('Détail de toutes les ventes de la période', styles['Heading3']))
    data = [['N° vente', 'Date', 'Client', 'Employé', 'Mode paiement', 'HT (€)', 'TVA (€)', 'Bénéfice (€)', 'TTC (€)']]
    for vente in ventes:
        data.append([
            Paragraph(vente.numero_vente or '-', styles['Cell']),
            vente.created_at.strftime('%d/%m/%Y %H:%M') if vente.created_at else '-',
            Paragraph(vente.client_label, styles['Cell']),
            Paragraph(sale_employee_label(vente), styles['Cell']),
            (vente.mode_paiement or '-').replace('_', ' '),
            f"{money_value(vente.total_ht):.2f}",
            f"{vente.total_tva_reelle:.2f}",
            f"{vente.total_benefice:.2f}",
            f"{money_value(vente.total_ttc):.2f}"
        ])
    if not ventes:
        data.append(['Aucune vente sur la période', '', '', '', '', '', '', '', ''])
    data.append([
        'TOTAL', '', '', '', '',
        f"{summary['ht']:.2f}",
        f"{summary['tva']:.2f}",
        f"{summary['benefice']:.2f}",
        f"{summary['ttc']:.2f}"
    ])
    table = Table(data, repeatRows=1, colWidths=[95, 75, 120, 110, 75, 70, 70, 70, 70])
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2c3e50')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('GRID', (0, 0), (-1, -1), 0.25, colors.HexColor('#D1D5DB')),
        ('FONTSIZE', (0, 0), (-1, -1), 6.5),
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ('ALIGN', (5, 0), (-1, -1), 'RIGHT'),
        ('ROWBACKGROUNDS', (0, 1), (-1, -2), [colors.white, colors.HexColor('#F9FAFB')]),
        ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#fef7e0')),
        ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
    ]))
    elements.append(table)

    doc.build(elements)
    output.seek(0)
    filename = f'declaration_impots_{declaration.reference}_{declaration.periode_debut.strftime("%Y%m%d")}_{declaration.periode_fin.strftime("%Y%m%d")}.pdf'
    return send_file(output, download_name=filename, as_attachment=True)

@admin.route('/impots/<int:id>/export/excel')
@login_required
@permission_required('module_impots')
def export_declaration_impot_excel(id):
    import io
    import pandas as pd
    from openpyxl.styles import Alignment, Font, PatternFill, Border, Side

    declaration = DeclarationImpot.query.get_or_404(id)
    ventes = get_ventes_declaration(declaration)
    summary = compute_impots_summary(ventes)
    tva_breakdown = compute_tva_breakdown(ventes)
    pharmacy_name = Setting.get_value('pharmacy_name', 'REFLEXPHARMA')
    generated_at = datetime.now()
    output = io.BytesIO()

    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        synth_rows = [
            {'Indicateur': 'Nombre de ventes', 'Valeur': summary['count']},
            {'Indicateur': 'Total HT (€)', 'Valeur': round(summary['ht'], 2)},
            {'Indicateur': 'TVA effective (€)', 'Valeur': round(summary['tva'], 2)},
            {'Indicateur': 'Bénéfice (€)', 'Valeur': round(summary['benefice'], 2)},
            {'Indicateur': 'Total TTC (€)', 'Valeur': round(summary['ttc'], 2)},
        ]
        pd.DataFrame(synth_rows).to_excel(writer, index=False, sheet_name='Synthese', startrow=7)

        tva_rows = [
            {'Taux de TVA (%)': row['taux'], 'Base HT (€)': round(row['ht'], 2), 'Montant TVA (€)': round(row['tva'], 2)}
            for row in tva_breakdown
        ] or [{'Taux de TVA (%)': '-', 'Base HT (€)': 0.0, 'Montant TVA (€)': 0.0}]
        pd.DataFrame(tva_rows).to_excel(writer, index=False, sheet_name='TVA par taux')

        vente_rows = [
            {
                'N° vente': vente.numero_vente or '-',
                'Date': vente.created_at.strftime('%d/%m/%Y %H:%M') if vente.created_at else '-',
                'Client': vente.client_label,
                'Employé': sale_employee_label(vente),
                'Mode paiement': (vente.mode_paiement or '-').replace('_', ' '),
                'HT (€)': round(money_value(vente.total_ht), 2),
                'TVA (€)': round(vente.total_tva_reelle, 2),
                'Bénéfice (€)': round(vente.total_benefice, 2),
                'TTC (€)': round(money_value(vente.total_ttc), 2),
            }
            for vente in ventes
        ]
        vente_rows.append({
            'N° vente': 'TOTAL',
            'Date': '',
            'Client': '',
            'Employé': '',
            'Mode paiement': '',
            'HT (€)': round(summary['ht'], 2),
            'TVA (€)': round(summary['tva'], 2),
            'Bénéfice (€)': round(summary['benefice'], 2),
            'TTC (€)': round(summary['ttc'], 2),
        })
        pd.DataFrame(vente_rows).to_excel(writer, index=False, sheet_name='Ventes')

        for sheet_name, worksheet in writer.sheets.items():
            worksheet.freeze_panes = 'A9' if sheet_name == 'Synthese' else 'A2'
            worksheet.column_dimensions['A'].width = 24
            for column in ['B', 'C', 'D', 'E', 'F', 'G', 'H', 'I']:
                worksheet.column_dimensions[column].width = 18
            header_row = 8 if sheet_name == 'Synthese' else 1
            for cell in worksheet[header_row]:
                cell.font = Font(bold=True, color='FFFFFF')
                cell.fill = PatternFill(start_color='A5670A', end_color='A5670A', fill_type='solid')
                cell.alignment = Alignment(horizontal='center')
            for row in worksheet.iter_rows(min_row=header_row + 1):
                for cell in row:
                    cell.border = Border(bottom=Side(style='thin', color='D1D5DB'))

        ventes_ws = writer.sheets['Ventes']
        for cell in ventes_ws[ventes_ws.max_row]:
            cell.font = Font(bold=True, color='7A4D00')
            cell.fill = PatternFill(start_color='FEF7E0', end_color='FEF7E0', fill_type='solid')

        statut_label = 'DÉCLARÉE' if declaration.est_declaree else 'EN PRÉPARATION'
        synth = writer.sheets['Synthese']
        synth['A1'] = f'Déclaration des taxes - {pharmacy_name}'
        synth['A1'].font = Font(bold=True, size=13, color='7A4D00')
        synth['A2'] = f'Référence : {declaration.reference} | Période : {declaration.periode_label} | Statut : {statut_label}'
        synth['A3'] = f'Date du tirage : {generated_at.strftime("%d/%m/%Y %H:%M")} | Tiré par : {current_user.nom} {current_user.prenom}'
        if declaration.est_declaree and declaration.declared_at:
            declarant = f'{declaration.declared_by.prenom} {declaration.declared_by.nom}' if declaration.declared_by else '-'
            synth['A4'] = f'Période marquée comme déclarée le {declaration.declared_at.strftime("%d/%m/%Y %H:%M")} par {declarant}'
        if declaration.note:
            synth['A5'] = f'Note : {declaration.note}'

    output.seek(0)
    filename = f'declaration_impots_{declaration.reference}_{declaration.periode_debut.strftime("%Y%m%d")}_{declaration.periode_fin.strftime("%Y%m%d")}.xlsx'
    return send_file(output, download_name=filename, as_attachment=True)
