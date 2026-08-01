"""Logique et exports du module Caisse (comptage physique de fin de journée +
contrôle par rapport aux ventes en espèces), séparés de blueprints/admin/views.py
sur le même principe que finance_reports.py.
"""

from datetime import datetime, timedelta

from models.vente import Vente
from models.declaration_caisse import DeclarationCaisse
from utils.currencies import devise_active

ECART_TOLERANCE = 0.01


def money_value(value):
    return float(value or 0)


def get_ventes_especes_jour(date_jour):
    """Ventes validées d'un jour donné (bornes incluses). Pas de filtre sur
    mode_paiement : ce champ n'est qu'une étiquette informative (especes,
    solde_client, mixte...) qui ne change rien au calcul du paiement -- voir
    create_vente() dans views.py, où la répartition groupe/solde client/espèces
    est indépendante de la valeur choisie. Le seul indicateur fiable de ce qui
    est réellement passé en espèces est montant_hors_solde (voir total_especes_jour)."""
    debut_dt = datetime(date_jour.year, date_jour.month, date_jour.day)
    fin_dt = datetime(date_jour.year, date_jour.month, date_jour.day, 23, 59, 59)
    return Vente.query.filter(
        Vente.statut == 'validee',
        Vente.created_at >= debut_dt,
        Vente.created_at <= fin_dt
    ).order_by(Vente.created_at.asc()).all()


def total_especes_jour(ventes):
    """Total réellement encaissé en espèces : montant_hors_solde (part de chaque
    vente non couverte par un solde client/groupe), pas total_ttc -- une vente
    partiellement payée par solde ne fait pas rentrer ce montant-là en caisse."""
    return sum(money_value(v.montant_hors_solde) for v in ventes)


def compute_controle_row(date_jour, declaration):
    ventes = get_ventes_especes_jour(date_jour)
    total_especes = total_especes_jour(ventes)
    montant_declare = money_value(declaration.montant_declare) if declaration else None
    ecart = (montant_declare - total_especes) if declaration else None

    if declaration is None:
        if ventes:
            statut = 'non_declare'
        else:
            statut = 'aucune_vente'
    elif abs(ecart) <= ECART_TOLERANCE:
        statut = 'ok'
    else:
        statut = 'ecart'

    return {
        'date_jour': date_jour,
        'declaration': declaration,
        'montant_declare': montant_declare,
        'nb_ventes': len(ventes),
        'total_especes': total_especes,
        'ecart': ecart,
        'statut': statut,
    }


def build_controle_rows(date_from, date_to):
    """Une ligne par jour de la période, déclaration ou non, dès qu'il y a une
    déclaration ou au moins une vente en espèces ce jour-là (les jours sans
    aucune activité n'apportent rien à un rapport de contrôle)."""
    declarations = {
        d.date_jour: d
        for d in DeclarationCaisse.query.filter(
            DeclarationCaisse.date_jour >= date_from,
            DeclarationCaisse.date_jour <= date_to
        ).all()
    }

    rows = []
    jour = date_from
    while jour <= date_to:
        declaration = declarations.get(jour)
        row = compute_controle_row(jour, declaration)
        if row['statut'] != 'aucune_vente' or declaration is not None:
            rows.append(row)
        jour += timedelta(days=1)
    return rows


STATUT_LABELS = {
    'ok': 'Conforme',
    'ecart': 'Écart',
    'non_declare': 'Non déclaré',
    'aucune_vente': 'Aucune vente',
}


def build_declarations_caisse_pdf(target, declarations, tire_par, pharmacy_name):
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
    from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

    devise = devise_active()
    doc = SimpleDocTemplate(target, pagesize=A4, topMargin=24, bottomMargin=24, leftMargin=24, rightMargin=24)
    styles = getSampleStyleSheet()
    styles.add(ParagraphStyle('Small', parent=styles['Normal'], fontSize=8, leading=10))
    largeur_utile = A4[0] - doc.leftMargin - doc.rightMargin

    total = sum(money_value(d.montant_declare) for d in declarations)
    elements = [
        Paragraph(f'Déclarations de caisse - {pharmacy_name}', styles['Title']),
        Paragraph(
            f'Date du tirage : {datetime.now().strftime("%d/%m/%Y %H:%M")} | Tiré par : {tire_par} | '
            f'Nombre de déclarations : {len(declarations)} | Total déclaré : {total:.2f} {devise}',
            styles['Small']
        ),
        Spacer(1, 10)
    ]

    cell_style = ParagraphStyle('CelluleCaisse', parent=styles['Normal'], fontSize=8, leading=10)
    data = [['Date', f'Montant déclaré ({devise})', 'Note', 'Déclaré par']]
    for d in declarations:
        data.append([
            d.date_jour.strftime('%d/%m/%Y'),
            f'{money_value(d.montant_declare):.2f}',
            Paragraph(d.note or '', cell_style),
            f'{d.created_by.prenom} {d.created_by.nom}' if d.created_by else '-',
        ])
    if not declarations:
        data.append(['Aucune déclaration', '', '', ''])

    elements.append(Table(data, colWidths=[largeur_utile * p for p in (0.15, 0.2, 0.45, 0.2)], repeatRows=1, style=TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2c3e50')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('GRID', (0, 0), (-1, -1), 0.25, colors.HexColor('#D1D5DB')),
        ('FONTSIZE', (0, 0), (-1, -1), 8),
        ('ALIGN', (1, 0), (1, -1), 'RIGHT'),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F9FAFB')]),
    ])))
    doc.build(elements)


def build_declarations_caisse_excel(target, declarations):
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    devise = devise_active()
    wb = Workbook()
    ws = wb.active
    ws.title = 'Déclarations caisse'

    total = sum(money_value(d.montant_declare) for d in declarations)
    ws['A1'] = 'Déclarations de caisse'
    ws['A1'].font = Font(bold=True, size=14, color='2C3E50')
    ws['A2'] = f'Généré le {datetime.now().strftime("%d/%m/%Y %H:%M")}'
    ws['A3'] = f'Nombre de déclarations : {len(declarations)} | Total déclaré : {total:.2f} {devise}'

    header_row = 5
    columns = ['Date', f'Montant déclaré ({devise})', 'Note', 'Déclaré par']
    header_fill = PatternFill(start_color='2C3E50', end_color='2C3E50', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF')
    for ci, col in enumerate(columns, start=1):
        cell = ws.cell(row=header_row, column=ci, value=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')

    row = header_row + 1
    for d in declarations:
        ws.cell(row=row, column=1, value=d.date_jour.strftime('%d/%m/%Y'))
        ws.cell(row=row, column=2, value=round(money_value(d.montant_declare), 2))
        ws.cell(row=row, column=3, value=d.note or '')
        ws.cell(row=row, column=4, value=f'{d.created_by.prenom} {d.created_by.nom}' if d.created_by else '-')
        row += 1

    ws.freeze_panes = f'A{header_row + 1}'
    for col_idx in range(1, len(columns) + 1):
        longueur = max(
            (len(str(ws.cell(row=r, column=col_idx).value or '')) for r in range(1, ws.max_row + 1)),
            default=0
        )
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max(12, longueur + 2), 45)

    wb.save(target)


def build_controle_caisse_pdf(target, rows, periode_label, tire_par, pharmacy_name):
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
    from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

    devise = devise_active()
    doc = SimpleDocTemplate(target, pagesize=A4, topMargin=24, bottomMargin=24, leftMargin=24, rightMargin=24)
    styles = getSampleStyleSheet()
    styles.add(ParagraphStyle('Small', parent=styles['Normal'], fontSize=8, leading=10))
    largeur_utile = A4[0] - doc.leftMargin - doc.rightMargin

    nb_ecarts = sum(1 for r in rows if r['statut'] == 'ecart')
    nb_non_declares = sum(1 for r in rows if r['statut'] == 'non_declare')
    elements = [
        Paragraph(f'Contrôle de caisse - {pharmacy_name}', styles['Title']),
        Paragraph(
            f'Période : {periode_label} | Date du tirage : {datetime.now().strftime("%d/%m/%Y %H:%M")} | '
            f'Tiré par : {tire_par} | Écarts : {nb_ecarts} | Jours non déclarés : {nb_non_declares}',
            styles['Small']
        ),
        Spacer(1, 10)
    ]

    data = [['Date', f'Déclaré ({devise})', f'Ventes espèces ({devise})', f'Écart ({devise})', 'Statut']]
    for r in rows:
        data.append([
            r['date_jour'].strftime('%d/%m/%Y'),
            f"{r['montant_declare']:.2f}" if r['montant_declare'] is not None else '-',
            f"{r['total_especes']:.2f}",
            f"{r['ecart']:+.2f}" if r['ecart'] is not None else '-',
            STATUT_LABELS[r['statut']],
        ])
    if not rows:
        data.append(['Aucune donnée sur cette période', '', '', '', ''])

    table = Table(data, colWidths=[largeur_utile * p for p in (0.2, 0.2, 0.24, 0.16, 0.2)], repeatRows=1)
    style = [
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2c3e50')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('GRID', (0, 0), (-1, -1), 0.25, colors.HexColor('#D1D5DB')),
        ('FONTSIZE', (0, 0), (-1, -1), 8),
        ('ALIGN', (1, 0), (3, -1), 'RIGHT'),
        ('ALIGN', (4, 0), (4, -1), 'CENTER'),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F9FAFB')]),
    ]
    for i, r in enumerate(rows, start=1):
        if r['statut'] == 'ecart':
            style.append(('BACKGROUND', (0, i), (-1, i), colors.HexColor('#fdecea')))
        elif r['statut'] == 'non_declare':
            style.append(('BACKGROUND', (0, i), (-1, i), colors.HexColor('#fef7e0')))
    table.setStyle(TableStyle(style))
    elements.append(table)
    doc.build(elements)


def build_controle_caisse_excel(target, rows, periode_label):
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    devise = devise_active()
    wb = Workbook()
    ws = wb.active
    ws.title = 'Contrôle caisse'

    nb_ecarts = sum(1 for r in rows if r['statut'] == 'ecart')
    nb_non_declares = sum(1 for r in rows if r['statut'] == 'non_declare')
    ws['A1'] = 'Contrôle de caisse'
    ws['A1'].font = Font(bold=True, size=14, color='2C3E50')
    ws['A2'] = f'Période : {periode_label}'
    ws['A3'] = f'Généré le {datetime.now().strftime("%d/%m/%Y %H:%M")}'
    ws['A4'] = f'Écarts : {nb_ecarts} | Jours non déclarés : {nb_non_declares}'

    header_row = 6
    columns = ['Date', f'Déclaré ({devise})', f'Ventes espèces ({devise})', f'Écart ({devise})', 'Statut']
    header_fill = PatternFill(start_color='2C3E50', end_color='2C3E50', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF')
    for ci, col in enumerate(columns, start=1):
        cell = ws.cell(row=header_row, column=ci, value=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')

    ecart_fill = PatternFill(start_color='FDECEA', end_color='FDECEA', fill_type='solid')
    non_declare_fill = PatternFill(start_color='FEF7E0', end_color='FEF7E0', fill_type='solid')

    row = header_row + 1
    for r in rows:
        ws.cell(row=row, column=1, value=r['date_jour'].strftime('%d/%m/%Y'))
        ws.cell(row=row, column=2, value=round(r['montant_declare'], 2) if r['montant_declare'] is not None else None)
        ws.cell(row=row, column=3, value=round(r['total_especes'], 2))
        ws.cell(row=row, column=4, value=round(r['ecart'], 2) if r['ecart'] is not None else None)
        ws.cell(row=row, column=5, value=STATUT_LABELS[r['statut']])
        if r['statut'] == 'ecart':
            for ci in range(1, len(columns) + 1):
                ws.cell(row=row, column=ci).fill = ecart_fill
        elif r['statut'] == 'non_declare':
            for ci in range(1, len(columns) + 1):
                ws.cell(row=row, column=ci).fill = non_declare_fill
        row += 1

    ws.freeze_panes = f'A{header_row + 1}'
    for col_idx in range(1, len(columns) + 1):
        longueur = max(
            (len(str(ws.cell(row=r, column=col_idx).value or '')) for r in range(1, ws.max_row + 1)),
            default=0
        )
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max(12, longueur + 2), 45)

    wb.save(target)
